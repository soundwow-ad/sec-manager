import io
import os
import calendar

import pandas as pd


def build_annual_seconds_summary(
    df_daily,
    year,
    monthly_capacity_loader,
    annual_summary_entity_labels,
    annual_summary_media_map,
    seconds_usage_types,
):
    if df_daily.empty or "日期" not in df_daily.columns or "使用店秒" not in df_daily.columns:
        return None
    df = df_daily.copy()
    df["日期"] = pd.to_datetime(df["日期"], errors="coerce")
    df = df.dropna(subset=["日期"])
    df["年"] = df["日期"].dt.year
    df["月"] = df["日期"].dt.month
    df = df[df["年"] == int(year)]
    if "媒體平台" not in df.columns:
        return None
    if "秒數用途" not in df.columns:
        # 若根本沒有該欄位，避免誤把未知資料當成銷售秒數
        df["秒數用途"] = ""

    def to_entity(mp):
        for ent, platforms in annual_summary_media_map.items():
            if mp in platforms:
                return ent
        return None

    df["實體"] = df["媒體平台"].map(to_entity)
    df = df[df["實體"].notna()]

    months = list(range(1, 13))
    month_cols = [f"{m}月" for m in months]
    top_rows = []
    monthly_cap_cache = {}
    if monthly_capacity_loader:
        for ent in annual_summary_entity_labels:
            row = {"項目": f"{ent}使用率"}
            for m in months:
                cap = 0
                for mp in annual_summary_media_map.get(ent, []):
                    daily = monthly_capacity_loader(mp, year, m)
                    if daily is not None and daily > 0:
                        ndays = calendar.monthrange(int(year), m)[1]
                        cap += int(daily) * ndays
                monthly_cap_cache[(ent, m)] = cap
                used = df[(df["實體"] == ent) & (df["月"] == m)]["使用店秒"].sum()
                pct = (used / cap * 100) if cap else 0
                row[f"{m}月"] = round(pct, 1)
            top_rows.append(row)
    top_usage_df = pd.DataFrame(top_rows, columns=["項目"] + month_cols) if top_rows else None

    entities_out = {}
    for ent in annual_summary_entity_labels:
        platforms = annual_summary_media_map.get(ent, [])
        df_ent = df[df["實體"] == ent]
        avg_monthly = 0
        if monthly_capacity_loader:
            for mp in platforms:
                daily = monthly_capacity_loader(mp, year, 1)
                if daily is not None and daily > 0:
                    avg_monthly += int(daily) * calendar.monthrange(int(year), 1)[1]
        by_type_rows = []
        for stype in seconds_usage_types:
            row = {"項目": stype}
            for m in months:
                row[f"{m}月"] = int(df_ent[(df_ent["月"] == m) & (df_ent["秒數用途"] == stype)]["使用店秒"].sum())
            by_type_rows.append(row)
        by_type_df = pd.DataFrame(by_type_rows, columns=["項目"] + month_cols)

        used_row = {"項目": "使用秒數"}
        for m in months:
            used_row[f"{m}月"] = int(df_ent[df_ent["月"] == m]["使用店秒"].sum())
        unused_row = {"項目": "未使用秒數"}
        rate_row = {"項目": f"{ent}使用率"}
        for m in months:
            cap = monthly_cap_cache.get((ent, m))
            if cap is None and monthly_capacity_loader:
                cap = 0
                for mp in platforms:
                    daily = monthly_capacity_loader(mp, year, m)
                    if daily is not None and daily > 0:
                        ndays = calendar.monthrange(int(year), m)[1]
                        cap += int(daily) * ndays
                monthly_cap_cache[(ent, m)] = cap
            else:
                cap = cap or 0
            used = used_row.get(f"{m}月", 0) or 0
            unused_row[f"{m}月"] = max(0, int(cap) - int(used))
            rate_row[f"{m}月"] = round((used / cap * 100), 1) if cap else 0
        entities_out[ent] = {
            "avg_monthly_seconds": avg_monthly,
            "by_type_df": by_type_df,
            "used_row": used_row,
            "unused_row": unused_row,
            "usage_rate_row": rate_row,
        }
    return {"top_usage_df": top_usage_df, "entities": entities_out}


def build_visualization_summary_excel(annual_viz, summary_year, annual_summary_entity_labels):
    try:
        from openpyxl import Workbook
        from openpyxl.drawing.image import Image as ExcelImage
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        import altair as alt
        try:
            import vl_convert as vlc  # type: ignore[import-untyped]
        except ImportError:
            vlc = None
    except ImportError:
        return None

    buf = io.BytesIO()
    try:
        wb = Workbook()
        wb.remove(wb.active)
        month_cols = [f"{m}月" for m in range(1, 13)]

        def _chart_to_image(chart, scale=2):
            if vlc is None:
                return None
            try:
                return io.BytesIO(vlc.vegalite_to_png(chart.to_json(), scale=scale))
            except Exception:
                return None

        def _style_cell(cell, is_header=False, is_percentage=False, value=None):
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
            if is_header:
                cell.font = Font(bold=True, size=10)
                cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
            else:
                cell.font = Font(size=9)
                if is_percentage and value is not None:
                    try:
                        val_float = float(str(value).replace("%", "").replace(",", ""))
                        if val_float >= 100:
                            cell.fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
                            cell.font = Font(size=9, color="FFFFFF", bold=True)
                        elif val_float >= 70:
                            cell.fill = PatternFill(start_color="FFD93D", end_color="FFD93D", fill_type="solid")
                            cell.font = Font(size=9, bold=True)
                        elif val_float >= 50:
                            cell.fill = PatternFill(start_color="6BCF7F", end_color="6BCF7F", fill_type="solid")
                    except (ValueError, TypeError):
                        pass

        def _add_dataframe_to_sheet(ws, df, start_row=1, start_col=1, apply_color=False):
            for col_idx, col_name in enumerate(df.columns, start=start_col):
                cell = ws.cell(row=start_row, column=col_idx)
                cell.value = str(col_name)
                _style_cell(cell, is_header=True)
            for row_idx, (_, row) in enumerate(df.iterrows(), start=start_row + 1):
                for col_idx, col_name in enumerate(df.columns, start=start_col):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    val = row[col_name]
                    cell.value = val
                    is_percentage = "使用率" in str(col_name) or (isinstance(val, str) and "%" in val)
                    _style_cell(cell, is_header=False, is_percentage=is_percentage if apply_color else False, value=val)
            for col_idx, col_name in enumerate(df.columns, start=start_col):
                col_letter = get_column_letter(col_idx)
                max_length = max(len(str(col_name)), max([len(str(row[col_name])) for _, row in df.iterrows()], default=0))
                ws.column_dimensions[col_letter].width = min(max_length + 2, 15)

        ws1 = wb.create_sheet("①媒體平台使用率")
        ws1["A1"] = f"① 各媒體平台使用率隨時間變化趨勢 - {summary_year}"
        ws1["A1"].font = Font(bold=True, size=14)
        ws1.merge_cells("A1:N1")

        if annual_viz.get("top_usage_df") is not None and not annual_viz["top_usage_df"].empty:
            top_df = annual_viz["top_usage_df"].copy()
            top_df["媒體平台"] = top_df["項目"].str.replace("使用率", "", regex=False)
            chart_df_platform = top_df.set_index("媒體平台")[month_cols].T
            chart_df_platform.index.name = "月份"
            try:
                chart_df_platform_melted = chart_df_platform.reset_index().melt(id_vars="月份", var_name="媒體平台", value_name="使用率")
                chart_df_platform_melted["使用率標籤"] = chart_df_platform_melted["使用率"].apply(lambda x: f"{x:.1f}%" if pd.notna(x) else "")
                line_chart = alt.Chart(chart_df_platform_melted).mark_line(point=True).encode(
                    x=alt.X("月份:O", title="月份"),
                    y=alt.Y("使用率:Q", title="使用率 (%)", axis=alt.Axis(format=".1f")),
                    color=alt.Color("媒體平台:N", title="媒體平台"),
                    tooltip=["月份", "媒體平台", alt.Tooltip("使用率:Q", format=".1f", title="使用率 (%)")],
                ).properties(width=700, height=400)
                text_chart = alt.Chart(chart_df_platform_melted).mark_text(align="center", baseline="bottom", dy=-8, fontSize=10).encode(
                    x=alt.X("月份:O", title="月份"),
                    y=alt.Y("使用率:Q", title="使用率 (%)", axis=alt.Axis(format=".1f")),
                    text=alt.Text("使用率標籤:N"),
                    color=alt.Color("媒體平台:N", legend=None),
                )
                img_data = _chart_to_image((line_chart + text_chart).properties(width=700, height=400))
                if img_data:
                    img = ExcelImage(img_data)
                    img.width = 700
                    img.height = 400
                    ws1.add_image(img, "A3")
            except Exception:
                pass
            _add_dataframe_to_sheet(ws1, top_df, start_row=25, apply_color=True)

        ws2 = wb.create_sheet("②秒數類型比例")
        ws2["A1"] = f"② 各秒數類型使用比例隨時間變化趨勢 - {summary_year}"
        ws2["A1"].font = Font(bold=True, size=14)
        ws2.merge_cells("A1:N1")

        by_type_agg = None
        for ent in annual_summary_entity_labels:
            block = annual_viz.get("entities", {}).get(ent)
            if not block or block.get("by_type_df") is None:
                continue
            bt = block["by_type_df"].set_index("項目")[month_cols]
            by_type_agg = bt.copy() if by_type_agg is None else by_type_agg + bt
        if by_type_agg is not None and not by_type_agg.empty:
            monthly_total = by_type_agg.sum(axis=0)
            proportion = by_type_agg.copy()
            for c in month_cols:
                proportion[c] = (by_type_agg[c] / monthly_total[c] * 100) if (monthly_total.get(c, 0) and monthly_total[c] > 0) else 0
            for col in proportion.columns:
                monthly_sum = proportion[col].sum()
                if monthly_sum > 0 and abs(monthly_sum - 100) > 0.01:
                    proportion[col] = proportion[col] / monthly_sum * 100
            proportion_df = proportion.reset_index()
            proportion_df.columns = ["秒數類型"] + month_cols
            _add_dataframe_to_sheet(ws2, proportion_df, start_row=25, apply_color=False)

        for ent in annual_summary_entity_labels:
            block = annual_viz.get("entities", {}).get(ent)
            if not block:
                continue
            ws_ent = wb.create_sheet(f"{ent}")
            ws_ent["A1"] = f"{summary_year} {ent}"
            ws_ent["A1"].font = Font(bold=True, size=12)
            ws_ent.merge_cells("A1:N1")
            _bt = block.get("by_type_df")
            if _bt is not None and not _bt.empty:
                ws_ent["A3"] = f"{ent} 秒數用途分列（1月～12月）"
                ws_ent["A3"].font = Font(bold=True, size=11)
                _add_dataframe_to_sheet(ws_ent, _bt, start_row=4, apply_color=False)
            summary_table = pd.DataFrame([block.get("used_row", {}), block.get("unused_row", {}), block.get("usage_rate_row", {})])
            if not summary_table.empty:
                start_row = len(_bt) + 6 if _bt is not None and not _bt.empty else 4
                ws_ent.cell(row=start_row, column=1).value = f"{ent} 使用/未使用/使用率（1月～12月）"
                ws_ent.cell(row=start_row, column=1).font = Font(bold=True, size=11)
                _add_dataframe_to_sheet(ws_ent, summary_table, start_row=start_row + 1, apply_color=True)

        wb.save(buf)
        buf.seek(0)
        return buf.getvalue()
    except Exception:
        return None


def build_visualization_summary_pdf(annual_viz, summary_year, annual_summary_entity_labels):
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak, Image
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.ttfonts import TTFont
    except ImportError:
        return None
    buf = io.BytesIO()
    windir = os.environ.get("WINDIR", "C:/Windows")
    pdf_font_name = None
    font_candidates = [
        (os.path.join(windir, "Fonts", "msjh.ttf"), "CJK"),
        (os.path.join(windir, "Fonts", "mingliu.ttc"), "CJK"),
        (os.path.join(windir, "Fonts", "msjh.ttc"), "CJK"),
    ]
    for font_path, name in font_candidates:
        if not os.path.isfile(font_path):
            continue
        try:
            if font_path.lower().endswith(".ttc"):
                pdfmetrics.registerFont(TTFont(name, font_path, subfontIndex=0))
            else:
                pdfmetrics.registerFont(TTFont(name, font_path))
            pdf_font_name = name
            break
        except Exception:
            continue
    if not pdf_font_name:
        try:
            from reportlab.pdfbase.cidfonts import UnicodeCIDFont

            pdfmetrics.registerFont(UnicodeCIDFont("HeiseiMin-W3"))
            pdf_font_name = "HeiseiMin-W3"
        except Exception:
            pass
    if not pdf_font_name:
        return None
    try:
        doc = SimpleDocTemplate(buf, pagesize=A4, rightMargin=30, leftMargin=30, topMargin=40, bottomMargin=30)
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(name="CJKTitle", parent=styles["Title"], fontName=pdf_font_name, fontSize=16)
        heading_style = ParagraphStyle(name="CJKHeading2", parent=styles["Heading2"], fontName=pdf_font_name, fontSize=12)
        story = [Paragraph(f"<b>📉 總結表視覺化 {summary_year}</b>", title_style), Spacer(1, 12)]
        story.append(Paragraph("<b>📊 總結表數字</b>", heading_style))
        story.append(Spacer(1, 6))
        for ent in annual_summary_entity_labels:
            block = annual_viz.get("entities", {}).get(ent)
            if not block:
                continue
            story.append(Paragraph(f"<b>{summary_year} {ent}</b>", heading_style))
            story.append(Spacer(1, 6))
            _bt = block.get("by_type_df")
            if _bt is not None and not _bt.empty:
                data = [list(_bt.columns)] + _bt.astype(str).values.tolist()
                t = Table(data, repeatRows=1)
                t.setStyle(TableStyle([("FONTNAME", (0, 0), (-1, -1), pdf_font_name), ("GRID", (0, 0), (-1, -1), 0.5, colors.grey), ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#e0e0e0"))]))
                story.append(t)
                story.append(Spacer(1, 8))
            summary_table = pd.DataFrame([block.get("used_row", {}), block.get("unused_row", {}), block.get("usage_rate_row", {})])
            if not summary_table.empty:
                data = [list(summary_table.columns)] + summary_table.astype(str).values.tolist()
                t = Table(data, repeatRows=1)
                t.setStyle(TableStyle([("FONTNAME", (0, 0), (-1, -1), pdf_font_name), ("GRID", (0, 0), (-1, -1), 0.5, colors.grey), ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#e0e0e0"))]))
                story.append(t)
                story.append(Spacer(1, 10))
        doc.build(story)
        buf.seek(0)
        return buf.getvalue()
    except Exception:
        return None
