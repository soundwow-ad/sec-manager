import streamlit as st
import pandas as pd
import numpy as np
import sqlite3
import requests
import json
import time
import os
import uuid
import io
import re
import hashlib
from datetime import datetime, timedelta, date
from collections import Counter
from itertools import combinations
import random
import calendar


# ==========================================
# 1. 設定區 (Configuration)
# ==========================================

# 平台店數對照表 (Store Count Logic)
# ⚠️ 重要：請根據實際情況修正這些數值
# 以下為「平台名稱」直接對應；未列在此處者改由 REGION_STORE_COUNTS 依區域對應
STORE_COUNTS = {
    "新鮮視全省": 3124,
    "新鮮視北北基": 1127,
    "新鮮視中彰投": 528,
    "全家廣播": 4200,
    # 若找不到對應，改依區域查 REGION_STORE_COUNTS，再無則預設為 1
}

# 區域店數對照（對齊獎金表；新鮮視/企頻 等皆依「區域」取店數）
REGION_STORE_COUNTS = {
    "全省": 3124,
    "北北基": 1127,
    "桃竹苗": 616,
    "中彰投": 528,
    "高高屏": 405,
    "雲嘉南": 365,
    "宜花東": 83,
}

# 平台產能設定 (每日營業時間，單位：小時)
# 用於計算每日最大可容納秒數
PLATFORM_CAPACITY = {
    "新鮮視全省": 18,      # 每日營業 18 小時
    "新鮮視北北基": 18,
    "新鮮視中彰投": 18,
    "全家廣播": 18,
    # 預設值：18 小時
}

DB_FILE = "inventory_data.db"

# ==========================================
# Ragic 欄位流水號（由畫面截圖整理）
# 後續「從 Ragic 讀取專案資訊／下載 CUE Excel」會使用這份對照表。
# ==========================================
RAGIC_FIELDS = {
    # 訂檔資訊
    "訂檔性質": "1015324",
    "訂檔單號": "1015325",
    "建立日期": "1015326",
    "修改日期": "1015327",
    "申請人": "1015328",
    "業務(開發客戶)": "1015329",
    "業務主管": "1015330",
    "公司": "1015331",
    "CUE誌": "1015332",
    "波段": "1015333",
    "客服": "1015334",
    "文案": "1015335",
    "CUE": "1015336",
    "總波段": "1015337",
    "訂檔人": "1015338",
    "節配": "1015339",
    "客戶編": "1015340",
    "客戶": "1015343",
    "產業別": "1015346",
    "客戶類別": "1015347",
    "產品名稱": "1015349",
    "平台": "1015351",
    "電台": "1015352",
    "執行開始日期": "1015353",
    "執行結束日期": "1015354",
    "企頻素材": "1015355",
    "CUE表秒數": "1015356",
    "CUE表總檔數": "1015357",
    "訂檔CUE表": "1015359",
    "戶口受檔/業務回受檔": "1015361",
    "距上檔時數": "1015362",
    "電台播報開立名稱": "1015365",
    "平台(結案區)": "1015366",
    "素材走法": "1015367",
    "素材走法附件": "1015368",
    "付款條件": "1015370",
    "現折%": "1015371",
    "收款日期": "1015373",
    "退佣%": "1015374",
    "退佣%+現折%": "1015377",
    "除價買收換為退佣%+現折%數": "1015379",
    "個案確認": "1015566",
    "成本確認": "1015567",
    "純廣告秒數": "1015871",
    "個案狀況": "1015360",
    "發票開立": "1015829",
    "發票開立日": "1015830",
    # 需求/說明
    "需求或狀況說明": "1015383",
    "建立時間": "1015584",
}

# 子表（#xxxxx# 類型）：收入、素材需求等
RAGIC_SUBTABLE_FIELDS = {
    # 訂檔收入相關資訊（子表）
    "收入_類別": "1015360",
    "收入_業績歸屬": "1015389",
    "收入_平台細項": "1015390",
    "收入_平台檔數": "1015411",
    "收入_秒數": "1015412",
    "收入_製作成本x金額(未稅)": "1015391",
    "收入_製作成本x金額(未稅)_2": "1015392",
    "收入_除價買收(未稅)": "1015393",
    "收入_成本": "1015546",
    "收入_其他成本": "1015547",
    "收入_平台檔數總計": "1015461",
    "收入_實收金額總計(未稅)": "1015462",
    "收入_除價買收總計(未稅)": "1015463",
    "收入_電台總成本": "1015759",
    # 訂檔素材及需求資訊（子表）
    "素材_素材檔": "1015380",
    "素材_廣告檔名": "1015381",
    "素材_文案": "1015382",
}
# ==========================================
# 2. 核心邏輯區 (Core Logic)
# ==========================================

def get_db_connection():
    """取得資料庫連線"""
    conn = sqlite3.connect(DB_FILE)
    return conn


# --- 表3 與整頁重跑加速：依 DB 修改時間快取讀取與重計算 ---
@st.cache_data(ttl=120)
def _load_orders_cached(db_mtime):
    """依 DB 檔案修改時間快取 orders 讀取，DB 更新後自動失效。"""
    conn = get_db_connection()
    df = pd.read_sql("SELECT * FROM orders", conn)
    conn.close()
    return df


@st.cache_data(ttl=120)
def _load_segments_cached(db_mtime):
    """依 DB 檔案修改時間快取 ad_flight_segments 讀取，DB 更新後自動失效。"""
    conn = get_db_connection()
    try:
        df = pd.read_sql("SELECT * FROM ad_flight_segments", conn)
    except Exception:
        df = pd.DataFrame()
    conn.close()
    return df


@st.cache_data(ttl=120)
def _explode_segments_to_daily_cached(df_segments):
    """快取 explode_segments_to_daily，相同 segments 不重算。"""
    if df_segments.empty:
        return pd.DataFrame()
    return explode_segments_to_daily(df_segments)


@st.cache_data(ttl=120)
def _build_table3_monthly_control_cached(db_mtime, year, month, monthly_capacity_tuple):
    """快取表3 建表結果，以 db_mtime+年月+容量為鍵，不 hash 大 DataFrame，換月才約 1 秒內。"""
    df_seg = _load_segments_cached(db_mtime)
    df_daily = _explode_segments_to_daily_cached(df_seg) if not df_seg.empty else pd.DataFrame()
    if df_daily.empty or df_seg.empty:
        return {}
    cap = dict(monthly_capacity_tuple) if monthly_capacity_tuple else None
    return build_table3_monthly_control(df_daily, df_seg, None, year, month, cap)

def init_db():
    """初始化 SQLite 資料庫（包含 schema 檢查和遷移）"""
    conn = get_db_connection()
    c = conn.cursor()
    
    # 預期的 orders 表欄位順序和名稱
    expected_cols = ["id", "platform", "client", "product", "sales", "company", 
                     "start_date", "end_date", "seconds", "spots", "amount_net", "updated_at"]
    
    # 檢查 orders 表是否存在
    c.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='orders'")
    exists = c.fetchone() is not None
    
    if exists:
        # 檢查現有表的 schema 是否符合預期
        table_info = c.execute("PRAGMA table_info(orders)").fetchall()
        current_cols = [row[1] for row in table_info]  # row[1] 是欄位名稱
        
        # 如果缺少必要欄位（不含 contract_id），重建表；若有 contract_id 以外的差異則不重建
        required_core = [x for x in expected_cols if x != 'contract_id']
        has_core = all(col in current_cols for col in required_core)
        if not has_core:
            if current_cols != expected_cols:
                print(f"⚠️ 偵測到 orders 表 schema 不一致，將重建表")
                print(f"   目前欄位: {current_cols}")
                print(f"   預期欄位: {expected_cols}")
                c.execute("DROP TABLE IF EXISTS orders")
                conn.commit()
                exists = False
        else:
            # 遷移：若無 contract_id 則新增（一份合約可拆多列）
            if 'contract_id' not in current_cols:
                c.execute("ALTER TABLE orders ADD COLUMN contract_id TEXT")
                conn.commit()
            # 遷移：若無 seconds_type 則新增（秒數用途：銷售/交換/贈送/補檔/賀歲/公益）
            if 'seconds_type' not in current_cols:
                c.execute("ALTER TABLE orders ADD COLUMN seconds_type TEXT")
                conn.commit()
            # 遷移：專案實收金額（同一專案/合約填同一數字）、拆分金額（依比例拆分，ROI 用此計算）
            if 'project_amount_net' not in current_cols:
                c.execute("ALTER TABLE orders ADD COLUMN project_amount_net REAL")
                conn.commit()
            if 'split_amount' not in current_cols:
                c.execute("ALTER TABLE orders ADD COLUMN split_amount REAL")
                conn.commit()
    
    # 建立訂單主表（如果不存在或已刪除）
    if not exists:
        c.execute('''
            CREATE TABLE orders (
                id TEXT PRIMARY KEY,
                platform TEXT,
                client TEXT,
                product TEXT,
                sales TEXT,
                company TEXT,
                start_date TEXT,
                end_date TEXT,
                seconds INTEGER,
                spots INTEGER,
                amount_net REAL,
                updated_at TIMESTAMP,
                contract_id TEXT,
                seconds_type TEXT,
                project_amount_net REAL,
                split_amount REAL
            )
        ''')
    
    # 建立平台設定表（用於儲存自訂店數）
    c.execute('''
        CREATE TABLE IF NOT EXISTS platform_settings (
            platform TEXT PRIMARY KEY,
            store_count INTEGER,
            daily_hours INTEGER
        )
    ''')
    # 建立核心事實表：ad_flight_segments（檔次段）
    c.execute('''
        CREATE TABLE IF NOT EXISTS ad_flight_segments (
            segment_id TEXT PRIMARY KEY,
            source_order_id TEXT,
            platform TEXT,
            channel TEXT,
            region TEXT,
            company TEXT,
            sales TEXT,
            client TEXT,
            product TEXT,
            seconds INTEGER,
            spots INTEGER,
            start_date DATE,
            end_date DATE,
            duration_days INTEGER,
            store_count INTEGER,
            total_spots INTEGER,
            total_store_seconds INTEGER,
            seconds_type TEXT,
            created_at TIMESTAMP,
            updated_at TIMESTAMP
        )
    ''')
    # 遷移：ad_flight_segments 若無 media_platform 則新增（表一媒體平台切換用）
    try:
        table_info = c.execute("PRAGMA table_info(ad_flight_segments)").fetchall()
        current_cols = [row[1] for row in table_info]
        if 'media_platform' not in current_cols:
            c.execute("ALTER TABLE ad_flight_segments ADD COLUMN media_platform TEXT")
            conn.commit()
    except Exception:
        pass
    
    # 表3 用：各媒體當月「原始每日可用秒數」（向全家/家樂福等購買的當月每日秒數）
    c.execute('''
        CREATE TABLE IF NOT EXISTS platform_monthly_capacity (
            media_platform TEXT NOT NULL,
            year INTEGER NOT NULL,
            month INTEGER NOT NULL,
            daily_available_seconds INTEGER NOT NULL,
            PRIMARY KEY (media_platform, year, month)
        )
    ''')
    # 媒體採購：各媒體每年每月「購買秒數」與「購買價格」（供 ROI 換算成本、並可同步每日可用秒數）
    c.execute('''
        CREATE TABLE IF NOT EXISTS platform_monthly_purchase (
            media_platform TEXT NOT NULL,
            year INTEGER NOT NULL,
            month INTEGER NOT NULL,
            purchased_seconds INTEGER NOT NULL,
            purchase_price REAL NOT NULL,
            PRIMARY KEY (media_platform, year, month)
        )
    ''')
    # 使用者帳號表（模擬登入用，之後可遷移外部 DB）
    c.execute('''
        CREATE TABLE IF NOT EXISTS users (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            username TEXT UNIQUE NOT NULL,
            password_hash TEXT NOT NULL,
            role TEXT NOT NULL,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    ''')
    conn.commit()
    # 若尚無任何帳號，建立預設管理員 admin / admin123
    c.execute("SELECT COUNT(*) FROM users")
    if c.fetchone()[0] == 0:
        _hash = _hash_password("admin123")
        c.execute("INSERT INTO users (username, password_hash, role) VALUES (?, ?, ?)", ("admin", _hash, "行政主管"))
        conn.commit()
    conn.close()

# --- 登入與權限管理 ---
ROLES = ["行政主管", "業務", "總經理"]
SALT = "secmanager_2026"

def _hash_password(password):
    return hashlib.sha256((SALT + password).encode()).hexdigest()

def auth_verify(username, password):
    """驗證帳密，成功回傳 dict {username, role}，失敗回傳 None"""
    conn = get_db_connection()
    c = conn.cursor()
    h = _hash_password(password)
    c.execute("SELECT username, role FROM users WHERE username=? AND password_hash=?", (username.strip(), h))
    row = c.fetchone()
    conn.close()
    return {"username": row[0], "role": row[1]} if row else None

def auth_list_users():
    """列出所有帳號（不含密碼）"""
    conn = get_db_connection()
    df = pd.read_sql("SELECT id, username, role, created_at FROM users ORDER BY id", conn)
    conn.close()
    return df

def auth_create_user(username, password, role):
    """新增帳號，回傳 (success: bool, message: str)"""
    u = str(username).strip()
    if not u or not password:
        return False, "帳號與密碼不可為空"
    if role not in ROLES:
        return False, "無效的權限"
    conn = get_db_connection()
    c = conn.cursor()
    try:
        c.execute("INSERT INTO users (username, password_hash, role) VALUES (?, ?, ?)",
                  (u, _hash_password(password), role))
        conn.commit()
        conn.close()
        return True, "已新增"
    except sqlite3.IntegrityError:
        conn.close()
        return False, "帳號已存在"
    except Exception as e:
        conn.rollback()
        conn.close()
        return False, str(e)

def auth_delete_user(username):
    """刪除帳號"""
    conn = get_db_connection()
    c = conn.cursor()
    c.execute("DELETE FROM users WHERE username=?", (str(username).strip(),))
    conn.commit()
    conn.close()

def auth_change_password(username, new_password):
    """變更密碼"""
    if not new_password:
        return False, "密碼不可為空"
    conn = get_db_connection()
    c = conn.cursor()
    c.execute("UPDATE users SET password_hash=? WHERE username=?", (_hash_password(new_password), str(username).strip()))
    conn.commit()
    conn.close()
    return True

def get_platform_monthly_purchase(media_platform, year, month):
    """取得某媒體某年某月的購買秒數與購買價格，回傳 (purchased_seconds, purchase_price) 或 None"""
    conn = get_db_connection()
    c = conn.cursor()
    c.execute(
        'SELECT purchased_seconds, purchase_price FROM platform_monthly_purchase WHERE media_platform=? AND year=? AND month=?',
        (media_platform, int(year), int(month))
    )
    row = c.fetchone()
    conn.close()
    return row if row is not None else None

def set_platform_monthly_purchase(media_platform, year, month, purchased_seconds, purchase_price):
    """設定某媒體某年某月的購買秒數與購買價格；並同步更新 platform_monthly_capacity（每日可用 = 購買秒數/當月天數）"""
    import calendar
    conn = get_db_connection()
    c = conn.cursor()
    c.execute('''
        INSERT OR REPLACE INTO platform_monthly_purchase (media_platform, year, month, purchased_seconds, purchase_price)
        VALUES (?, ?, ?, ?, ?)
    ''', (media_platform, int(year), int(month), int(purchased_seconds), float(purchase_price)))
    ndays = calendar.monthrange(int(year), int(month))[1]
    daily_seconds = int(purchased_seconds) // ndays if ndays else 0
    c.execute('''
        INSERT OR REPLACE INTO platform_monthly_capacity (media_platform, year, month, daily_available_seconds)
        VALUES (?, ?, ?, ?)
    ''', (media_platform, int(year), int(month), daily_seconds))
    conn.commit()
    conn.close()

def load_platform_monthly_purchase_for_year(media_platform, year):
    """載入某媒體某年 1~12 月的購買資料，回傳 dict: month -> (purchased_seconds, purchase_price)"""
    conn = get_db_connection()
    c = conn.cursor()
    c.execute(
        'SELECT month, purchased_seconds, purchase_price FROM platform_monthly_purchase WHERE media_platform=? AND year=?',
        (media_platform, int(year))
    )
    out = {row[0]: (row[1], row[2]) for row in c.fetchall()}
    conn.close()
    return out

def load_platform_monthly_purchase_all_media_for_year(year):
    """載入某年所有媒體 1~12 月購買資料，回傳 dict: media_platform -> { month -> (purchased_seconds, purchase_price) }"""
    conn = get_db_connection()
    c = conn.cursor()
    c.execute(
        'SELECT media_platform, month, purchased_seconds, purchase_price FROM platform_monthly_purchase WHERE year=?',
        (int(year),)
    )
    out = {}
    for row in c.fetchall():
        mp, mo, sec, pr = row[0], row[1], row[2], row[3]
        if mp not in out:
            out[mp] = {}
        out[mp][mo] = (sec, pr)
    conn.close()
    return out


def generate_mock_platform_purchase_for_year(year):
    """
    產生某年度、各媒體 1～12 月的模擬採購資料（購買秒數與購買價格），數值合理、不爆量。
    會寫入 platform_monthly_purchase 並同步 platform_monthly_capacity。
    回傳 (success: bool, message: str)
    """
    import calendar
    # 各媒體基準：月購買秒數（店秒）、約略單價（元/秒），依月份 ±10% 變化
    base_per_media = {
        '全家廣播(企頻)': (1_600_000, 2.0),
        '全家新鮮視': (1_300_000, 2.2),
        '家樂福超市': (900_000, 2.4),
        '家樂福量販店': (700_000, 2.1),
    }
    try:
        for mp in MEDIA_PLATFORM_OPTIONS:
            base_sec, base_price_per_sec = base_per_media.get(mp, (1_000_000, 2.0))
            for m in range(1, 13):
                # 依月份略變：約 0.92～1.08 倍，讓每月不同但穩定
                var = 0.92 + (hash((year, mp, m)) % 17) / 100.0
                sec = int(base_sec * var)
                sec = max(100_000, min(sec, 5_000_000))
                price_per_sec = base_price_per_sec * (0.95 + (hash((year, mp, m + 10)) % 11) / 100.0)
                price_per_sec = max(0.8, min(price_per_sec, 4.0))
                price = int(sec * price_per_sec)
                price = max(50_000, min(price, 15_000_000))
                set_platform_monthly_purchase(mp, year, m, sec, price)
        return True, f"已產生 {len(MEDIA_PLATFORM_OPTIONS)} 個媒體、{year} 年 1～12 月模擬採購資料（已寫入並同步表3 每日可用秒數）"
    except Exception as e:
        return False, str(e)

def generate_mock_platform_purchase_for_year_with_capacity_check(year):
    """
    產生某年度、各媒體 1～12 月的模擬採購資料，但確保採購秒數 >= 實際使用秒數，
    避免覆蓋容量設定後導致使用率破千。
    購買價格依各媒體各月實收金額加入隨機比例（約 40% 機率產生正 ROI），呈現狀態多樣性。
    回傳 (success: bool, message: str)
    """
    import calendar
    try:
        conn = get_db_connection()
        df_seg_full = pd.read_sql("SELECT * FROM ad_flight_segments WHERE media_platform IS NOT NULL", conn)
        df_ord = pd.read_sql("SELECT id, split_amount FROM orders", conn)
        conn.close()
        usage_dict = {}
        revenue_dict = {}  # mp -> { month -> 實收金額 }
        if not df_seg_full.empty and not df_ord.empty:
            df_seg = df_seg_full[['source_order_id', 'media_platform', 'start_date', 'end_date']].merge(
                df_ord, left_on='source_order_id', right_on='id', how='left')
            df_seg['split_amount'] = pd.to_numeric(df_seg['split_amount'], errors='coerce').fillna(0)
            df_seg['start_date'] = pd.to_datetime(df_seg['start_date'], errors='coerce')
            df_seg['end_date'] = pd.to_datetime(df_seg['end_date'], errors='coerce')
            df_seg = df_seg.dropna(subset=['start_date', 'end_date'])
            df_daily = explode_segments_to_daily(df_seg_full)
            if not df_daily.empty and '媒體平台' in df_daily.columns and '使用店秒' in df_daily.columns and '日期' in df_daily.columns:
                df_daily['日期'] = pd.to_datetime(df_daily['日期'], errors='coerce')
                df_daily = df_daily.dropna(subset=['日期'])
                df_daily['年'] = df_daily['日期'].dt.year
                df_daily['月'] = df_daily['日期'].dt.month
                df_y = df_daily[df_daily['年'] == year]
                if not df_y.empty:
                    usage_by_media_month = df_y.groupby(['媒體平台', '月'])['使用店秒'].sum().reset_index()
                    for _, row in usage_by_media_month.iterrows():
                        mp, month = row['媒體平台'], int(row['月'])
                        if mp not in usage_dict:
                            usage_dict[mp] = {}
                        usage_dict[mp][month] = float(row['使用店秒'] or 0)
            # 依 segment 日期與月份重疊，按比例分配 split_amount 到各媒體各月
            for _, seg in df_seg.iterrows():
                mp = seg['media_platform']
                amt = float(seg['split_amount'] or 0)
                if amt <= 0:
                    continue
                s, e = seg['start_date'], seg['end_date']
                total_days = max(1, (e - s).days + 1)
                for m in range(1, 13):
                    ms = pd.Timestamp(year, m, 1)
                    _, nd = calendar.monthrange(year, m)
                    me = pd.Timestamp(year, m, nd)
                    overlap_start = max(s, ms)
                    overlap_end = min(e, me)
                    if overlap_start <= overlap_end:
                        overlap_days = (overlap_end - overlap_start).days + 1
                        prorate = amt * (overlap_days / total_days)
                        if mp not in revenue_dict:
                            revenue_dict[mp] = {}
                        revenue_dict[mp][m] = revenue_dict[mp].get(m, 0) + prorate
        
        base_per_media = {
            '全家廣播(企頻)': (1_600_000, 2.0),
            '全家新鮮視': (1_300_000, 2.2),
            '家樂福超市': (900_000, 2.4),
            '家樂福量販店': (700_000, 2.1),
        }
        for mp in MEDIA_PLATFORM_OPTIONS:
            base_sec, base_price_per_sec = base_per_media.get(mp, (1_000_000, 2.0))
            for m in range(1, 13):
                var = 0.92 + (hash((year, mp, m)) % 17) / 100.0
                sec = int(base_sec * var)
                sec = max(100_000, min(sec, 5_000_000))
                if mp in usage_dict and m in usage_dict[mp]:
                    used_sec = usage_dict[mp][m]
                    min_sec = int(used_sec * 1.2)
                    sec = max(sec, min_sec)
                rev = revenue_dict.get(mp, {}).get(m, 0)
                if rev > 0:
                    # 隨機比例：約 40% 機率 購買成本 < 實收 → 正 ROI
                    margin = random.uniform(-0.35, 0.55)
                    price = max(10_000, int(rev * (1 + margin)))
                else:
                    price_per_sec = base_price_per_sec * (0.95 + (hash((year, mp, m + 10)) % 11) / 100.0)
                    price_per_sec = max(0.8, min(price_per_sec, 4.0))
                    price = max(50_000, min(int(sec * price_per_sec), 15_000_000))
                conn = get_db_connection()
                c = conn.cursor()
                c.execute('''
                    INSERT OR REPLACE INTO platform_monthly_purchase (media_platform, year, month, purchased_seconds, purchase_price)
                    VALUES (?, ?, ?, ?, ?)
                ''', (mp, int(year), int(m), int(sec), float(price)))
                conn.commit()
                conn.close()
        return True, f"已產生 {len(MEDIA_PLATFORM_OPTIONS)} 個媒體、{year} 年 1～12 月模擬採購資料（採購秒數已確保 >= 實際使用秒數；約 40% 機率正 ROI）"
    except Exception as e:
        return False, f"產生採購資料失敗：{e}"

def load_platform_settings():
    """從資料庫載入平台設定（優先使用資料庫中的設定）"""
    conn = get_db_connection()
    c = conn.cursor()
    settings = {}
    for row in c.execute('SELECT platform, store_count, daily_hours FROM platform_settings'):
        settings[row[0]] = {'store_count': row[1], 'daily_hours': row[2]}
    conn.close()
    return settings

def get_platform_monthly_capacity(media_platform, year, month):
    """取得某媒體、某年某月的「當月每日可用秒數」（向全家/家樂福等購買的每日秒數），無設定則回傳 None"""
    conn = get_db_connection()
    c = conn.cursor()
    c.execute(
        'SELECT daily_available_seconds FROM platform_monthly_capacity WHERE media_platform=? AND year=? AND month=?',
        (media_platform, int(year), int(month))
    )
    row = c.fetchone()
    conn.close()
    return row[0] if row is not None else None

def set_platform_monthly_capacity(media_platform, year, month, daily_available_seconds):
    """設定某媒體、某年某月的當月每日可用秒數"""
    conn = get_db_connection()
    c = conn.cursor()
    c.execute('''
        INSERT OR REPLACE INTO platform_monthly_capacity (media_platform, year, month, daily_available_seconds)
        VALUES (?, ?, ?, ?)
    ''', (media_platform, int(year), int(month), int(daily_available_seconds)))
    conn.commit()
    conn.close()

def load_platform_monthly_capacity_for(year, month):
    """載入某年某月所有媒體的每日可用秒數設定，回傳 dict: media_platform -> daily_available_seconds"""
    conn = get_db_connection()
    c = conn.cursor()
    c.execute('SELECT media_platform, daily_available_seconds FROM platform_monthly_capacity WHERE year=? AND month=?', (int(year), int(month)))
    out = {row[0]: row[1] for row in c.fetchall()}
    conn.close()
    return out

def save_platform_settings(platform, store_count, daily_hours):
    """儲存平台設定到資料庫"""
    conn = get_db_connection()
    c = conn.cursor()
    c.execute('''
        INSERT OR REPLACE INTO platform_settings (platform, store_count, daily_hours)
        VALUES (?, ?, ?)
    ''', (platform, store_count, daily_hours))
    conn.commit()
    conn.close()

def parse_platform_region(raw_platform):
    """
    將原始平台名稱拆解為 (platform, channel, region)
    例如：'新鮮視全省' → ('全家', '新鮮視', '全省')
    """
    if not raw_platform or pd.isna(raw_platform):
        return '其他', '其他', '未知'
    
    raw_platform = str(raw_platform)
    # 錯字對應：宜花束 → 宜花東
    if '宜花束' in raw_platform:
        raw_platform = raw_platform.replace('宜花束', '宜花東')
    
    # 判斷平台和頻道
    if '新鮮視' in raw_platform:
        platform = '全家'
        channel = '新鮮視'
    elif '企頻' in raw_platform or ('廣播' in raw_platform and '全家' in raw_platform):
        platform = '全家'
        channel = '企頻'
    elif '家樂福' in raw_platform:
        platform = '家樂福'
        channel = '廣播'
    elif raw_platform.strip() in ('企頻', 'RADIO', '企業頻道', '全家廣播'):
        # Google Sheet 常只填「企頻」或「RADIO」等，仍視為全家企頻
        platform = '全家'
        channel = '企頻'
    else:
        platform = '其他'
        channel = '其他'
    
    # 判斷區域
    region = '未知'
    for r in ['全省', '北北基', '中彰投', '桃竹苗', '高高屏', '雲嘉南', '宜花東']:
        if r in raw_platform:
            region = r
            break
    
    return platform, channel, region

# 表一「媒體平台」顯示名稱：全家廣播(企頻)、全家新鮮視、家樂福超市、家樂福量販店
MEDIA_PLATFORM_OPTIONS = ['全家廣播(企頻)', '全家新鮮視', '家樂福超市', '家樂福量販店']

def get_media_platform_display(platform, channel, raw_platform=''):
    """
    依 platform / channel / 原始平台名稱 回傳表一用「媒體平台」顯示名稱。
    回傳值為 MEDIA_PLATFORM_OPTIONS 之一或 '其他'。
    """
    raw = str(raw_platform or '')
    if platform == '全家' and channel == '企頻':
        return '全家廣播(企頻)'
    if platform == '全家' and channel == '新鮮視':
        return '全家新鮮視'
    if platform == '家樂福':
        return '家樂福量販店' if '量販' in raw else '家樂福超市'
    return '其他'


def should_multiply_store_count(media_platform: str) -> bool:
    """
    使用秒數計算規則（重要）：
    - 全家廣播(企頻)、全家新鮮視：使用店秒 = 檔次 × 秒數 × 店數
    - 其他（如 家樂福超市/量販店、診所/門診等）：使用秒數 = 檔次 × 秒數（不乘店數）

    注意：系統內部仍沿用欄名「使用店秒」，但在不乘店數的平台其意義等同「使用秒數」。
    """
    mp = (media_platform or '').strip()
    return mp in ('全家廣播(企頻)', '全家新鮮視')

def get_store_count(platform, custom_settings=None):
    """取得平台店數（優先使用自訂設定，其次平台鍵，再依區域對照，最後預設 1）"""
    if custom_settings and platform in custom_settings:
        return custom_settings[platform]['store_count']
    if platform in STORE_COUNTS:
        return STORE_COUNTS[platform]
    # 依區域對照：新鮮視/企頻 等「平台名含區域」皆可由此取得店數
    try:
        _, _, region = parse_platform_region(platform)
        if region and region != '未知' and region in REGION_STORE_COUNTS:
            return REGION_STORE_COUNTS[region]
    except Exception:
        pass
    # 家樂福超市/家樂福量販店 等未列在 STORE_COUNTS 時，fallback 至家樂福
    if platform and '家樂福' in str(platform):
        return STORE_COUNTS.get('家樂福', 1)
    return 1

def get_daily_capacity(platform, custom_settings=None):
    """計算平台每日最大容量（店數 × 每日小時數 × 3600秒）"""
    store_count = get_store_count(platform, custom_settings)
    
    # 取得每日營業小時數
    if custom_settings and platform in custom_settings:
        daily_hours = custom_settings[platform]['daily_hours']
    else:
        daily_hours = PLATFORM_CAPACITY.get(platform, 18)
    
    # 計算每日最大秒數容量
    return store_count * daily_hours * 3600

def df_to_excel_bytes(df, sheet_name="Sheet1"):
    """
    將 DataFrame 轉換為 Excel (.xlsx) 格式的 bytes
    這是解決 CSV 編碼問題的最佳方案
    """
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name=sheet_name)
    output.seek(0)
    return output.getvalue()


def sanitize_dataframe_for_display(df):
    """
    清理 DataFrame，將複雜類型轉換為字符串，以便 PyArrow 可以正確序列化
    用於修復 st.dataframe() 的 PyArrow 錯誤
    
    參數:
        df: pandas DataFrame
    
    返回:
        清理後的 DataFrame
    """
    df = df.copy()
    for col in df.columns:
        if df[col].dtype == 'object':
            # 將所有複雜類型（列表、字典等）轉換為字符串
            # 使用更安全的方式處理 NaN 和 None
            def safe_convert(x):
                try:
                    # 先檢查是否為 NaN/None
                    if x is None:
                        return ''
                    if pd.isna(x):
                        return ''
                    # 嘗試轉換為字符串
                    return str(x)
                except (TypeError, ValueError):
                    # 如果轉換失敗，返回空字符串
                    return ''
            
            df[col] = df[col].apply(safe_convert)
    return df


def _styler_one_decimal(df):
    """各分頁表格用：數值欄位顯示最多小數點第一位，超過三位數時自動加千分位。回傳 Styler 供 st.dataframe 使用。"""
    if df is None:
        return None
    if df.empty:
        return df.style
    num_cols = df.select_dtypes(include=[np.number]).columns.tolist()
    if not num_cols:
        return df.style
    # 使用 {:,.1f} 讓超過三位數的數字自動加千分位（如 1,234.5）
    return df.style.format({c: "{:,.1f}" for c in num_cols})


def _display_monthly_table_split(df, month_cols, style_func=None, height=None, key_prefix=""):
    """
    將包含 12 個月欄位的表格拆分成上下半年兩個表格顯示，避免左右滑動。
    將 12 個月分成 2 組：上半年（1-6月）、下半年（7-12月），垂直排列顯示。
    
    參數:
        df: DataFrame，必須包含 month_cols 中的欄位
        month_cols: 月份欄位列表，例如 ['1月', '2月', ..., '12月']
        style_func: 可選的樣式函數，接受 DataFrame 並回傳 Styler
        height: 可選的表格高度
        key_prefix: 用於生成唯一 key 的前綴
    """
    if df.empty or not month_cols:
        return
    
    # 將 12 個月分成 2 組：上半年和下半年
    groups = [
        (month_cols[0:6], "上半年（1月～6月）"),   # 1-6月
        (month_cols[6:12], "下半年（7月～12月）"),  # 7-12月
    ]
    
    # 取得非月份欄位（例如「項目」欄位）
    non_month_cols = [c for c in df.columns if c not in month_cols]
    
    # 垂直排列顯示兩個表格
    for idx, (group_months, label) in enumerate(groups):
        # 選取該組的欄位
        display_cols = non_month_cols + group_months
        df_subset = df[[c for c in display_cols if c in df.columns]].copy()
        
        if df_subset.empty:
            continue
        
        # 顯示標題
        st.markdown(f"**{label}**")
        
        # 套用樣式
        if style_func:
            styled_df = style_func(df_subset)
        else:
            styled_df = df_subset.style
        
        # 顯示表格（一個一列，垂直排列）
        st.dataframe(
            styled_df,
            use_container_width=True,
            height=height,
            key=f"{key_prefix}_split_{idx}"
        )


def read_cue_excel(file_content, max_rows=100):
    """
    讀取 CUE Excel 檔案內容
    
    參數:
        file_content: Excel 檔案的 bytes 內容
        max_rows: 最多顯示的行數（避免過大）
    
    返回:
        dict: {
            'sheets': [sheet_name1, sheet_name2, ...],
            'data': {
                'sheet_name1': DataFrame,
                'sheet_name2': DataFrame,
                ...
            },
            'error': None or error_message
        }
    """
    try:
        # 使用 BytesIO 讀取 Excel
        excel_file = io.BytesIO(file_content)
        
        # 讀取所有工作表
        excel_file.seek(0)
        xls = pd.ExcelFile(excel_file, engine='openpyxl')
        
        result = {
            'sheets': xls.sheet_names,
            'data': {},
            'error': None
        }
        
        # 讀取每個工作表（限制行數以避免過大）
        for sheet_name in xls.sheet_names:
            try:
                df = pd.read_excel(excel_file, sheet_name=sheet_name, header=None, engine='openpyxl')
                # 限制行數
                if len(df) > max_rows:
                    df = df.head(max_rows)
                    result['data'][sheet_name] = df
                    result['truncated'] = True
                else:
                    result['data'][sheet_name] = df
            except Exception as e:
                result['data'][sheet_name] = pd.DataFrame()
                result['error'] = f"讀取工作表 '{sheet_name}' 失敗: {str(e)}"
        
        excel_file.close()
        return result
        
    except Exception as e:
        return {
            'sheets': [],
            'data': {},
            'error': f"讀取 Excel 檔案失敗: {str(e)}"
        }


# ================= Cueapp Excel 專用解析（東吳／聲活／鉑霖三種格式）=================
def _parse_cueapp_period_dongwu(row_b5_value):
    """從東吳格式 B5 儲存格解析 Period : YYYY. MM. DD - YYYY. MM. DD"""
    if pd.isna(row_b5_value):
        return None, None
    s = str(row_b5_value).strip()
    # 可能是 datetime 或 "2026. 01. 31 - 2026. 02. 06"
    if hasattr(row_b5_value, 'date'):
        return row_b5_value.date(), row_b5_value.date()
    m = re.search(r'(\d{4})\s*[.\-/]\s*(\d{1,2})\s*[.\-/]\s*(\d{1,2})\s*[-~－]\s*(\d{4})\s*[.\-/]\s*(\d{1,2})\s*[.\-/]\s*(\d{1,2})', s)
    if m:
        try:
            start = datetime(int(m.group(1)), int(m.group(2)), int(m.group(3)))
            end = datetime(int(m.group(4)), int(m.group(5)), int(m.group(6)))
            return start.date(), end.date()
        except (ValueError, TypeError):
            pass
    return None, None

def _parse_cueapp_period_shenghuo_bolin(df, search_rows=(3, 4, 5)):
    """從聲活/鉑霖格式中找「執行期間：YYYY.MM.DD - YYYY.MM.DD」"""
    for ri in search_rows:
        if ri >= len(df):
            continue
        row_text = df.iloc[ri].fillna('').astype(str).str.cat(sep=' ')
        m = re.search(r'執行期間[：:]\s*(\d{4})[.\-/](\d{1,2})[.\-/](\d{1,2})\s*[-~－]\s*(\d{4})[.\-/](\d{1,2})[.\-/](\d{1,2})', row_text)
        if m:
            try:
                start = datetime(int(m.group(1)), int(m.group(2)), int(m.group(3))).date()
                end = datetime(int(m.group(4)), int(m.group(5)), int(m.group(6))).date()
                return start, end
            except (ValueError, TypeError):
                pass
    return None, None

def _cell_val(v):
    """Excel 儲存格轉單一值；日期轉 date。"""
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return None
    if hasattr(v, 'date'):
        return v.date() if hasattr(v, 'date') else v
    return v

def _safe_spots(val):
    """將儲存格轉為檔次整數，無效則 0。"""
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return 0
    try:
        n = int(round(float(val)))
        return n if 0 <= n <= 10000 else 0
    except (ValueError, TypeError):
        return 0

def _extract_seconds_from_cell(val):
    """從「15秒」「15秒廣告」等字串抽出秒數。"""
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return 0
    s = str(val).strip()
    m = re.search(r'(\d+)\s*秒', s)
    if m:
        try:
            sec = int(m.group(1))
            if 5 <= sec <= 120:
                return sec
        except ValueError:
            pass
    return 0

def parse_cueapp_excel(file_content):
    """
    解析 Cue Sheet Pro (cueapp) 產生的 Excel，支援東吳、聲活、鉑霖三種格式。
    與 parse_cue_excel_for_table1 回傳格式一致，供表1／檔次段使用。
    回傳: list of dict (ad_unit)，若無法辨識為 cueapp 或解析失敗則回傳 []。
    """
    result = []
    try:
        excel_file = io.BytesIO(file_content)
        xls = pd.ExcelFile(excel_file, engine='openpyxl')
    except Exception:
        return []

    for sheet_name in xls.sheet_names:
        try:
            excel_file.seek(0)
            df = pd.read_excel(excel_file, sheet_name=sheet_name, header=None, engine='openpyxl')
            if df.empty or len(df) < 9:
                continue
            # 辨識格式
            row0_text = df.iloc[0].fillna('').astype(str).str.cat(sep=' ')
            fmt = None
            if 'Media Schedule' in row0_text or (len(df.columns) > 0 and str(df.iloc[0, 0]).strip() == 'Media Schedule'):
                fmt = 'dongwu'
            elif '聲活數位' in row0_text:
                fmt = 'shenghuo'
            elif '鉑霖行動行銷' in row0_text or '鉑霖' in row0_text:
                fmt = 'bolin'

            if fmt is None:
                # 工作表名為 1月、2月 等且內容像東吳（有 Period 在 row 5）
                if re.match(r'^\d+月$', str(sheet_name).strip()):
                    b5 = df.iloc[4, 1] if df.shape[1] > 1 else None
                    start, end = _parse_cueapp_period_dongwu(b5)
                    if start and end:
                        fmt = 'dongwu'
                if fmt is None:
                    continue

            start_date, end_date = None, None
            date_start_col = None
            eff_days = None
            header_row_idx = None

            def _find_schedule_header_row(_df: pd.DataFrame):
                def _row_text(i: int) -> str:
                    try:
                        return _df.iloc[i].fillna('').astype(str).str.cat(sep=' ')
                    except Exception:
                        return ''
                for i in range(min(40, len(_df))):
                    t = _row_text(i)
                    if ('頻道' in t and '播出地區' in t and '秒數' in t) or ('Station' in t and 'Location' in t and ('Size' in t or '秒數' in t)):
                        return i
                return None

            def _parse_day_cell(v):
                v = _cell_val(v)
                if isinstance(v, (datetime, date)):
                    # 1900-01-20 這種，取 day=20
                    try:
                        return int(v.day)
                    except Exception:
                        return None
                try:
                    import numbers as _numbers
                    is_num = isinstance(v, (_numbers.Integral, _numbers.Real))
                except Exception:
                    is_num = isinstance(v, (int, float))
                if is_num and not pd.isna(v):
                    try:
                        n = int(round(float(v)))
                        return n if 1 <= n <= 31 else None
                    except Exception:
                        return None
                if isinstance(v, str):
                    s = v.strip()
                    if s.isdigit():
                        n = int(s)
                        return n if 1 <= n <= 31 else None
                return None

            def _infer_year_from_df(_df: pd.DataFrame):
                try:
                    for i in range(min(25, len(_df))):
                        for j in range(min(15, _df.shape[1])):
                            s = str(_df.iloc[i, j]) if _df.iloc[i, j] is not None else ''
                            m = re.search(r'(20\d{2})', s)
                            if m:
                                y = int(m.group(1))
                                if 2000 <= y <= 2100:
                                    return y
                except Exception:
                    pass
                return None

            def _infer_month_for_col(_df: pd.DataFrame, header_i: int, col_j: int):
                # 先找同欄往上最近的「x月」
                for i in range(max(0, header_i - 6), header_i):
                    try:
                        s = str(_df.iloc[i, col_j]).strip()
                        m = re.search(r'(\d{1,2})\s*月', s)
                        if m:
                            mm = int(m.group(1))
                            if 1 <= mm <= 12:
                                return mm
                    except Exception:
                        continue
                # 再找同列往左最近的「x月」
                for j in range(col_j, -1, -1):
                    try:
                        s = str(_df.iloc[header_i - 1, j]).strip()
                        m = re.search(r'(\d{1,2})\s*月', s)
                        if m:
                            mm = int(m.group(1))
                            if 1 <= mm <= 12:
                                return mm
                    except Exception:
                        continue
                return None

            if fmt == 'dongwu':
                b5 = df.iloc[4, 1] if df.shape[1] > 1 else None
                start_date, end_date = _parse_cueapp_period_dongwu(b5)
                if not start_date or not end_date:
                    continue
                date_start_col = 7
                header_row_idx = 6
                # 找「檔次」欄：通常在最右側日期欄之後
                for c in range(df.shape[1] - 1, date_start_col - 1, -1):
                    try:
                        val = str(df.iloc[header_row_idx, c]).strip() + str(df.iloc[header_row_idx + 1, c]).strip()
                        if '檔次' in val:
                            eff_days = c - date_start_col
                            break
                    except IndexError:
                        continue
                if eff_days is None:
                    eff_days = max(0, df.shape[1] - date_start_col - 1)
            else:
                start_date, end_date = _parse_cueapp_period_shenghuo_bolin(df)
                # Schedule/Media Schedule 常見沒有「檔次」欄，改用「日期數字欄」推導
                header_row_idx = _find_schedule_header_row(df)
                if header_row_idx is None:
                    continue
                # 找「秒數/Size」欄位置，日期欄從其右側開始
                sec_col = None
                for j in range(min(25, df.shape[1])):
                    s = str(df.iloc[header_row_idx, j]).strip()
                    if ('秒數' in s) or (s.lower() == 'size') or ('size' in s.lower()):
                        sec_col = j
                        break
                if sec_col is None:
                    continue
                date_start_col = sec_col + 1
                # 解析日期欄（以 header row 的 day number 連續欄位為準）
                day_cols = []
                for j in range(date_start_col, min(df.shape[1], date_start_col + 80)):
                    d = _parse_day_cell(df.iloc[header_row_idx, j])
                    if d is None:
                        if day_cols:
                            break
                        continue
                    day_cols.append((j, d))
                if not day_cols:
                    continue
                eff_days = len(day_cols)

                # 嘗試推導 start/end（若原本解析不到）
                year = _infer_year_from_df(df) or (start_date.year if start_date else None)
                if year is None:
                    year = datetime.now().year
                # 逐欄推導月份：優先用「x月」標記，否則用遞增/回跳推斷
                months = []
                last_day = None
                last_month = None
                base_month = start_date.month if start_date else None
                for (j, d) in day_cols:
                    mm = _infer_month_for_col(df, header_row_idx, j) or base_month
                    if mm is None:
                        if last_month is None:
                            mm = 1
                        else:
                            mm = last_month
                    if last_day is not None and d < last_day and (mm == last_month):
                        # day 回跳但月沒變，推進一個月
                        mm = 1 if last_month == 12 else (last_month + 1)
                    months.append(mm)
                    last_day = d
                    last_month = mm

                dates = []
                for (_, d), mm in zip(day_cols, months):
                    try:
                        dates.append(date(year, int(mm), int(d)))
                    except Exception:
                        dates.append(None)
                # 濾掉 None
                dates = [dt for dt in dates if dt is not None]
                if not dates:
                    continue
                start_date = start_date or min(dates)
                end_date = end_date or max(dates)

            if eff_days is None or eff_days <= 0:
                continue
            # dates_str：若為 schedule 型，使用 header 的 dates；否則用連續日期
            dates_str = None
            if fmt != 'dongwu' and header_row_idx is not None and date_start_col is not None:
                # 重新嘗試以 header day_cols 組 dates_str（避免 start/end 連續推斷不準）
                try:
                    # 以同樣邏輯重取 day_cols（已算過 eff_days）
                    day_cols2 = []
                    for j in range(date_start_col, min(df.shape[1], date_start_col + 80)):
                        d = _parse_day_cell(df.iloc[header_row_idx, j])
                        if d is None:
                            if day_cols2:
                                break
                            continue
                        day_cols2.append((j, d))
                    if day_cols2:
                        year2 = _infer_year_from_df(df) or (start_date.year if start_date else datetime.now().year)
                        months2 = []
                        last_day2 = None
                        last_month2 = start_date.month if start_date else None
                        for (j, d) in day_cols2:
                            mm = _infer_month_for_col(df, header_row_idx, j) or last_month2
                            if mm is None:
                                mm = 1
                            if last_day2 is not None and d < last_day2 and (mm == last_month2):
                                mm = 1 if last_month2 == 12 else (last_month2 + 1)
                            months2.append(mm)
                            last_day2 = d
                            last_month2 = mm
                        dates2 = []
                        for (_, d), mm in zip(day_cols2, months2):
                            try:
                                dates2.append(date(int(year2), int(mm), int(d)))
                            except Exception:
                                pass
                        if dates2:
                            dates_str = [dt.strftime('%Y-%m-%d') for dt in dates2]
                            eff_days = len(dates_str)
                except Exception:
                    dates_str = None
            if not dates_str:
                date_list = pd.date_range(start_date, end_date, freq='D')
                if len(date_list) != eff_days:
                    date_list = date_list[:eff_days]
                dates_str = [d.strftime('%Y-%m-%d') for d in date_list]

            data_start_row = header_row_idx + 2
            platform_info = _extract_platform_from_sheet(df, sheet_name)
            seconds_info = _extract_seconds_from_sheet(df, sheet_name)
            default_seconds = seconds_info.get('seconds', 0)

            for r in range(data_start_row, min(data_start_row + 200, len(df))):
                row = df.iloc[r]
                try:
                    e_val = row.iloc[4] if len(row) > 4 else None
                    e_str = str(e_val).strip() if e_val is not None else ''
                    if 'Total' in e_str or 'total' in e_str or e_str == 'Total':
                        break
                    first_cell = str(row.iloc[0]).strip() if len(row) > 0 else ''
                    if not first_cell or first_cell == 'nan':
                        continue
                    region_cell = row.iloc[1] if len(row) > 1 else ''
                    region = str(region_cell).strip() if region_cell is not None and str(region_cell) != 'nan' else platform_info.get('region', '全省')
                    # 秒數欄：dongwu/shenghuo/bolin 可能在第 4 欄或 header 找到的 sec_col
                    sec_cell = None
                    try:
                        if fmt != 'dongwu' and date_start_col is not None and date_start_col >= 1:
                            sec_cell = row.iloc[date_start_col - 1]
                        else:
                            sec_cell = row.iloc[4] if len(row) > 4 else None
                    except Exception:
                        sec_cell = row.iloc[4] if len(row) > 4 else None
                    sec = _extract_seconds_from_cell(sec_cell)
                    if sec <= 0:
                        sec = default_seconds
                    daily_spots = []
                    for c in range(date_start_col, date_start_col + min(eff_days, len(dates_str))):
                        if c < len(row):
                            daily_spots.append(_safe_spots(row.iloc[c]))
                        else:
                            daily_spots.append(0)
                    if len(daily_spots) < len(dates_str):
                        daily_spots.extend([0] * (len(dates_str) - len(daily_spots)))
                    daily_spots = daily_spots[:len(dates_str)]
                    if len([s for s in daily_spots if s > 0]) < 1:
                        continue
                    split_groups = _split_by_spots_change(daily_spots, dates_str, dates_str[0] if dates_str else None, dates_str[-1] if dates_str else None)
                    for group in split_groups:
                        ad_unit = {
                            'platform': platform_info.get('platform', '未知'),
                            'platform_category': platform_info.get('category', '其他'),
                            'seconds': sec,
                            'region': region,
                            'ad_name': first_cell,
                            'daily_spots': group.get('daily_spots_list', [group['daily_spots']] * group['days']),
                            'dates': group.get('dates', []),
                            'start_date': group.get('start_date', ''),
                            'end_date': group.get('end_date', ''),
                            'total_spots': sum(group.get('daily_spots_list', [])),
                            'days': group.get('days', 0),
                            'source_sheet': sheet_name,
                            'source_row': r,
                            'split_reason': group.get('split_reason', 'none'),
                            'split_groups': [group],
                        }
                        if ad_unit['total_spots'] == 0:
                            ad_unit['total_spots'] = sum(ad_unit['daily_spots'])
                        result.append(ad_unit)
                except (IndexError, KeyError, ValueError, TypeError):
                    continue
        except Exception:
            continue

    try:
        excel_file.close()
    except Exception:
        pass
    return result


# ================= Excel 解析工具函數（參考 V29 邏輯）=================
SECONDS_BLACKLIST = {5, 10, 15, 20, 30, 40, 60}
YEAR_BLACKLIST = {114, 115, 116, 2025, 2026}

def safe_int_v29(v, target=None):
    """安全地將值轉換為整數（V29 邏輯）"""
    try:
        f = float(v)
        if abs(f - round(f)) > 1e-3:
            return None
        f = int(round(f))
        
        if target and f != target:
            if f in SECONDS_BLACKLIST:
                return None
            if f in YEAR_BLACKLIST:
                return None
        
        if 0 < f <= 50000:
            return f
    except:
        return None
    return None

def is_noise_row_v29(text):
    """判斷是否為噪音行（V29 邏輯）"""
    noise = ['元', '$', '含稅', '未稅', 'VAT', 'COST', 'PRICE', '報價', '金額', '製作費', '費用', '日期', '結案', '發票']
    return any(x in text for x in noise)

def is_store_count_row_v29(text, nums):
    """判斷是否為店數行（V29 邏輯）"""
    keywords = ['門市', '店數', '間門市', '約', '覆蓋', '店家', '家數']
    if any(k in text for k in keywords):
        if len(nums) <= 2 and max(nums) > 100:
            return True
    return False

def semantic_bonus_v29(text):
    """語義加分（V29 邏輯）"""
    bonus = 0
    if any(x in text for x in ['全家', '家樂福', '區域', '北', '中', '南', '通路', 'RADIO', 'VISION', '廣播', '店舖']):
        bonus += 3
    if any(x in text for x in ['每日', '明細', 'LIST']):
        bonus -= 2
    return bonus

def extract_row_signatures_v29(df, sheet_name, target=None):
    """
    從 DataFrame 中提取行簽名（參考 V29 邏輯）
    用於識別可能的每日檔次資料
    
    返回:
        list of dict: 每行包含 row_idx, nums, text, level, unit_val, bonus 等資訊
    """
    rows = []
    for idx in range(len(df)):
        row = df.iloc[idx]
        nums = [safe_int_v29(v, target) for v in row if safe_int_v29(v, target) is not None]
        if len(nums) < 1:
            continue
        
        text = row.astype(str).str.cat(sep=' ').upper()
        if is_noise_row_v29(text):
            continue
        if is_store_count_row_v29(text, nums):
            continue
        
        if len(nums) > 2:
            big_nums = [n for n in nums if n > 1000]
            small_nums = [n for n in nums if n <= 200]
            if big_nums and small_nums:
                if target and target not in big_nums:
                    nums = small_nums
        
        unit_val = None
        if len(nums) >= 2:
            c = Counter(nums)
            most_common, count = c.most_common(1)[0]
            if count >= 3 or count / len(nums) > 0.3:
                if target and most_common in SECONDS_BLACKLIST and most_common != target:
                    pass
                elif target and most_common in YEAR_BLACKLIST and most_common != target:
                    pass
                else:
                    unit_val = most_common
        
        level = "L3"
        if len(nums) == 1:
            level = "L1"
        else:
            max_n = max(nums)
            if max_n >= sum(nums) * 0.4:
                level = "L2"
        
        rows.append({
            "sheet": sheet_name,
            "row_idx": idx,
            "sum": sum(nums),
            "nums": nums,
            "unit_val": unit_val,
            "count": len(nums),
            "text": text,
            "bonus": semantic_bonus_v29(text),
            "level": level,
            "raw_row": row.tolist()  # 保留原始行資料
        })
    return rows

def parse_excel_daily_ads(file_content, target_spots=None):
    """
    解析 Excel CUE 表，提取每日廣告檔次資訊（人機協作版本）
    
    參數:
        file_content: Excel 檔案的 bytes 內容
        target_spots: 目標檔次（用於驗證）
    
    返回:
        dict: {
            'file_hash': str,  # 檔案 hash
            'file_name': str,
            'ai_interpretations': [
                {
                    'sheet': str,
                    'row_idx': int,
                    'col_idx': int,
                    'date': str,  # 推測的日期
                    'ad_name': str,  # 推測的廣告名稱
                    'spots': int,  # AI 判斷的檔次
                    'seconds': int,  # AI 判斷的秒數
                    'confidence': str,  # 信心等級
                    'rule_used': str,  # 使用的規則
                    'reason': str,  # 判斷原因
                    'raw_value': any,  # 原始值
                }
            ],
            'raw_data': {...},  # 原始 Excel 資料
            'error': str or None
        }
    """
    # 計算檔案 hash
    file_hash = hashlib.md5(file_content).hexdigest()
    
    result = {
        'file_hash': file_hash,
        'file_name': '',
        'ai_interpretations': [],
        'raw_data': {},
        'error': None
    }
    
    try:
        excel_file = io.BytesIO(file_content)
        excel_file.seek(0)
        xls = pd.ExcelFile(excel_file, engine='openpyxl')
        
        result['file_name'] = 'cue_file.xlsx'  # 預設名稱
        
        # 讀取所有工作表
        for sheet_name in xls.sheet_names:
            try:
                df = pd.read_excel(excel_file, sheet_name=sheet_name, header=None, engine='openpyxl')
                result['raw_data'][sheet_name] = df.to_dict('records')  # 轉換為可序列化的格式
                
                # 使用 V29 邏輯提取行簽名
                row_signatures = extract_row_signatures_v29(df, sheet_name, target_spots)
                
                # 從行簽名中推測可能的每日檔次
                for sig in row_signatures:
                    # 簡單推測：如果 unit_val 存在且合理，可能是每日檔次
                    if sig['unit_val'] and 1 <= sig['unit_val'] <= 1000:
                        interpretation = {
                            'sheet': sheet_name,
                            'row_idx': sig['row_idx'],
                            'col_idx': -1,  # 需要進一步分析
                            'date': '',  # 需要從上下文推測
                            'ad_name': '',  # 需要從上下文推測
                            'spots': sig['unit_val'],
                            'seconds': None,  # 需要進一步分析
                            'confidence': 'medium' if sig['bonus'] > 0 else 'low',
                            'rule_used': f"unit_val_extraction_v29",
                            'reason': f"Row {sig['row_idx']+1}: 發現重複數值 {sig['unit_val']} (出現 {sig['count']} 次), level={sig['level']}, bonus={sig['bonus']}",
                            'raw_value': sig['unit_val'],
                            'raw_row': sig['raw_row']
                        }
                        result['ai_interpretations'].append(interpretation)
                    
                    # 如果 sum 接近 target，也可能是候選
                    if target_spots and sig['sum'] > 0:
                        diff_ratio = abs(sig['sum'] - target_spots) / target_spots if target_spots > 0 else 1
                        if diff_ratio < 0.1:  # 誤差小於 10%
                            interpretation = {
                                'sheet': sheet_name,
                                'row_idx': sig['row_idx'],
                                'col_idx': -1,
                                'date': '',
                                'ad_name': '',
                                'spots': sig['sum'],
                                'seconds': None,
                                'confidence': 'high' if diff_ratio < 0.05 else 'medium',
                                'rule_used': f"sum_match_target_v29",
                                'reason': f"Row {sig['row_idx']+1}: 總和 {sig['sum']} 接近目標 {target_spots} (誤差 {diff_ratio*100:.1f}%)",
                                'raw_value': sig['sum'],
                                'raw_row': sig['raw_row']
                            }
                            result['ai_interpretations'].append(interpretation)
            
            except Exception as e:
                result['error'] = f"處理工作表 '{sheet_name}' 時發生錯誤: {str(e)}"
        
        excel_file.close()
        return result
        
    except Exception as e:
        result['error'] = f"讀取 Excel 檔案失敗: {str(e)}"
        return result

def parse_cue_excel_for_table1(file_content, order_info=None):
    """
    從 CUE Excel 檔案中解析出表1所需的結構化資料
    
    此函數會：
    1. 識別平台（全家新鮮視、全家廣播/企頻、家樂福、診所、其他）
    2. 提取秒數
    3. 提取地區
    4. 提取每日檔次，並識別檔次變化的日期（用於拆分）
    
    參數:
        file_content: Excel 檔案的 bytes 內容
        order_info: 訂單資訊（包含客戶、產品、業務等，可選）
    
    返回:
        list of dict: 每個 dict 代表一個需要拆分的「廣告單位」
        [
            {
                'platform': '全家新鮮視',  # 平台名稱
                'platform_category': '全家新鮮視',  # 平台分類（用於分組）
                'seconds': 15,  # 秒數
                'region': '全省',  # 地區
                'ad_name': '廣告名稱',  # 廣告名稱（可選）
                'daily_spots': [16, 16, 16, 16, 16, 16, 16],  # 每日檔次列表
                'dates': ['2026-01-31', '2026-02-01', ...],  # 對應的日期列表
                'start_date': '2026-01-31',  # 起始日
                'end_date': '2026-02-06',  # 終止日
                'total_spots': 112,  # 總檔次
                'source_sheet': '0131-0206',  # 來源工作表
                'source_row': 12,  # 來源行號
                'split_reason': 'daily_spots_change',  # 拆分原因
                'split_groups': [  # 如果檔次有變化，會分成多組
                    {
                        'start_date': '2026-01-31',
                        'end_date': '2026-02-03',
                        'daily_spots': 16,
                        'days': 4
                    },
                    {
                        'start_date': '2026-02-04',
                        'end_date': '2026-02-06',
                        'daily_spots': 12,
                        'days': 3
                    }
                ]
            }
        ]
    """
    result = []
    
    try:
        # 優先嘗試 Cue Sheet Pro (cueapp) 產生的三種格式：東吳、聲活、鉑霖
        result = parse_cueapp_excel(file_content)
        if result:
            if order_info:
                for ad_unit in result:
                    ad_unit.update({
                        'client': order_info.get('client', ''),
                        'product': order_info.get('product', ''),
                        'sales': order_info.get('sales', ''),
                        'company': order_info.get('company', ''),
                        'order_id': order_info.get('order_id', ''),
                        'amount_net': order_info.get('amount_net', 0),
                    })
            return result

        excel_file = io.BytesIO(file_content)
        excel_file.seek(0)
        xls = pd.ExcelFile(excel_file, engine='openpyxl')
        
        # 讀取所有工作表（非 cueapp 格式時沿用既有邏輯）
        for sheet_name in xls.sheet_names:
            try:
                df = pd.read_excel(excel_file, sheet_name=sheet_name, header=None, engine='openpyxl')
                
                # 移除完全為空的列
                df = df.loc[:, ~df.isna().all()]
                
                # 解析工作表名稱（通常包含日期範圍，例如：0131-0206）
                sheet_date_range = _parse_sheet_date_range(sheet_name)
                
                # 尋找平台資訊（通常在標題行或特定行）
                platform_info = _extract_platform_from_sheet(df, sheet_name)
                
                # 尋找秒數資訊
                seconds_info = _extract_seconds_from_sheet(df, sheet_name)
                
                # 尋找每日檔次資料行
                daily_spots_rows = _extract_daily_spots_rows(df, sheet_name, sheet_date_range)
                
                # 為每個每日檔次行建立廣告單位
                for spots_row in daily_spots_rows:
                    # 識別檔次變化，進行拆分
                    split_groups = _split_by_spots_change(
                        spots_row['daily_spots'],
                        spots_row['dates'],
                        spots_row.get('start_date'),
                        spots_row.get('end_date')
                    )
                    
                    # 為每個拆分組建立一個記錄
                    for group in split_groups:
                        ad_unit = {
                            'platform': platform_info.get('platform', '未知'),
                            'platform_category': platform_info.get('category', '其他'),
                            'seconds': seconds_info.get('seconds', 0),
                            'region': platform_info.get('region', '未知'),
                            'ad_name': spots_row.get('ad_name', ''),
                            'daily_spots': group['daily_spots_list'] if 'daily_spots_list' in group else [group['daily_spots']] * group['days'],
                            'dates': group['dates'],
                            'start_date': group['start_date'],
                            'end_date': group['end_date'],
                            'total_spots': sum(group['daily_spots_list']) if 'daily_spots_list' in group else group['daily_spots'] * group['days'],
                            'days': group['days'],
                            'source_sheet': sheet_name,
                            'source_row': spots_row.get('row_idx', -1),
                            'split_reason': group.get('split_reason', 'none'),
                            'split_groups': [group]  # 單一組
                        }
                        
                        # 合併訂單資訊（若有提供）
                        if order_info:
                            ad_unit.update({
                                'client': order_info.get('client', ''),
                                'product': order_info.get('product', ''),
                                'sales': order_info.get('sales', ''),
                                'company': order_info.get('company', ''),
                                'order_id': order_info.get('order_id', ''),
                                'amount_net': order_info.get('amount_net', 0)
                            })
                        
                        result.append(ad_unit)
            
            except Exception as e:
                print(f"處理工作表 '{sheet_name}' 時發生錯誤: {str(e)}")
                continue
        
        excel_file.close()
        return result
        
    except Exception as e:
        print(f"讀取 Excel 檔案失敗: {str(e)}")
        return result

def _parse_sheet_date_range(sheet_name):
    """
    從工作表名稱解析日期範圍
    例如：'0131-0206' → {'start': '2026-01-31', 'end': '2026-02-06'}
    """
    # 嘗試匹配日期範圍格式（例如：0131-0206, 01/31-02/06）
    patterns = [
        r'(\d{2})(\d{2})-(\d{2})(\d{2})',  # 0131-0206
        r'(\d{2})/(\d{2})-(\d{2})/(\d{2})',  # 01/31-02/06
    ]
    
    for pattern in patterns:
        m = re.search(pattern, sheet_name)
        if m:
            # 假設是當年度（可從訂單資訊取得）
            current_year = datetime.now().year
            if len(m.groups()) == 4:
                start_month = int(m.group(1))
                start_day = int(m.group(2))
                end_month = int(m.group(3))
                end_day = int(m.group(4))
                
                try:
                    start_date = datetime(current_year, start_month, start_day)
                    end_date = datetime(current_year, end_month, end_day)
                    return {
                        'start': start_date.strftime('%Y-%m-%d'),
                        'end': end_date.strftime('%Y-%m-%d')
                    }
                except:
                    pass
    
    return None

def _extract_platform_from_sheet(df, sheet_name):
    """
    從工作表內容中提取平台資訊
    返回: {'platform': '全家新鮮視', 'category': '全家新鮮視', 'region': '全省'}
    """
    platform_keywords = {
        '全家新鮮視': ['新鮮視', 'VISION', '全家便利商店店鋪'],
        '全家廣播': ['全家廣播', '企頻', 'RADIO', '企業頻道', '【全台全家共', '全家便利商店店鋪廣播'],
        '家樂福': ['家樂福', 'CARREFOUR', '量販通路', '量販店', '超市'],
        '診所': ['診所', 'CLINIC', '醫療', '醫院']
    }
    
    region_keywords = ['全省', '北北基', '中彰投', '桃竹苗', '高高屏', '雲嘉南', '宜花東']
    
    # 掃描前 30 行尋找平台資訊（增加掃描範圍）
    for idx in range(min(30, len(df))):
        row_text = ' '.join(df.iloc[idx].astype(str).tolist())
        row_text_upper = row_text.upper()
        
        # 檢查平台關鍵字（優先順序：全家新鮮視 > 全家廣播 > 家樂福 > 診所）
        # 注意：需要先檢查「全家廣播」，因為「全家」可能同時匹配「全家新鮮視」和「全家廣播」
        platform_found = None
        for platform in ['全家廣播', '全家新鮮視', '家樂福', '診所']:
            keywords = platform_keywords.get(platform, [])
            if any(kw in row_text_upper or kw in row_text for kw in keywords):
                platform_found = platform
                break
        
        if platform_found:
            # 檢查地區
            region = '全省'  # 預設
            for r in region_keywords:
                if r in row_text:
                    region = r
                    break
            
            return {
                'platform': platform_found,
                'category': platform_found,  # 用於分組
                'region': region
            }
    
    return {'platform': '未知', 'category': '其他', 'region': '未知'}

def _extract_seconds_from_sheet(df, sheet_name):
    """
    從工作表內容中提取秒數資訊
    返回: {'seconds': 15}
    """
    # 掃描前 20 行尋找秒數資訊
    for idx in range(min(20, len(df))):
        row_text = ' '.join(df.iloc[idx].astype(str).tolist())
        
        # 尋找 "15秒"、"15\"、"15秒廣告" 等格式
        patterns = [
            r'(\d+)\s*秒',
            r'(\d+)\s*"',
            r'廣告秒數[：:]\s*(\d+)',
            r'秒數[：:]\s*(\d+)'
        ]
        
        for pattern in patterns:
            m = re.search(pattern, row_text)
            if m:
                try:
                    seconds = int(m.group(1))
                    if 5 <= seconds <= 120:  # 合理的秒數範圍
                        return {'seconds': seconds}
                except:
                    pass
    
    return {'seconds': 0}

def _extract_daily_spots_rows(df, sheet_name, date_range=None):
    """
    從工作表中提取每日檔次資料行
    返回: [
        {
            'row_idx': 12,
            'ad_name': '15秒廣告',
            'daily_spots': [16, 16, 16, 16, 16, 16, 16],
            'dates': ['2026-01-31', '2026-02-01', ...],
            'start_date': '2026-01-31',
            'end_date': '2026-02-06'
        }
    ]
    """
    result = []
    
    # 尋找日期標題行（通常包含日期數字，例如：31, 1, 2, 3...）
    date_header_row_idx = None
    date_columns = []
    
    for idx in range(min(30, len(df))):
        row = df.iloc[idx]
        # 尋找包含連續數字的行（可能是日期）
        nums = []
        for col_idx, val in enumerate(row):
            try:
                num = int(float(val))
                if 1 <= num <= 31:  # 日期範圍
                    nums.append((col_idx, num))
            except:
                pass
        
        if len(nums) >= 5:  # 至少 5 個日期數字
            date_header_row_idx = idx
            date_columns = [col_idx for col_idx, _ in nums]
            break
    
    if date_header_row_idx is None or not date_columns:
        return result
    
    # 從日期標題行推斷日期
    dates = []
    if date_range:
        start_date = pd.to_datetime(date_range['start'])
        end_date = pd.to_datetime(date_range['end'])
        date_list = pd.date_range(start_date, end_date, freq='D')
        dates = [d.strftime('%Y-%m-%d') for d in date_list]
    else:
        # 如果沒有日期範圍，嘗試從標題行推斷
        # 這裡需要更複雜的邏輯，暫時跳過
        return result
    
    # 尋找檔次資料行（在日期標題行之後）
    for idx in range(date_header_row_idx + 1, min(date_header_row_idx + 50, len(df))):
        row = df.iloc[idx]
        
        # 提取對應日期欄位的檔次數值
        daily_spots = []
        for col_idx in date_columns[:len(dates)]:
            try:
                val = row.iloc[col_idx]
                if pd.notna(val):
                    spots = int(float(val))
                    if 0 <= spots <= 1000:  # 合理的檔次範圍
                        daily_spots.append(spots)
                    else:
                        daily_spots.append(0)
                else:
                    daily_spots.append(0)
            except:
                daily_spots.append(0)
        
        # 如果找到有效的檔次資料（至少 3 天有資料）
        if len([s for s in daily_spots if s > 0]) >= 3:
            # 提取廣告名稱（通常在檔次行的第一列）
            ad_name = ''
            try:
                first_col = str(row.iloc[0]).strip()
                if first_col and first_col != 'nan':
                    ad_name = first_col
            except:
                pass
            
            result.append({
                'row_idx': idx,
                'ad_name': ad_name,
                'daily_spots': daily_spots,
                'dates': dates[:len(daily_spots)],
                'start_date': dates[0] if dates else '',
                'end_date': dates[len(daily_spots)-1] if dates and len(daily_spots) > 0 else ''
            })
    
    return result

def _split_by_spots_change(daily_spots, dates, start_date=None, end_date=None):
    """
    根據每日檔次變化進行拆分
    例如：[44, 44, 44, 44, 40, 40, 40] 會拆成兩組
    
    返回: [
        {
            'start_date': '2026-01-31',
            'end_date': '2026-02-03',
            'daily_spots': 44,
            'daily_spots_list': [44, 44, 44, 44],
            'dates': ['2026-01-31', '2026-02-01', '2026-02-02', '2026-02-03'],
            'days': 4,
            'split_reason': 'daily_spots_change'
        },
        {
            'start_date': '2026-02-04',
            'end_date': '2026-02-06',
            'daily_spots': 40,
            'daily_spots_list': [40, 40, 40],
            'dates': ['2026-02-04', '2026-02-05', '2026-02-06'],
            'days': 3,
            'split_reason': 'daily_spots_change'
        }
    ]
    """
    if not daily_spots or not dates:
        return []
    
    groups = []
    current_group = {
        'daily_spots': daily_spots[0],
        'daily_spots_list': [daily_spots[0]],
        'dates': [dates[0]],
        'start_date': dates[0]
    }
    
    for i in range(1, len(daily_spots)):
        if daily_spots[i] != current_group['daily_spots']:
            # 檔次變化，結束當前組，開始新組
            current_group['end_date'] = dates[i-1]
            current_group['days'] = len(current_group['daily_spots_list'])
            current_group['split_reason'] = 'daily_spots_change'
            groups.append(current_group)
            
            # 開始新組
            current_group = {
                'daily_spots': daily_spots[i],
                'daily_spots_list': [daily_spots[i]],
                'dates': [dates[i]],
                'start_date': dates[i]
            }
        else:
            # 檔次相同，繼續當前組
            current_group['daily_spots_list'].append(daily_spots[i])
            current_group['dates'].append(dates[i])
    
    # 添加最後一組
    if current_group:
        current_group['end_date'] = dates[-1]
        current_group['days'] = len(current_group['daily_spots_list'])
        current_group['split_reason'] = 'daily_spots_change' if len(groups) > 0 else 'none'
        groups.append(current_group)
    
    return groups

def build_table1_from_cue_excel(cue_data_list, custom_settings=None):
    """
    從 CUE Excel 解析結果建立表1
    
    參數:
        cue_data_list: parse_cue_excel_for_table1() 的返回結果
        custom_settings: 自訂平台設定
    
    返回:
        DataFrame: 表1格式的資料
    """
    if not cue_data_list:
        return pd.DataFrame()
    
    result_rows = []
    
    for ad_unit in cue_data_list:
        # 計算店數
        platform_display = ad_unit.get('platform', '未知')
        try:
            p, ch, _ = parse_platform_region(platform_display)
            mp = get_media_platform_display(p, ch, platform_display)
        except Exception:
            mp = '其他'
        store_count = get_store_count(platform_display, custom_settings) if should_multiply_store_count(mp) else 1
        
        # 計算統計欄位
        daily_spots = ad_unit.get('daily_spots', [])
        days = ad_unit.get('days', len(daily_spots))
        total_spots = ad_unit.get('total_spots', sum(daily_spots))
        seconds = ad_unit.get('seconds', 0)
        total_seconds = total_spots * seconds
        total_store_seconds = total_seconds * store_count
        
        # 建立基本欄位
        base_row = {
            '業務': ad_unit.get('sales', ''),
            '主管': '',
            '合約編號': ad_unit.get('order_id', ''),
            '實收金額': int(ad_unit.get('amount_net', 0) or 0),
            '除佣實收': int(ad_unit.get('amount_net', 0) or 0),
            '製作成本': '',
            '獎金%': '',
            '核定獎金': '',
            '加發獎金': '',
            '業務基金': '',
            '協力基金': '',
            '秒數用途': '銷售秒數',
            '提交日': '',  # 需要從 orders 表取得
            'HYUNDAI_CUSTIN': ad_unit.get('client', ''),  # 客戶名稱
            '秒數': seconds,
            '素材': ad_unit.get('product', ''),
            '起始日': ad_unit.get('start_date', ''),
            '終止日': ad_unit.get('end_date', ''),
            '走期天數': days,
            '區域': ad_unit.get('region', '未知'),
            '平台': platform_display,
            '平台分類': ad_unit.get('platform_category', '其他'),
            '媒體平台': get_media_platform_display(*parse_platform_region(platform_display), platform_display),
        }
        
        # 每日24小時檔次分配（預留）
        hour_columns = [6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 18, 19, 20, 21, 22, 23, 0, 1]
        for hour in hour_columns:
            base_row[str(hour)] = ''
        
        # 統計欄位
        base_row['每天總檔次'] = daily_spots[0] if daily_spots else 0  # 使用第一天的檔次作為代表
        base_row['委刊總檔數'] = total_spots
        base_row['總秒數'] = total_seconds
        base_row['店數'] = store_count
        base_row['使用總秒數'] = total_store_seconds
        
        # 日期欄位（需要從所有廣告單位中收集所有日期）
        # 這裡先建立一個簡單版本，後續可以改進
        
        result_rows.append(base_row)
    
    df_table1 = pd.DataFrame(result_rows)
    
    # 處理日期欄位（需要收集所有日期）
    all_dates = set()
    for ad_unit in cue_data_list:
        dates = ad_unit.get('dates', [])
        all_dates.update([pd.to_datetime(d) for d in dates if d])
    
    if all_dates:
        sorted_dates = sorted(all_dates)
        weekday_map = {0: '一', 1: '二', 2: '三', 3: '四', 4: '五', 5: '六', 6: '日'}
        
        date_column_names = []
        for d in sorted_dates:
            weekday = weekday_map[d.weekday()]
            date_key = f"{d.month}/{d.day}({weekday})"
            if date_key not in date_column_names:
                date_column_names.append(date_key)
        
        # 初始化所有日期欄位
        for date_key in date_column_names:
            df_table1[date_key] = ''
        
        # 填入每日檔次
        for idx, ad_unit in enumerate(cue_data_list):
            dates = ad_unit.get('dates', [])
            daily_spots = ad_unit.get('daily_spots', [])
            
            for date_str, spots in zip(dates, daily_spots):
                try:
                    d = pd.to_datetime(date_str)
                    weekday = weekday_map[d.weekday()]
                    date_key = f"{d.month}/{d.day}({weekday})"
                    if date_key in df_table1.columns:
                        df_table1.loc[idx, date_key] = spots
                except:
                    pass
    
    return df_table1

# ==========================================
# 模擬資料產生（2026 年，供介面呈現）
# ==========================================

# 模擬用常數（對齊 Cue 表規格）
MOCK_REGIONS = ["北區", "桃竹苗", "中區", "雲嘉南", "高屏", "東區", "全省"]
MOCK_PLATFORM_RAW = [
    "新鮮視全省", "新鮮視北北基", "新鮮視中彰投", "新鮮視桃竹苗", "新鮮視雲嘉南", "新鮮視高高屏", "新鮮視宜花東",
    "企頻全省", "企頻北北基", "企頻中彰投", "企頻桃竹苗", "企頻雲嘉南", "企頻高高屏", "企頻宜花東",
    "全家廣播", "全家廣播北北基", "全家廣播中彰投", "全家廣播桃竹苗",
    "家樂福", "家樂福全省", "家樂福超市", "家樂福量販店",
]
MOCK_SECONDS = [5, 10, 15, 20, 25, 30, 35, 40, 45, 50, 55, 60]
MOCK_CLIENTS = ["統一企業", "富邦投信", "國泰人壽", "台灣大哥大", "遠傳電信", "中華電信", "可口可樂", "味全", "桂格", "黑松", "義美", "光泉", "大潤發", "全聯", "家樂福", "PChome", "momo", "玉山銀行", "中信金", "台新金"]
MOCK_PRODUCTS = ["春節檔期", "中秋促銷", "年貨大街", "週年慶", "品牌形象", "新品上市", "促銷活動", "認簽專案", "聯播方案", "區域方案", "30秒廣告", "15秒廣告", "10秒廣告"]
MOCK_SALES = ["王小明", "李小華", "張小美", "陳小傑", "林小芳", "黃小偉", "劉小玲"]
# 公司別：東吳、聲活、鉑霖（三家分公司）
MOCK_COMPANY = ["東吳", "聲活", "鉑霖"]
# 秒數用途類型（同一合約內可有多列、每列一種類型；未拆分的一列只會有一種）
SECONDS_USAGE_TYPES = ["銷售秒數", "交換秒數", "贈送秒數", "補檔秒數", "賀歲秒數", "公益秒數"]
# Google Sheet 或舊資料常見簡寫 → 正式類型（總結表／秒數用途分列用）
SECONDS_TYPE_ALIASES = {
    '銷售': '銷售秒數', '交換': '交換秒數', '贈送': '贈送秒數',
    '補檔': '補檔秒數', '賀歲': '賀歲秒數', '公益': '公益秒數',
}

def _normalize_seconds_type(val):
    """將秒數用途正規化為 SECONDS_USAGE_TYPES 其一，避免「銷售」等簡寫導致總結表銷售秒數為 0。"""
    if not val or (isinstance(val, float) and pd.isna(val)):
        return '銷售秒數'
    s = str(val).strip()
    if s in SECONDS_USAGE_TYPES:
        return s
    return SECONDS_TYPE_ALIASES.get(s, '銷售秒數')

def generate_mock_orders_2026(n=200):
    """
    產生 2026 年模擬訂單，模擬「一份合約因多平台/多區域/多秒數/多檔次而拆成多列」的真實情境。
    先產生若干「合約」，每份合約再拆成多筆訂單列（不同平台、區域、秒數、檔次、秒數用途），總筆數約 n。
    同一合約內可有多種秒數用途；未拆分的一列只會有一種類型。
    模擬「專案實收金額」（同一合約同值），實收金額不模擬（填 0）。
    回傳 list of tuples:
    (id, platform, client, product, sales, company, start_date, end_date, seconds, spots, amount_net=0, updated_at, contract_id, seconds_type, project_amount_net)
    """
    random.seed()
    orders = []
    base_date = datetime(2026, 1, 1)
    updated_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    # 合約數：約 n/8 份合約，每份合約 4~15 列（不同媒體/區域/秒數/檔次）
    n_contracts = max(15, min(40, n // 8))
    row_count = 0
    for c in range(n_contracts):
        contract_id = f"mock_{2026}_c{c+1:03d}"
        client = random.choice(MOCK_CLIENTS)
        product = random.choice(MOCK_PRODUCTS)
        sales = random.choice(MOCK_SALES)
        company = random.choice(MOCK_COMPANY)
        # 同一合約共用同一檔期（可略為錯開模擬不同檔次時段）
        start_offset = random.randint(0, 300)
        duration_days = random.randint(14, min(31, 365 - start_offset))
        start_dt = base_date + timedelta(days=start_offset)
        end_dt = start_dt + timedelta(days=duration_days - 1)
        if end_dt.year > 2026 or end_dt.month > 12:
            end_dt = datetime(2026, 12, 31)
        start_date = start_dt.strftime("%Y-%m-%d")
        end_date = end_dt.strftime("%Y-%m-%d")
        # 每份合約拆成多列（剩餘筆數少時可 1~3 列，避免 randint 空範圍）
        remaining = n - row_count
        if remaining <= 0:
            break
        low_rows = max(1, min(4, remaining))
        high_rows = min(15, remaining)
        n_rows = random.randint(low_rows, high_rows)
        # 專案實收金額：同一合約一筆總額，每列都填同一數字；實收金額不模擬（0）
        project_amount_net = max(120000, random.randint(120, 400) * 1000 * n_rows)
        for r in range(n_rows):
            if row_count >= n:
                break
            uid = f"mock_{2026}_c{c+1:03d}_{r+1:02d}"
            platform = random.choice(MOCK_PLATFORM_RAW)
            seconds = random.choice(MOCK_SECONDS)
            spots = random.randint(2, 36)
            if spots % 2 != 0:
                spots += 1
            seconds_type = random.choice(SECONDS_USAGE_TYPES)  # 同一合約內每列可不同類型
            orders.append((uid, platform, client, product, sales, company, start_date, end_date, seconds, spots, 0, updated_at, contract_id, seconds_type, project_amount_net))
            row_count += 1
        if row_count >= n:
            break
    return orders

def generate_mock_capacity_for_year(year=2026, target_usage_min=50, target_usage_max=120):
    """
    根據已產生的檔次段，計算每個媒體平台每個月的使用秒數，然後設定容量使使用率控制在 target_usage_min ~ target_usage_max %。
    回傳 (success: bool, message: str)
    """
    try:
        import calendar
        # 從 DB 讀取 segments 並展開為每日資料
        conn = get_db_connection()
        df_seg = pd.read_sql("SELECT * FROM ad_flight_segments WHERE media_platform IS NOT NULL", conn)
        conn.close()
        if df_seg.empty:
            return False, "尚無檔次段資料，請先產生模擬訂單"
        df_daily = explode_segments_to_daily(df_seg)
        if df_daily.empty or '媒體平台' not in df_daily.columns or '使用店秒' not in df_daily.columns or '日期' not in df_daily.columns:
            return False, "無法從檔次段展開每日資料"
        df_daily['日期'] = pd.to_datetime(df_daily['日期'], errors='coerce')
        df_daily = df_daily.dropna(subset=['日期'])
        df_daily['年'] = df_daily['日期'].dt.year
        df_daily['月'] = df_daily['日期'].dt.month
        df_y = df_daily[df_daily['年'] == year]
        if df_y.empty:
            return False, f"{year} 年尚無每日資料"
        # 計算每個媒體平台、每個月的使用秒數總和
        usage_by_media_month = df_y.groupby(['媒體平台', '月'])['使用店秒'].sum().reset_index()
        # 設定容量：目標使用率 50-120%，容量 = 使用秒數 / 目標使用率
        capacity_set = set()
        for _, row in usage_by_media_month.iterrows():
            mp = row['媒體平台']
            month = int(row['月'])
            used_sec = float(row['使用店秒'] or 0)
            if used_sec <= 0 or pd.isna(mp):
                continue
            target_rate = random.uniform(target_usage_min / 100, target_usage_max / 100)
            monthly_cap = used_sec / target_rate
            ndays = calendar.monthrange(year, month)[1]
            daily_cap = max(1, int(monthly_cap / ndays)) if ndays > 0 else max(1, int(monthly_cap / 30))
            set_platform_monthly_capacity(mp, year, month, daily_cap)
            capacity_set.add((mp, month))
        return True, f"已設定 {len(capacity_set)} 筆容量（{year} 年，使用率控制在 {target_usage_min}%-{target_usage_max}%）"
    except Exception as e:
        return False, f"產生容量設定失敗：{e}"


def load_mock_data_to_db(n=200):
    """
    產生 n 筆 2026 模擬資料並寫入 orders，同時建立 ad_flight_segments。
    模擬專案實收金額（同一合約同值），實收金額不模擬（0）；寫入後依比例自動計算拆分金額。
    回傳 (success: bool, message: str)
    """
    init_db()
    orders = generate_mock_orders_2026(n=n)
    conn = get_db_connection()
    c = conn.cursor()
    try:
        c.execute("BEGIN TRANSACTION")
        c.execute("DELETE FROM orders")
        # t = (id, platform, ..., amount_net=0, updated_at, contract_id, seconds_type, project_amount_net)；拆分金額先 NULL，稍後計算
        c.executemany("""
            INSERT OR REPLACE INTO orders
            (id, platform, client, product, sales, company, start_date, end_date, seconds, spots, amount_net, updated_at, contract_id, seconds_type, project_amount_net, split_amount)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
        """, [(*t, None) for t in orders])
        conn.commit()
        conn.close()
        conn_read = get_db_connection()
        df_orders = pd.read_sql("SELECT * FROM orders", conn_read)
        conn_read.close()
        custom_settings = load_platform_settings()
        build_ad_flight_segments(df_orders, custom_settings, write_to_db=True)
        # 依專案實收金額與使用秒數比例計算並寫回拆分金額
        contracts_with_project = df_orders.loc[df_orders['project_amount_net'].notna() & (pd.to_numeric(df_orders['project_amount_net'], errors='coerce') > 0), 'contract_id'].dropna().unique()
        for cid in contracts_with_project:
            if cid:
                _compute_and_save_split_amount_for_contract(str(cid))
        return True, f"已產生 {len(orders)} 筆 2026 年模擬資料（專案實收金額＋自動計算拆分金額）"
    except Exception as e:
        conn.rollback()
        conn.close()
        return False, str(e)


def load_mock_data_with_capacity_to_db(n=200, year=2026):
    """
    產生 n 筆模擬資料並寫入 orders，同時產生模擬容量設定和採購資料，使使用率控制在 0-120%。
    回傳 (success: bool, message: str)
    """
    init_db()
    # 先產生訂單和檔次段
    success1, msg1 = load_mock_data_to_db(n=n)
    if not success1:
        return False, f"產生訂單失敗：{msg1}"
    # 再產生容量設定（根據實際使用秒數，使使用率在 50-120%）
    success2, msg2 = generate_mock_capacity_for_year(year=year, target_usage_min=50, target_usage_max=120)
    if not success2:
        return False, f"產生容量設定失敗：{msg2}"
    # 產生模擬採購資料，但採購秒數要 >= 實際使用秒數，避免覆蓋容量後導致使用率破千
    success3, msg3 = generate_mock_platform_purchase_for_year_with_capacity_check(year)
    if not success3:
        return False, f"產生採購資料失敗：{msg3}"
    return True, f"{msg1}；{msg2}；{msg3}"


def _extract_google_sheet_id(url_or_id):
    """從 Google 試算表網址或直接 ID 取得 sheet ID。"""
    s = (url_or_id or "").strip()
    if not s:
        return None
    # 已是 ID（無 / 的長字串）
    if "/" not in s and len(s) > 20:
        return s
    m = re.search(r'/d/([a-zA-Z0-9_-]{40,})', s)
    return m.group(1) if m else None


def _fetch_google_sheet_as_dataframe(sheet_id, gid=0):
    """
    透過「匯出為 CSV」取得試算表內容（需試算表為「知道連結的使用者」可檢視）。
    回傳 (df, error_msg)。error_msg 為 None 表示成功。
    """
    url = f"https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=csv&gid={gid}"
    try:
        r = requests.get(url, timeout=30)
        r.raise_for_status()
        # 可能有多個 BOM / 編碼
        content = r.content
        if content.startswith(b'\xef\xbb\xbf'):
            content = content[3:]
        df_raw = pd.read_csv(io.BytesIO(content), encoding='utf-8', header=None, dtype=str)
    except Exception as e:
        return None, str(e)
    if df_raw.empty or len(df_raw) < 2:
        return None, "試算表為空或列數不足"
    # 找表頭列：第一列同時出現「平台」與「起始日」（或「終止日」）
    header_row = None
    for i in range(min(10, len(df_raw))):
        row_str = " ".join(df_raw.iloc[i].astype(str).fillna(""))
        if "平台" in row_str and ("起始日" in row_str or "終止日" in row_str):
            header_row = i
            break
    if header_row is None:
        return None, "找不到表1結構的表頭列（需含：平台、起始日/終止日）"
    df = pd.read_csv(io.BytesIO(content), encoding='utf-8', header=header_row, dtype=str)
    df = df.dropna(how='all', axis=1).dropna(how='all', axis=0)
    df.columns = [str(c).strip() for c in df.columns]
    return df, None


def _normalize_date(val):
    """將 2026/1/1、2026-01-01 等轉成 YYYY-MM-DD。"""
    if pd.isna(val) or val == '' or val == 'nan':
        return ''
    val = str(val).strip()
    if not val:
        return ''
    try:
        dt = pd.to_datetime(val, errors='coerce')
        if pd.isna(dt):
            return ''
        return dt.strftime('%Y-%m-%d')
    except Exception:
        return ''


def _sheet_row_to_order(row, row_index, col_map):
    """
    將表1結構的一列轉成 orders 一筆 (id, platform, client, product, sales, company,
    start_date, end_date, seconds, spots, amount_net, updated_at, contract_id, seconds_type)。
    col_map: { 'platform': '平台', 'client': 'HYUNDAI_CUSTIN', ... }
    """
    def get(k, default=''):
        key = col_map.get(k, k)  # 無對照時用 k 當欄位名
        if key not in row.index:
            return default
        v = row.get(key, default)
        return '' if pd.isna(v) or v == 'nan' else str(v).strip()

    platform = get('platform') or get('平台')
    if not platform:
        return None
    start_date = _normalize_date(get('start_date') or get('起始日'))
    end_date = _normalize_date(get('end_date') or get('終止日'))
    if not start_date or not end_date:
        return None
    try:
        seconds = int(float(get('seconds') or get('秒數') or 0))
    except (ValueError, TypeError):
        seconds = 0
    try:
        spots = int(float(get('spots') or get('每天總檔次') or get('委刊總檔數') or get('委刋總檔數') or 0))
    except (ValueError, TypeError):
        spots = 0
    try:
        amount_net = float(get('amount_net') or get('實收金額') or 0)
    except (ValueError, TypeError):
        amount_net = 0
    client = get('client') or get('HYUNDAI_CUSTIN') or get('客戶')
    product = get('product') or get('素材')
    sales = get('sales') or get('業務')
    company = get('company') or get('公司')
    contract_id = get('contract_id') or get('合約編號')
    seconds_type = _normalize_seconds_type(get('seconds_type') or get('秒數用途') or '銷售秒數')
    try:
        project_amount_net = float(get('project_amount_net') or get('專案實收金額') or 0)
    except (ValueError, TypeError):
        project_amount_net = 0
    if project_amount_net <= 0:
        project_amount_net = None
    updated_at = get('updated_at') or get('提交日') or datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    if not updated_at or updated_at == '':
        updated_at = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    else:
        updated_at = _normalize_date(updated_at)
        if not updated_at:
            updated_at = datetime.now().strftime('%Y-%m-%d')
        updated_at = updated_at + " 00:00:00" if len(updated_at) == 10 else updated_at
    order_id = f"gs_{row_index}_{contract_id or row_index}_{platform}_{start_date}".replace(" ", "_")[:200]
    return (order_id, platform, client or '', product or '', sales or '', company or '',
            start_date, end_date, seconds, spots, amount_net, updated_at, contract_id or None, seconds_type or '銷售秒數', project_amount_net)


def import_google_sheet_to_orders(url_or_id, replace_existing=True):
    """
    從 Google 試算表（表1結構）匯入至 orders，並建立 ad_flight_segments。
    url_or_id: 試算表完整網址或 Sheet ID。
    replace_existing: True 則先清空 orders 再匯入；False 則追加。
    回傳 (success: bool, message: str)
    """
    sheet_id = _extract_google_sheet_id(url_or_id)
    if not sheet_id:
        return False, "請輸入有效的 Google 試算表網址或 ID"
    df, err = _fetch_google_sheet_as_dataframe(sheet_id)
    if err:
        return False, f"無法讀取試算表：{err}"
    # 欄位對照：表頭可能為中文（專案實收金額：同合約填同一數字，匯入後系統會依比例計算拆分金額）
    col_map = {
        'platform': '平台', 'company': '公司', 'sales': '業務', 'contract_id': '合約編號',
        'client': 'HYUNDAI_CUSTIN', 'product': '素材', 'start_date': '起始日', 'end_date': '終止日',
        'seconds': '秒數', 'spots': '每天總檔次', 'amount_net': '實收金額', 'seconds_type': '秒數用途',
        'updated_at': '提交日', '客戶': 'HYUNDAI_CUSTIN', '委刊總檔數': '委刊總檔數', '委刋總檔數': '委刋總檔數',
        'project_amount_net': '專案實收金額', '專案實收金額': '專案實收金額',
    }
    orders = []
    for i, (_, row) in enumerate(df.iterrows()):
        t = _sheet_row_to_order(row, i, col_map)
        if t is not None:
            orders.append(t)
    if not orders:
        return False, "沒有可匯入的資料列（需有平台、起始日、終止日且為有效日期）"
    init_db()
    conn = get_db_connection()
    c = conn.cursor()
    try:
        if replace_existing:
            c.execute("DELETE FROM orders")
        for t in orders:
            # t = (id, platform, ..., contract_id, seconds_type, project_amount_net)；拆分金額先 NULL，匯入後依專案實收計算
            project_val = t[14] if len(t) > 14 else None
            c.execute("""
                INSERT OR REPLACE INTO orders
                (id, platform, client, product, sales, company, start_date, end_date, seconds, spots, amount_net, updated_at, contract_id, seconds_type, project_amount_net, split_amount)
                VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
            """, (*t[:14], project_val, None))
        conn.commit()
        conn.close()
        conn_read = get_db_connection()
        df_orders = pd.read_sql("SELECT * FROM orders", conn_read)
        conn_read.close()
        custom_settings = load_platform_settings()
        build_ad_flight_segments(df_orders, custom_settings, write_to_db=True)
        # 有填專案實收金額的合約：依使用秒數比例計算並寫回拆分金額
        contracts_with_project = df_orders.loc[df_orders['project_amount_net'].notna() & (pd.to_numeric(df_orders['project_amount_net'], errors='coerce') > 0), 'contract_id'].dropna().unique()
        for cid in contracts_with_project:
            if cid:
                _compute_and_save_split_amount_for_contract(str(cid))
        return True, f"已匯入 {len(orders)} 筆（表1結構）；若有專案實收金額已自動計算拆分金額）"
    except Exception as e:
        conn.rollback()
        conn.close()
        return False, str(e)


def calculate_inventory(df_orders=None, custom_settings=None, use_segments=True):
    """
    核心運算引擎：將檔次段展開為「每日庫存」
    對應 Excel 表2 與 表3 的邏輯
    
    參數:
        df_orders: 訂單 DataFrame（如果 use_segments=False 時使用）
        custom_settings: 自訂平台設定
        use_segments: 是否從 segments 表計算（預設 True，推薦）
    """
    if use_segments:
        # 從 segments 表讀取（正確做法）
        conn = get_db_connection()
        try:
            df_segments = pd.read_sql("SELECT * FROM ad_flight_segments", conn)
            conn.close()
            
            if df_segments.empty:
                return pd.DataFrame()
            
            return explode_segments_to_daily(df_segments)
        except Exception as e:
            conn.close()
            # 如果 segments 表不存在或為空，回退到舊方法
            if df_orders is not None:
                return calculate_inventory(df_orders, custom_settings, use_segments=False)
            return pd.DataFrame()
    else:
        # 舊方法：直接從 orders 計算（向後相容）
        if df_orders is None or df_orders.empty:
            return pd.DataFrame()
            
        daily_records = []
        
        for _, row in df_orders.iterrows():
            try:
                # 跳過無效日期
                if pd.isna(row['start_date']) or pd.isna(row['end_date']):
                    continue
                if row['start_date'] == '' or row['end_date'] == '':
                    continue
                    
                s_date = pd.to_datetime(row['start_date'], errors='coerce')
                e_date = pd.to_datetime(row['end_date'], errors='coerce')
                
                if pd.isna(s_date) or pd.isna(e_date):
                    continue
                
                # 依媒體平台決定是否乘店數（避免家樂福/診所等被乘上店數）
                p, ch, _ = parse_platform_region(row.get('platform'))
                mp = get_media_platform_display(p, ch, row.get('platform', ''))
                store_count = get_store_count(row['platform'], custom_settings) if should_multiply_store_count(mp) else 1
                
                # 計算每日消耗：全家類為「使用店秒」，其餘為「使用秒數」（仍沿用欄名使用店秒）
                daily_usage_store_seconds = row['seconds'] * row['spots'] * store_count
                daily_usage_raw_seconds = row['seconds'] * row['spots']
                
                # 裂解日期 (Explode)
                date_range = pd.date_range(s_date, e_date, inclusive='both')
                for d in date_range:
                    daily_records.append({
                        '日期': d,
                        '平台': row['platform'],
                        '公司': row['company'],
                        '業務': row['sales'],
                        '客戶': row['client'],
                        '產品': row['product'],
                        '使用店秒': daily_usage_store_seconds,
                        '原始秒數': daily_usage_raw_seconds,
                        '秒數': row['seconds'],
                        '檔次': row['spots'],
                        '訂單ID': row['id']
                    })
            except Exception as e:
                continue
                
        return pd.DataFrame(daily_records)

def build_ad_flight_segments(df_orders, custom_settings=None, write_to_db=True):
    """
    建立檔次段（Flight Segments）核心事實表
    
    規則：
    1. 只保留三家平台：全家廣播（企頻）、全家新鮮視、家樂福
    2. 同一訂單（source_order_id），如果以下任一項不同，就拆成不同 segment：
       - 不同秒數
       - 不同區域（region）
       - 不同檔次（spots）
       - 日期不連續（間隔超過1天）
    
    參數:
        df_orders: 訂單 DataFrame
        custom_settings: 自訂平台設定
        write_to_db: 是否寫入資料庫（預設 True）
    
    返回:
        DataFrame: 檔次段資料
    """
    if df_orders.empty:
        return pd.DataFrame()
    
    segments = []
    
    # 按照訂單 ID 分組（每個訂單獨立處理）
    for order_id, group in df_orders.groupby('id'):
        # 按日期排序
        group = group.sort_values('start_date')
        
        for _, row in group.iterrows():
            try:
                # 跳過無效資料
                if pd.isna(row['seconds']) or row['seconds'] <= 0:
                    continue
                if pd.isna(row['spots']) or row['spots'] <= 0:
                    continue
                
                # 解析平台/頻道/區域，並取得表一用「媒體平台」顯示名稱
                platform, channel, region = parse_platform_region(row['platform'])
                media_platform = get_media_platform_display(platform, channel, row.get('platform', ''))
                
                # 只保留全家和家樂福
                if platform not in ['全家', '家樂福']:
                    continue
                
                # 處理日期
                s_date = pd.to_datetime(row['start_date'], errors='coerce')
                e_date = pd.to_datetime(row['end_date'], errors='coerce')
                
                if pd.isna(s_date) or pd.isna(e_date):
                    continue
                
                # 計算相關數值（僅全家廣播/新鮮視乘店數；其餘媒體不乘店數）
                store_count = get_store_count(row['platform'], custom_settings) if should_multiply_store_count(media_platform) else 1
                days = (e_date - s_date).days + 1
                total_spots = row['spots'] * days
                total_store_seconds = row['seconds'] * total_spots * store_count
                
                # 建立 segment
                segment_id = str(uuid.uuid4())
                segments.append({
                    'segment_id': segment_id,
                    'source_order_id': order_id,
                    'platform': platform,
                    'channel': channel,
                    'region': region,
                    'media_platform': media_platform,
                    'company': row.get('company', ''),
                    'sales': row.get('sales', ''),
                    'client': row.get('client', ''),
                    'product': row.get('product', ''),
                    'seconds': int(row['seconds']),
                    'spots': int(row['spots']),
                    'start_date': s_date.date(),
                    'end_date': e_date.date(),
                    'duration_days': days,
                    'store_count': store_count,
                    'total_spots': total_spots,
                    'total_store_seconds': total_store_seconds,
                    'seconds_type': _normalize_seconds_type(row.get('seconds_type')),
                    'created_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                    'updated_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                })
            except Exception as e:
                continue
    
    df_segments = pd.DataFrame(segments)
    
    # 寫入資料庫
    if write_to_db and not df_segments.empty:
        conn = get_db_connection()
        c = conn.cursor()
        try:
            c.execute('BEGIN TRANSACTION')
            # 清空舊資料（全量更新）
            c.execute('DELETE FROM ad_flight_segments')
            # 寫入新資料
            c.executemany('''
                INSERT INTO ad_flight_segments
                (segment_id, source_order_id, platform, channel, region, media_platform, company, sales,
                 client, product, seconds, spots, start_date, end_date, duration_days,
                 store_count, total_spots, total_store_seconds, seconds_type, created_at, updated_at)
                VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
            ''', [
                (
                    seg['segment_id'], seg['source_order_id'], seg['platform'], seg['channel'],
                    seg['region'], seg.get('media_platform', ''), seg['company'], seg['sales'], seg['client'], seg['product'],
                    seg['seconds'], seg['spots'], seg['start_date'], seg['end_date'],
                    seg['duration_days'], seg['store_count'], seg['total_spots'],
                    seg['total_store_seconds'], seg['seconds_type'], seg['created_at'], seg['updated_at']
                )
                for seg in segments
            ])
            conn.commit()
            conn.close()
        except Exception as e:
            conn.rollback()
            conn.close()
            print(f"寫入 segments 失敗: {e}")
    
    return df_segments

def _resolve_media_platform_for_daily(seg):
    """從 segment 取得媒體平台，供 df_daily 使用；DB 讀回時 media_platform 可能為 NaN/空，改由 platform/channel 推算。"""
    mp = seg.get('media_platform')
    if mp is not None and str(mp).strip() and not (isinstance(mp, float) and pd.isna(mp)):
        return str(mp).strip()
    return get_media_platform_display(
        seg.get('platform') if pd.notna(seg.get('platform')) else '',
        seg.get('channel') if pd.notna(seg.get('channel')) else '',
        ''
    )

def explode_segments_to_daily(df_segments):
    """
    將檔次段展開為每日資料
    這是 daily_inventory 的正確來源
    """
    daily_records = []
    
    for _, seg in df_segments.iterrows():
        try:
            s_date = pd.to_datetime(seg['start_date'])
            e_date = pd.to_datetime(seg['end_date'])
            
            # 展開日期
            date_range = pd.date_range(s_date, e_date, inclusive='both')
            for d in date_range:
                daily_store_seconds = seg['seconds'] * seg['spots'] * seg['store_count']
                
                # 組合平台顯示名稱（向後相容）
                platform_display = f"{seg['platform']}-{seg['channel']}"
                if seg['region'] != '未知':
                    platform_display = f"{seg['platform']}-{seg['channel']}-{seg['region']}"
                
                daily_records.append({
                    '日期': d,
                    '平台': platform_display,  # 向後相容，使用組合名稱
                    '媒體平台': _resolve_media_platform_for_daily(seg),
                    '秒數用途': _normalize_seconds_type(seg.get('seconds_type')),  # 年度總結表依類型分列用
                    '公司': seg['company'],
                    '業務': seg['sales'],
                    '客戶': seg['client'],
                    '產品': seg['product'],
                    '使用店秒': daily_store_seconds,
                    '原始秒數': seg['seconds'] * seg['spots'],
                    '秒數': seg['seconds'],
                    '檔次': seg['spots'],
                    'segment_id': seg['segment_id'],
                    '訂單ID': seg['source_order_id']
                })
        except Exception as e:
            continue
    
    return pd.DataFrame(daily_records)

def _segment_platform_display(seg):
    """表2 明細用平台顯示名稱（對齊 Excel：新鮮視北北基、企頻北北基、家樂福超市等；不顯示「全家廣播(企頻)」）"""
    platform = seg.get('platform', '')
    channel = seg.get('channel', '')
    region = seg.get('region', '未知')
    media = seg.get('media_platform', '')
    if platform == '全家':
        if region and region != '未知':
            return (channel or '全家') + region
        # 企頻／新鮮視：無區域時顯示頻道名，不顯示「全家廣播(企頻)」
        if channel == '企頻':
            return '企頻'
        if channel == '新鮮視':
            return '新鮮視'
        return media or (channel or '全家')
    if platform == '家樂福':
        return media or '家樂福'
    return media or f"{platform}-{channel}"

def build_table2_summary_by_company(df_segments, df_daily, df_orders, media_platform=None):
    """
    表2 區塊一：依公司統計總覽
    欄位：公司、實收金額、除佣實收、委刊總檔數、使用總秒數、1/1(四)…每日使用店秒、小計列
    media_platform: 若指定（如 '全家新鮮視'、'全家廣播(企頻)'）則只統計該媒體；None 表示全部。
    """
    if df_segments.empty or df_daily.empty:
        return pd.DataFrame()
    # 依媒體篩選（與表3 一致：media_platform 空時用 platform/channel 推算）
    def _resolve_mp(r):
        return r.get('media_platform') or get_media_platform_display(r.get('platform'), r.get('channel'), r.get('platform', ''))
    if media_platform:
        df_segments = df_segments[df_segments.apply(_resolve_mp, axis=1) == media_platform].copy()
        if '媒體平台' in df_daily.columns:
            df_daily = df_daily[df_daily['媒體平台'] == media_platform].copy()
        if df_segments.empty or df_daily.empty:
            return pd.DataFrame()
    companies = df_segments['company'].dropna().unique()
    companies = [c for c in companies if c]
    if not companies:
        return pd.DataFrame()
    # 合約編號對應金額（每合約只算一次）
    try:
        df_ord = df_orders[['id', 'amount_net', 'contract_id']].drop_duplicates(subset=['id'])
    except Exception:
        df_ord = df_orders[['id', 'amount_net']].copy().drop_duplicates(subset=['id'])
        df_ord['contract_id'] = None
    seg_ord = df_segments[['source_order_id', 'company', 'total_spots', 'total_store_seconds']].merge(
        df_ord, left_on='source_order_id', right_on='id', how='left')
    # 實收金額：依合約編號去重後加總（有 contract_id 用 contract_id 去重，否則用 source_order_id）
    seg_ord['_contract_key'] = seg_ord.get('contract_id').fillna(seg_ord['source_order_id'])
    by_company = seg_ord.groupby('company').agg(
        total_spots=('total_spots', 'sum'),
        total_store_seconds=('total_store_seconds', 'sum'),
    ).reset_index()
    # 每公司實收：依合約編號去重後加總（同一合約只算一次）
    def _sum_amt_unique_contract(g):
        return g.drop_duplicates('_contract_key')['amount_net'].sum()
    amt_by_co = seg_ord.groupby('company').apply(_sum_amt_unique_contract).reindex(companies).fillna(0)
    by_company['實收金額'] = by_company['company'].map(amt_by_co).fillna(0).astype(int)
    by_company['除佣實收'] = by_company['實收金額']
    by_company['委刊總檔數'] = by_company['total_spots'].fillna(0).astype(int)
    by_company['使用總秒數'] = by_company['total_store_seconds'].fillna(0).astype(int)
    # 每日使用店秒
    if '日期' not in df_daily.columns or '使用店秒' not in df_daily.columns or '公司' not in df_daily.columns:
        date_cols = []
    else:
        daily_agg = df_daily.groupby(['公司', '日期'])['使用店秒'].sum().reset_index()
        daily_agg['日期'] = pd.to_datetime(daily_agg['日期'])
        all_dates = sorted(daily_agg['日期'].dropna().unique())
        weekday_map = {0: '一', 1: '二', 2: '三', 3: '四', 4: '五', 5: '六', 6: '日'}
        date_cols = [f"{d.month}/{d.day}({weekday_map[d.weekday()]})" for d in all_dates]
        pivot_daily = daily_agg.pivot(index='公司', columns='日期', values='使用店秒').reindex(companies).fillna(0)
        for d in all_dates:
            key = f"{d.month}/{d.day}({weekday_map[d.weekday()]})"
            if d in pivot_daily.columns:
                by_company[key] = pivot_daily.loc[by_company['company'], d].fillna(0).astype(int).values
            else:
                by_company[key] = 0
    base_cols = ['公司', '實收金額', '除佣實收', '委刊總檔數', '使用總秒數']
    out = by_company[['company', '實收金額', '除佣實收', '委刊總檔數', '使用總秒數']].copy()
    out.columns = base_cols
    for c in date_cols:
        if c in by_company.columns:
            out[c] = by_company[c].fillna(0).astype(int)
    # 小計列
    subtotal = {'公司': '小計', '實收金額': out['實收金額'].sum(), '除佣實收': out['除佣實收'].sum(),
                '委刊總檔數': out['委刊總檔數'].sum(), '使用總秒數': out['使用總秒數'].sum()}
    for c in date_cols:
        subtotal[c] = out[c].sum() if c in out.columns else 0
    out = pd.concat([out, pd.DataFrame([subtotal])], ignore_index=True)
    return out

def build_table2_details_by_company(df_segments, df_daily, df_orders):
    """
    表2 區塊二/三：依業務統計明細（依公司分組）
    每公司一表：平台、公司、業務、合約編號、實收金額、除佣實收、提交日、客戶名稱、秒數、委刊總檔數、使用總秒數、每日使用店秒、小計列
    """
    if df_segments.empty:
        return {}
    try:
        df_ord = df_orders[['id', 'amount_net', 'updated_at', 'contract_id']].drop_duplicates(subset=['id'])
    except Exception:
        df_ord = df_orders[['id', 'amount_net', 'updated_at']].drop_duplicates(subset=['id'])
        df_ord['contract_id'] = None
    df_ord['提交日'] = pd.to_datetime(df_ord['updated_at'], errors='coerce').dt.strftime('%Y/%m/%d')
    seg = df_segments.merge(df_ord, left_on='source_order_id', right_on='id', how='left')
    seg['合約編號'] = seg.get('contract_id').fillna(seg['source_order_id'])
    seg['平台顯示'] = seg.apply(_segment_platform_display, axis=1)
    # 每日使用店秒（依 segment_id）
    result = {}
    daily_pivot = pd.DataFrame()
    if not df_daily.empty and 'segment_id' in df_daily.columns and '日期' in df_daily.columns:
        _piv = df_daily.groupby(['segment_id', '日期'])['使用店秒'].sum().unstack(fill_value=0)
        weekday_map = {0: '一', 1: '二', 2: '三', 3: '四', 4: '五', 5: '六', 6: '日'}
        _piv.columns = [f"{c.month}/{c.day}({weekday_map.get(c.weekday(), '')})" if hasattr(c, 'month') else str(c) for c in _piv.columns]
        daily_pivot = _piv
    for company in seg['company'].dropna().unique():
        if not company:
            continue
        s = seg[seg['company'] == company].copy()
        s = s.rename(columns={'client': '客戶名稱', 'total_spots': '委刊總檔數', 'total_store_seconds': '使用總秒數'})
        detail = pd.DataFrame({
            '公司': s['company'].values,
            '平台': s['平台顯示'].values,
            '業務': s['sales'].values,
            '合約編號': s['合約編號'].astype(str).values,
            '實收金額': s['amount_net'].fillna(0).astype(int).values,
            '除佣實收': s['amount_net'].fillna(0).astype(int).values,
            '提交日': s['提交日'].fillna('').values,
            '客戶名稱': s['客戶名稱'].fillna('').values,
            '秒數': s['seconds'].fillna(0).astype(int).values,
            '委刊總檔數': s['委刊總檔數'].fillna(0).astype(int).values,
            '使用總秒數': s['使用總秒數'].fillna(0).astype(int).values,
        })
        # 每日使用店秒：依 segment_id 對應
        if not daily_pivot.empty and 'segment_id' in s.columns:
            for col in daily_pivot.columns:
                detail[col] = 0
            for i, seg_id in enumerate(s['segment_id'].values):
                if seg_id in daily_pivot.index:
                    row_vals = daily_pivot.loc[seg_id]
                    for col in daily_pivot.columns:
                        if col in detail.columns:
                            detail.iloc[i, detail.columns.get_loc(col)] = int(row_vals.get(col, 0))
        # 小計列
        sub = {'公司': company, '平台': '', '業務': '', '合約編號': '小計', '實收金額': detail['實收金額'].sum(),
               '除佣實收': detail['除佣實收'].sum(), '提交日': '', '客戶名稱': '', '秒數': '', '委刊總檔數': detail['委刊總檔數'].sum(), '使用總秒數': detail['使用總秒數'].sum()}
        if not daily_pivot.empty:
            for col in daily_pivot.columns:
                sub[col] = detail[col].sum() if col in detail.columns else 0
        detail = pd.concat([detail, pd.DataFrame([sub])], ignore_index=True)
        result[company] = detail
    return result

def build_table3_monthly_control(df_daily, df_segments, custom_settings=None, year=None, month=None, monthly_capacity=None):
    """
    表3 每月秒數控管表：依媒體平台區分，含執行秒、可用秒數、使用率、可排日（顏色）
    year, month: 若指定則只顯示該年該月；monthly_capacity: dict media_platform -> 當月每日可用秒數（向媒體購買），有則可用秒數用此值。
    回傳 dict: media_platform -> DataFrame（4 列：執行秒、可用秒數、使用率、可排日）
    """
    if df_daily.empty or df_segments.empty:
        return {}
    if '媒體平台' not in df_daily.columns:
        return {}
    daily_hours_default = 18
    monthly_capacity = monthly_capacity or {}
    df_daily = df_daily.copy()
    df_daily['日期'] = pd.to_datetime(df_daily['日期'], errors='coerce')
    df_segments = df_segments.copy()
    df_segments['start_date'] = pd.to_datetime(df_segments['start_date'], errors='coerce')
    df_segments['end_date'] = pd.to_datetime(df_segments['end_date'], errors='coerce')
    all_dates = sorted(df_daily['日期'].dropna().unique())
    # 篩選指定年月
    if year is not None and month is not None:
        y, m = int(year), int(month)
        all_dates = [d for d in all_dates if d.year == y and d.month == m]
        if not all_dates:
            import calendar
            ndays = calendar.monthrange(y, m)[1]
            all_dates = [pd.Timestamp(year=y, month=m, day=day) for day in range(1, ndays + 1)]
    if not all_dates:
        return {}
    media_platforms = [p for p in MEDIA_PLATFORM_OPTIONS if p in df_daily['媒體平台'].unique()]
    if not media_platforms:
        media_platforms = df_daily['媒體平台'].dropna().unique().tolist()
    result = {}
    def _resolve_media_platform(r):
        """與 explode_segments_to_daily 一致：media_platform 空值時用 platform/channel 推算"""
        return r.get('media_platform') or get_media_platform_display(r.get('platform'), r.get('channel'), r.get('platform', ''))
    for mp in media_platforms:
        dd = df_daily[df_daily['媒體平台'] == mp]
        # 篩選該媒體的 segments（與 df_daily 的 媒體平台 推算方式一致，避免 DB 中 media_platform 為空時漏掉企頻）
        seg_mp = df_segments[df_segments.apply(lambda r: _resolve_media_platform(r) == mp, axis=1)].copy()
        # 每日使用秒數（依媒體平台彙總）
        used_by_date = dd.groupby('日期')['使用店秒'].sum().reindex(all_dates).fillna(0)
        # 每日可用秒數：若有設定「當月每日可用秒數」則用該值，否則用 segment 產能（向量化，避免 O(dates×segments) 雙層迴圈）
        set_cap = monthly_capacity.get(mp)
        if set_cap is not None and set_cap > 0:
            cap_series = pd.Series([int(set_cap)] * len(all_dates), index=all_dates)
        else:
            starts = np.array(seg_mp['start_date'].values, dtype=np.datetime64)
            ends = np.array(seg_mp['end_date'].values, dtype=np.datetime64)
            scs = (seg_mp['store_count'].fillna(0).astype(int) * daily_hours_default * 3600).values
            cap_list = []
            for d in all_dates:
                d64 = np.datetime64(d)
                mask = (starts <= d64) & (d64 <= ends)
                cap_list.append(np.sum(scs[mask]))
            cap_series = pd.Series(cap_list, index=all_dates)
        used_by_date = used_by_date.reindex(all_dates).fillna(0)
        # 使用率 %（可用秒數為 0 時為 0）
        util_series = (used_by_date / cap_series.replace(0, 1)).fillna(0) * 100
        # 建四列：執行秒、可用秒數、使用率、可排日；日期欄以 月/日(星期) 顯示；% 欄小數一位
        weekday_cn = ['一', '二', '三', '四', '五', '六', '日']
        date_cols = [f"{d.month}/{d.day}({weekday_cn[d.weekday()]})" for d in all_dates]
        total_used = used_by_date.sum()
        total_cap = cap_series.sum()
        pct_used = round(total_used / (total_cap or 1) * 100, 1)
        pct_unused = round((total_cap - total_used) / (total_cap or 1) * 100, 1)
        row_used = {'授權': '總經理', '項目': '執行秒', '秒數': int(total_used), '%': f"{pct_used:.1f}"}
        row_cap = {'授權': '總經理', '項目': '可用秒數', '秒數': int(total_cap), '%': f"{pct_unused:.1f}"}
        row_util = {'授權': '總經理', '項目': '使用率', '秒數': '', '%': '100.0'}
        row_color = {'授權': '業務', '項目': '可排日', '秒數': '', '%': ''}
        for i in range(len(all_dates)):
            d = all_dates[i]
            row_used[date_cols[i]] = int(used_by_date.iloc[i]) if i < len(used_by_date) else 0
            row_cap[date_cols[i]] = int(cap_series.iloc[i]) if i < len(cap_series) else 0
            u = util_series.iloc[i] if i < len(util_series) else 0
            row_util[date_cols[i]] = f"{round(float(u), 1)}%" if pd.notna(u) else "0%"
            row_color[date_cols[i]] = float(u) if pd.notna(u) else 0
        result[mp] = pd.DataFrame([row_used, row_cap, row_util, row_color])
    return result


# 年度使用秒數總表：實體對應（企頻、新鮮視、家樂福、診所）
ANNUAL_SUMMARY_ENTITY_LABELS = ['企頻', '新鮮視', '家樂福', '診所']
ANNUAL_SUMMARY_MEDIA_MAP = {
    '企頻': ['全家廣播(企頻)'],
    '新鮮視': ['全家新鮮視'],
    '家樂福': ['家樂福超市', '家樂福量販店'],
    '診所': [],  # 無對應平台則顯示 0
}

# ========== 實驗分頁：依時間的庫存警示與分析 ==========
# 【核心假設】當月秒數若未使用於月底結算視為 100% 浪費（不可逆）；秒數價值隨接近月底而衰減；目標為最小化月底浪費；爆量仍監控但屬次要。
EMERGENCY_DAYS = 7  # T0 緊急期天數（today ~ today+N 為唯一可補救窗口）
TIME_WEIGHT = {"past": 1.0, "emergency": 0.9, "buffer": 0.3}
TARGET_USAGE = 0.8
TOLERANCE = 0.2
SAFE_LIMIT = 0.95
OVER_BUFFER = 0.1
# 「約 X 檔全省 15 秒」換算：1 檔 = 15 秒 × 全省店數（店秒）
SECONDS_PER_SPOT_15S = 15 * 4200  # 全省約 4200 店


def build_daily_inventory_and_metrics(df_daily, year, month, today, emergency_days=EMERGENCY_DAYS, monthly_capacity_loader=None, media_platform=None):
    """
    建構日粒度事實表與戰略指標。回傳 (daily_inventory DataFrame, metrics dict)。
    daily_inventory 欄位: date, total_capacity_seconds, used_seconds, unused_seconds, usage_rate, days_to_month_end, time_bucket
    metrics: past_wasted_seconds, emergency_unused_seconds, twwi, remaining_days, required_daily_seconds,
             under_risk, over_risk, strategy_state, emergency_dates, past_dates, buffer_dates
    media_platform: 若指定則只計算該媒體；None 表示全媒體合計。
    """
    y, m = int(year), int(month)
    ndays = calendar.monthrange(y, m)[1]
    month_dates = [datetime(y, m, d).date() for d in range(1, ndays + 1)]
    if not monthly_capacity_loader:
        monthly_capacity_loader = lambda mp, yr, mo: get_platform_monthly_capacity(mp, yr, mo)
    daily_cap_total = 0
    platforms_to_sum = [media_platform] if media_platform else MEDIA_PLATFORM_OPTIONS
    for mp in platforms_to_sum:
        cap = monthly_capacity_loader(mp, y, m)
        if cap is not None and cap > 0:
            daily_cap_total += int(cap)
    df = df_daily.copy()
    if media_platform and '媒體平台' in df.columns:
        df = df[df['媒體平台'] == media_platform]
    if df.empty or '日期' not in df.columns or '使用店秒' not in df.columns:
        used_by_date = {d: 0 for d in month_dates}
    else:
        df['日期'] = pd.to_datetime(df['日期'], errors='coerce').dt.date
        used_by_date = df.groupby('日期')['使用店秒'].sum().to_dict()
    last_day = datetime(y, m, ndays).date()
    rows = []
    for d in month_dates:
        cap = daily_cap_total
        used = int(used_by_date.get(d, 0))
        unused = max(0, cap - used) if cap else 0
        usage_rate = (used / cap) if cap else 0
        days_to_end = (last_day - d).days
        if d < today:
            bucket = "past"
        elif d <= today + timedelta(days=emergency_days) and d >= today:
            bucket = "emergency"
        else:
            bucket = "buffer"
        rows.append({
            "date": d,
            "total_capacity_seconds": cap,
            "used_seconds": used,
            "unused_seconds": unused,
            "usage_rate": usage_rate,
            "days_to_month_end": days_to_end,
            "time_bucket": bucket,
        })
    daily_inventory = pd.DataFrame(rows)
    past_wasted = int(daily_inventory[daily_inventory["time_bucket"] == "past"]["unused_seconds"].sum())
    emergency_df = daily_inventory[daily_inventory["time_bucket"] == "emergency"]
    emergency_unused = int(emergency_df["unused_seconds"].sum())
    twwi = sum(
        row["unused_seconds"] * TIME_WEIGHT.get(row["time_bucket"], 0.3)
        for _, row in daily_inventory.iterrows()
    )
    remaining_days = max(0, len(emergency_df))
    required_daily_seconds = (emergency_unused / remaining_days) if remaining_days else 0
    month_usage_rate = (daily_inventory["used_seconds"].sum() / (daily_inventory["total_capacity_seconds"].sum() or 1))
    under_risk = max(0, TARGET_USAGE - month_usage_rate) / TOLERANCE if TOLERANCE else 0
    over_risk = max(0, month_usage_rate - SAFE_LIMIT) / OVER_BUFFER if OVER_BUFFER else 0
    under_high = under_risk >= 0.5
    over_high = over_risk >= 0.5
    if under_high and over_high:
        strategy_state = "ANOMALY"
    elif under_high and not over_high:
        strategy_state = "SELL"
    elif not under_high and over_high:
        strategy_state = "HOLD"
    else:
        strategy_state = "NORMAL"
    metrics = {
        "past_wasted_seconds": past_wasted,
        "emergency_unused_seconds": emergency_unused,
        "twwi": twwi,
        "remaining_days": remaining_days,
        "required_daily_seconds": required_daily_seconds,
        "under_risk": under_risk,
        "over_risk": over_risk,
        "strategy_state": strategy_state,
        "month_usage_rate": month_usage_rate,
        "month_total_capacity": daily_inventory["total_capacity_seconds"].sum(),
        "month_total_used": daily_inventory["used_seconds"].sum(),
        "emergency_dates": emergency_df["date"].tolist(),
        "past_dates": daily_inventory[daily_inventory["time_bucket"] == "past"]["date"].tolist(),
        "buffer_dates": daily_inventory[daily_inventory["time_bucket"] == "buffer"]["date"].tolist(),
    }
    return daily_inventory, metrics


def _seconds_to_spot_label(seconds, sec_per_spot=SECONDS_PER_SPOT_15S, short=False):
    """轉譯為「約 X 檔全省 15 秒」；short=True 時改為「約 X 檔(15秒)」以適應狹窄版面。"""
    if sec_per_spot <= 0:
        return f"{int(seconds):,} 店秒"
    n = round(seconds / sec_per_spot)
    if short:
        return f"約 {n} 檔(15秒)"
    return f"約 {n} 檔全省 15 秒"


# ========== ROI 分頁：依現有資料計算投報率（不寫入資料庫）==========
# 成本：來自「媒體秒數與採購」分頁（購買價格）
# 實收：來自表1 訂單（依各媒體使用秒數比例拆分，或使用拆分金額）
# ROI = (實收 - 購買成本) / 購買成本
SYSTEM_MEDIA_COST_PER_SECOND = {}


def _compute_and_save_split_amount_for_contract(contract_key):
    """
    依「專案實收金額」與各訂單使用秒數比例，計算並寫回每筆訂單的 拆分金額。
    同一合約（contract_key）內，拆分金額 = 專案實收金額 × (該訂單總店秒 / 合約總店秒)。
    """
    if not contract_key:
        return
    try:
        conn = get_db_connection()
        df_ord = pd.read_sql(
            "SELECT id, contract_id, project_amount_net FROM orders WHERE contract_id = ? OR id = ?",
            conn, params=(str(contract_key), str(contract_key))
        )
        if df_ord.empty:
            conn.close()
            return
        project_amt = pd.to_numeric(df_ord['project_amount_net'].iloc[0], errors='coerce')
        if pd.isna(project_amt) or project_amt <= 0:
            conn.close()
            return
        order_ids = df_ord['id'].tolist()
        placeholders = ",".join(["?"] * len(order_ids))
        df_seg = pd.read_sql(
            f"SELECT source_order_id, total_store_seconds FROM ad_flight_segments WHERE source_order_id IN ({placeholders})",
            conn, params=order_ids
        )
        conn.close()
        if df_seg.empty:
            return
        order_seconds = df_seg.groupby('source_order_id')['total_store_seconds'].sum().to_dict()
        total_sec = sum(order_seconds.values()) or 1
        conn = get_db_connection()
        for oid in order_ids:
            sec = order_seconds.get(oid, 0) or 0
            split_val = project_amt * (sec / total_sec)
            conn.execute("UPDATE orders SET split_amount = ? WHERE id = ?", (round(split_val, 2), oid))
        conn.commit()
        conn.close()
    except Exception:
        try:
            conn.close()
        except Exception:
            pass


def get_revenue_per_media_allocated_by_seconds():
    """
    從訂單＋檔次段計算「各媒體實收金額」：優先使用「拆分金額」；若無則依各媒體使用秒數佔該合約總秒數比例拆分「實收金額」。
    回傳 dict: media_platform -> 分配後實收金額(int)。若無資料或錯誤則回傳 {}。
    """
    try:
        conn = get_db_connection()
        df_ord = pd.read_sql("SELECT id, contract_id, amount_net, split_amount FROM orders", conn)
        df_seg = pd.read_sql("SELECT source_order_id, media_platform, total_store_seconds FROM ad_flight_segments WHERE media_platform IS NOT NULL AND total_store_seconds IS NOT NULL", conn)
        conn.close()
    except Exception:
        return {}
    if df_ord.empty or df_seg.empty:
        return {}
    df_seg = df_seg.merge(df_ord, left_on='source_order_id', right_on='id', how='left')
    df_seg['split_amount'] = pd.to_numeric(df_seg['split_amount'], errors='coerce').fillna(0)
    # 若有任一訂單有拆分金額，則各媒體實收 = 依訂單之 split_amount 依 media_platform 加總
    use_split = (df_seg['split_amount'] > 0).any()
    if use_split:
        rev_by_media = df_seg.groupby('media_platform')['split_amount'].sum()
        return {k: int(round(v)) for k, v in rev_by_media.items() if v and v > 0}
    # 否則沿用原邏輯：依合約實收金額按秒數比例拆分
    df_seg['contract_key'] = df_seg['contract_id'].fillna(df_seg['source_order_id'])
    df_seg['amount_net'] = pd.to_numeric(df_seg['amount_net'], errors='coerce').fillna(0)
    contract_revenue = df_ord.copy()
    contract_revenue['contract_key'] = contract_revenue['contract_id'].fillna(contract_revenue['id'])
    contract_revenue['amount_net'] = pd.to_numeric(contract_revenue['amount_net'], errors='coerce').fillna(0)
    contract_total = contract_revenue.groupby('contract_key')['amount_net'].sum().to_dict()
    seg_seconds = df_seg.groupby(['contract_key', 'media_platform'])['total_store_seconds'].sum().reset_index()
    contract_seconds = df_seg.groupby('contract_key')['total_store_seconds'].sum().to_dict()
    revenue_per_media = {}
    for (contract_key, media_platform), grp in seg_seconds.groupby(['contract_key', 'media_platform']):
        media_sec = int(grp['total_store_seconds'].sum())
        total_sec = contract_seconds.get(contract_key, 0) or 1
        rev = contract_total.get(contract_key, 0)
        allocated = rev * (media_sec / total_sec)
        revenue_per_media[media_platform] = revenue_per_media.get(media_platform, 0) + allocated
    return {k: int(round(v)) for k, v in revenue_per_media.items()}


def get_revenue_per_media_by_period(period_type, year, month=None):
    """
    依時間區間計算各媒體實收金額。
    period_type: 'month' | 'quarter' | 'year' | 'all'
    year, month: 指定參考年、月（month 用於 month/quarter 維度）
    回傳 dict: media_platform -> 實收金額(int)
    """
    import calendar
    try:
        conn = get_db_connection()
        df_ord = pd.read_sql("SELECT id, contract_id, amount_net, split_amount FROM orders", conn)
        df_seg = pd.read_sql(
            "SELECT source_order_id, media_platform, total_store_seconds, start_date, end_date FROM ad_flight_segments "
            "WHERE media_platform IS NOT NULL AND total_store_seconds IS NOT NULL", conn
        )
        conn.close()
    except Exception:
        return {}
    if df_ord.empty or df_seg.empty:
        return {}
    df_seg['start_date'] = pd.to_datetime(df_seg['start_date'], errors='coerce')
    df_seg['end_date'] = pd.to_datetime(df_seg['end_date'], errors='coerce')
    df_seg = df_seg.dropna(subset=['start_date', 'end_date'])
    # 依 period_type 決定區間
    if period_type == 'month' and month is not None:
        _, ndays = calendar.monthrange(int(year), int(month))
        period_start = pd.Timestamp(year, month, 1)
        period_end = pd.Timestamp(year, month, ndays)
    elif period_type == 'quarter' and month is not None:
        q = (int(month) - 1) // 3 + 1
        start_m = (q - 1) * 3 + 1
        end_m = q * 3
        period_start = pd.Timestamp(year, start_m, 1)
        _, ndays = calendar.monthrange(year, end_m)
        period_end = pd.Timestamp(year, end_m, ndays)
    elif period_type == 'year':
        period_start = pd.Timestamp(year, 1, 1)
        period_end = pd.Timestamp(year, 12, 31)
    else:
        period_start = pd.Timestamp(2000, 1, 1)
        period_end = pd.Timestamp(2100, 12, 31)
    # 篩選與區間有重疊的 segment
    mask = (df_seg['start_date'] <= period_end) & (df_seg['end_date'] >= period_start)
    df_seg = df_seg[mask]
    if df_seg.empty:
        return {}
    df_seg = df_seg.merge(df_ord, left_on='source_order_id', right_on='id', how='left')
    df_seg['split_amount'] = pd.to_numeric(df_seg['split_amount'], errors='coerce').fillna(0)
    use_split = (df_seg['split_amount'] > 0).any()
    if use_split:
        rev_by_media = df_seg.groupby('media_platform')['split_amount'].sum()
        return {k: int(round(v)) for k, v in rev_by_media.items() if v and v > 0}
    df_seg['contract_key'] = df_seg['contract_id'].fillna(df_seg['source_order_id'])
    df_seg['amount_net'] = pd.to_numeric(df_seg['amount_net'], errors='coerce').fillna(0)
    contract_total = df_ord.copy()
    contract_total['contract_key'] = contract_total['contract_id'].fillna(contract_total['id'])
    contract_total['amount_net'] = pd.to_numeric(contract_total['amount_net'], errors='coerce').fillna(0)
    contract_total = contract_total.groupby('contract_key')['amount_net'].sum().to_dict()
    seg_seconds = df_seg.groupby(['contract_key', 'media_platform'])['total_store_seconds'].sum().reset_index()
    contract_seconds = df_seg.groupby('contract_key')['total_store_seconds'].sum().to_dict()
    revenue_per_media = {}
    for (contract_key, media_platform), grp in seg_seconds.groupby(['contract_key', 'media_platform']):
        media_sec = int(grp['total_store_seconds'].sum())
        total_sec = contract_seconds.get(contract_key, 0) or 1
        rev = contract_total.get(contract_key, 0)
        allocated = rev * (media_sec / total_sec)
        revenue_per_media[media_platform] = revenue_per_media.get(media_platform, 0) + allocated
    return {k: int(round(v)) for k, v in revenue_per_media.items()}


def get_cost_per_media_by_period(period_type, year, month=None):
    """
    依時間區間彙總各媒體購買成本（秒數＋價格）。
    period_type: 'month' | 'quarter' | 'year' | 'all'
    回傳 dict: media_platform -> (purchased_seconds, purchase_cost)
    """
    import calendar
    try:
        conn = get_db_connection()
        if period_type == 'all':
            df = pd.read_sql(
                "SELECT media_platform, purchased_seconds, purchase_price FROM platform_monthly_purchase", conn
            )
        else:
            df = pd.read_sql(
                "SELECT media_platform, year, month, purchased_seconds, purchase_price FROM platform_monthly_purchase WHERE year=?",
                conn, params=(int(year),)
            )
        conn.close()
    except Exception:
        return {}
    if df.empty:
        return {}
    df['purchased_seconds'] = pd.to_numeric(df['purchased_seconds'], errors='coerce').fillna(0)
    df['purchase_price'] = pd.to_numeric(df['purchase_price'], errors='coerce').fillna(0)
    if period_type == 'month' and month is not None:
        df = df[df['month'] == int(month)]
    elif period_type == 'quarter' and month is not None:
        q = (int(month) - 1) // 3 + 1
        start_m, end_m = (q - 1) * 3 + 1, q * 3
        df = df[(df['month'] >= start_m) & (df['month'] <= end_m)]
    elif period_type == 'year':
        pass
    elif period_type == 'all':
        pass
    else:
        return {}
    out = df.groupby('media_platform').agg({'purchased_seconds': 'sum', 'purchase_price': 'sum'}).to_dict('index')
    return {k: (int(v['purchased_seconds']), float(v['purchase_price'])) for k, v in out.items()}


def _get_roi_all_period_date_range():
    """
    取得「累計至今」的實際統計日期範圍。
    回傳 (start_str, end_str) 如 ("2024/1/1", "2026/12/31")，無資料則回傳 (None, None)
    """
    try:
        conn = get_db_connection()
        # 採購資料：year, month
        df_pur = pd.read_sql("SELECT year, month FROM platform_monthly_purchase", conn)
        df_seg = pd.read_sql("SELECT start_date, end_date FROM ad_flight_segments WHERE start_date IS NOT NULL AND end_date IS NOT NULL", conn)
        conn.close()
        dates = []
        if not df_pur.empty:
            for _, r in df_pur.iterrows():
                try:
                    dates.append(pd.Timestamp(int(r['year']), int(r['month']), 1))
                    _, nd = calendar.monthrange(int(r['year']), int(r['month']))
                    dates.append(pd.Timestamp(int(r['year']), int(r['month']), nd))
                except Exception:
                    pass
        if not df_seg.empty:
            df_seg['start_date'] = pd.to_datetime(df_seg['start_date'], errors='coerce')
            df_seg['end_date'] = pd.to_datetime(df_seg['end_date'], errors='coerce')
            dates.extend(df_seg['start_date'].dropna().tolist())
            dates.extend(df_seg['end_date'].dropna().tolist())
        if not dates:
            return None, None
        start_d = min(dates)
        end_d = max(dates)
        return start_d.strftime("%Y/%m/%d"), end_d.strftime("%Y/%m/%d")
    except Exception:
        return None, None


def _calculate_roi_by_period(period_type, year, month, period_label):
    """
    依時間維度計算各媒體 ROI。
    period_type: 'month' | 'quarter' | 'year' | 'all'
    period_label: 顯示用標籤，如 "2026年1月"、"2026 Q1"、"2026年"、"累計至今"
    回傳 list of dict
    """
    revenue_per_media = get_revenue_per_media_by_period(period_type, year, month)
    cost_per_media = get_cost_per_media_by_period(period_type, year, month)
    media_set = set(MEDIA_PLATFORM_OPTIONS)
    media_set.update(revenue_per_media.keys())
    media_set.update(cost_per_media.keys())
    rows = []
    for mp in sorted(media_set):
        cost_row = cost_per_media.get(mp)
        if cost_row is None or not cost_row[0] or cost_row[0] <= 0:
            continue
        purchased_sec, purchase_cost = cost_row[0], cost_row[1]
        revenue = int(revenue_per_media.get(mp, 0) or 0)
        roi = ((revenue - purchase_cost) / purchase_cost) if purchase_cost > 0 else 0
        rows.append({
            "媒體": mp,
            "時間區間": period_label,
            "購買秒數": int(purchased_sec),
            "購買成本（元）": round(purchase_cost, 0),
            "實收金額（元）": revenue,
            "ROI（投報率）": round(roi, 2),
        })
    return rows


def _calculate_roi_from_actual_data(year, month, revenue_per_media):
    """
    依現有資料計算各媒體 ROI。
    成本：該媒體該年該月之購買價格（來自 platform_monthly_purchase）
    實收：revenue_per_media（來自訂單，已依秒數比例或拆分金額分配）
    ROI = (實收 - 購買成本) / 購買成本
    回傳 list of dict，每項含：媒體, 購買秒數, 購買成本（元）, 實收金額（元）, ROI（投報率）
    """
    media_set = set(MEDIA_PLATFORM_OPTIONS)
    for mp in revenue_per_media:
        media_set.add(mp)
    rows = []
    for mp in sorted(media_set):
        row_data = get_platform_monthly_purchase(mp, year, month)
        if row_data is None or not row_data[0] or row_data[0] <= 0:
            continue  # 無採購資料則跳過
        purchased_sec, purchase_price = row_data[0], row_data[1]
        purchase_cost = float(purchase_price) if purchase_price else 0
        revenue = int(revenue_per_media.get(mp, 0) or 0)
        roi = ((revenue - purchase_cost) / purchase_cost) if purchase_cost > 0 else 0
        rows.append({
            "媒體": mp,
            "購買秒數": int(purchased_sec),
            "購買成本（元）": round(purchase_cost, 0),
            "實收金額（元）": revenue,
            "ROI（投報率）": round(roi, 2),
        })
    return rows


def build_annual_seconds_summary(df_daily, year, monthly_capacity_loader=None):
    """
    建年度使用秒數總表（對齊 Excel 年度使用秒數總表）。
    含：頂部使用率（企頻/新鮮視/家樂福/診所 × 1月~12月）、各實體區塊（承包、平均每月店秒、秒數用途分列、使用秒數、未使用秒數、使用率）。
    monthly_capacity_loader: 可選，函數 (media_platform, y, m) -> daily_available_seconds，無則不顯示容量/未使用/使用率。
    回傳: dict with 'top_usage_df', 'entities': { entity_label: { 'avg_monthly_seconds', 'by_type_df', 'used_row', 'unused_row', 'usage_rate_row' } }
    """
    import calendar
    if df_daily.empty or '日期' not in df_daily.columns or '使用店秒' not in df_daily.columns:
        return None
    df = df_daily.copy()
    df['日期'] = pd.to_datetime(df['日期'], errors='coerce')
    df = df.dropna(subset=['日期'])
    df['年'] = df['日期'].dt.year
    df['月'] = df['日期'].dt.month
    df = df[df['年'] == int(year)]
    if '媒體平台' not in df.columns:
        return None
    if '秒數用途' not in df.columns:
        df['秒數用途'] = '銷售秒數'

    # 媒體平台 -> 實體
    def to_entity(mp):
        for ent, platforms in ANNUAL_SUMMARY_MEDIA_MAP.items():
            if mp in platforms:
                return ent
        return None

    df['實體'] = df['媒體平台'].map(to_entity)
    df = df[df['實體'].notna()]

    months = list(range(1, 13))
    month_cols = [f"{m}月" for m in months]

    # 頂部使用率：企頻使用率、新鮮視使用率、家樂福使用率、診所使用率 × 1月~12月
    top_rows = []
    monthly_cap_cache = {}
    if monthly_capacity_loader:
        for ent in ANNUAL_SUMMARY_ENTITY_LABELS:
            row = {'項目': f'{ent}使用率'}
            for m in months:
                cap = 0
                for mp in ANNUAL_SUMMARY_MEDIA_MAP.get(ent, []):
                    daily = monthly_capacity_loader(mp, year, m)
                    if daily is not None and daily > 0:
                        ndays = calendar.monthrange(int(year), m)[1]
                        cap += int(daily) * ndays
                monthly_cap_cache[(ent, m)] = cap
                used = df[(df['實體'] == ent) & (df['月'] == m)]['使用店秒'].sum()
                pct = (used / cap * 100) if cap else 0
                row[f'{m}月'] = round(pct, 1)
            top_rows.append(row)
    top_usage_df = pd.DataFrame(top_rows, columns=['項目'] + month_cols) if top_rows else None

    # 各實體區塊
    entities_out = {}
    for ent in ANNUAL_SUMMARY_ENTITY_LABELS:
        platforms = ANNUAL_SUMMARY_MEDIA_MAP.get(ent, [])
        df_ent = df[df['實體'] == ent]
        # 平均每月店秒：用 1 月容量代表（或各月平均）
        avg_monthly = 0
        if monthly_capacity_loader:
            for mp in platforms:
                daily = monthly_capacity_loader(mp, year, 1)
                if daily is not None and daily > 0:
                    avg_monthly += int(daily) * calendar.monthrange(int(year), 1)[1]
        # 依秒數用途分列 × 1月~12月
        by_type_rows = []
        for stype in SECONDS_USAGE_TYPES:
            row = {'項目': stype}
            for m in months:
                val = df_ent[(df_ent['月'] == m) & (df_ent['秒數用途'] == stype)]['使用店秒'].sum()
                row[f'{m}月'] = int(val)
            by_type_rows.append(row)
        by_type_df = pd.DataFrame(by_type_rows, columns=['項目'] + month_cols)

        # 使用秒數（每月合計）
        used_row = {'項目': '使用秒數'}
        for m in months:
            used_row[f'{m}月'] = int(df_ent[df_ent['月'] == m]['使用店秒'].sum())
        # 未使用秒數、使用率
        unused_row = {'項目': '未使用秒數'}
        rate_row = {'項目': f'{ent}使用率'}
        for m in months:
            cap = monthly_cap_cache.get((ent, m))
            if cap is None and monthly_capacity_loader:
                cap = 0
                for mp in platforms:
                    daily = monthly_capacity_loader(mp, year, m)
                    if daily is not None and daily > 0:
                        ndays = calendar.monthrange(int(year), m)[1]
                        cap += int(daily) * ndays
                monthly_cap_cache[(ent, m)] = cap
            else:
                cap = cap or 0
            used = used_row.get(f'{m}月', 0) or 0
            unused_row[f'{m}月'] = max(0, int(cap) - int(used))
            rate_row[f'{m}月'] = round((used / cap * 100), 1) if cap else 0
        entities_out[ent] = {
            'avg_monthly_seconds': avg_monthly,
            'by_type_df': by_type_df,
            'used_row': used_row,
            'unused_row': unused_row,
            'usage_rate_row': rate_row,
        }
    return {'top_usage_df': top_usage_df, 'entities': entities_out}


def _build_visualization_summary_excel(annual_viz, summary_year):
    """
    將總結表視覺化分頁產出為 Excel 二進位內容。
    包含：① 各媒體平台使用率 ② 各秒數類型使用比例，以及對應的總結表數字。
    包含完整的圖表（轉換為圖片插入Excel）。
    回傳 bytes，若失敗回傳 None。
    """
    try:
        from openpyxl import Workbook
        from openpyxl.drawing.image import Image as ExcelImage
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        import altair as alt
        try:
            import vl_convert as vlc
        except ImportError:
            vlc = None
    except ImportError:
        return None
    
    buf = io.BytesIO()
    try:
        wb = Workbook()
        wb.remove(wb.active)  # 移除預設工作表
        
        month_cols = [f"{m}月" for m in range(1, 13)]
        
        # 輔助函數：將 Altair 圖表轉換為圖片
        def _chart_to_image(chart, scale=2):
            """將 Altair 圖表轉換為 PNG bytes"""
            if vlc is None:
                return None
            try:
                png_data = vlc.vegalite_to_png(chart.to_json(), scale=scale)
                return io.BytesIO(png_data)
            except Exception:
                return None
        
        # 輔助函數：設置單元格樣式
        def _style_cell(cell, is_header=False, is_percentage=False, value=None):
            """設置單元格樣式和顏色"""
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            if is_header:
                cell.font = Font(bold=True, size=10)
                cell.fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
            else:
                cell.font = Font(size=9)
                # 根據數值設置顏色
                if is_percentage and value is not None:
                    try:
                        val_float = float(str(value).replace('%', '').replace(',', ''))
                        if val_float >= 100:
                            cell.fill = PatternFill(start_color='FF6B6B', end_color='FF6B6B', fill_type='solid')
                            cell.font = Font(size=9, color='FFFFFF', bold=True)
                        elif val_float >= 70:
                            cell.fill = PatternFill(start_color='FFD93D', end_color='FFD93D', fill_type='solid')
                            cell.font = Font(size=9, bold=True)
                        elif val_float >= 50:
                            cell.fill = PatternFill(start_color='6BCF7F', end_color='6BCF7F', fill_type='solid')
                            cell.font = Font(size=9)
                    except (ValueError, TypeError):
                        pass
        
        # 輔助函數：添加數據表格到工作表
        def _add_dataframe_to_sheet(ws, df, start_row=1, start_col=1, apply_color=False):
            """將DataFrame添加到工作表"""
            # 添加標題行
            for col_idx, col_name in enumerate(df.columns, start=start_col):
                cell = ws.cell(row=start_row, column=col_idx)
                cell.value = str(col_name)
                _style_cell(cell, is_header=True)
            
            # 添加數據行
            for row_idx, (idx, row) in enumerate(df.iterrows(), start=start_row + 1):
                for col_idx, col_name in enumerate(df.columns, start=start_col):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    val = row[col_name]
                    cell.value = val
                    is_percentage = '使用率' in str(col_name) or '使用率' in str(idx) or (isinstance(val, str) and '%' in val)
                    _style_cell(cell, is_header=False, is_percentage=is_percentage if apply_color else False, value=val)
            
            # 自動調整列寬
            for col_idx, col_name in enumerate(df.columns, start=start_col):
                col_letter = get_column_letter(col_idx)
                max_length = max(len(str(col_name)), max([len(str(row[col_name])) for _, row in df.iterrows()], default=0))
                ws.column_dimensions[col_letter].width = min(max_length + 2, 15)
        
        # ① 各媒體平台使用率
        ws1 = wb.create_sheet("①媒體平台使用率")
        ws1['A1'] = f"① 各媒體平台使用率隨時間變化趨勢 - {summary_year}"
        ws1['A1'].font = Font(bold=True, size=14)
        ws1.merge_cells('A1:N1')
        
        if annual_viz.get('top_usage_df') is not None and not annual_viz['top_usage_df'].empty:
            top_df = annual_viz['top_usage_df'].copy()
            top_df['媒體平台'] = top_df['項目'].str.replace("使用率", "", regex=False)
            chart_df_platform = top_df.set_index("媒體平台")[month_cols].T
            chart_df_platform.index.name = "月份"
            
            # 創建折線圖
            try:
                chart_df_platform_melted = chart_df_platform.reset_index().melt(id_vars='月份', var_name='媒體平台', value_name='使用率')
                chart_df_platform_melted['使用率標籤'] = chart_df_platform_melted['使用率'].apply(lambda x: f"{x:.1f}%" if pd.notna(x) else "")
                
                line_chart = alt.Chart(chart_df_platform_melted).mark_line(point=True).encode(
                    x=alt.X('月份:O', title='月份'),
                    y=alt.Y('使用率:Q', title='使用率 (%)', axis=alt.Axis(format='.1f')),
                    color=alt.Color('媒體平台:N', title='媒體平台'),
                    tooltip=['月份', '媒體平台', alt.Tooltip('使用率:Q', format='.1f', title='使用率 (%)')]
                ).properties(width=700, height=400)
                
                text_chart = alt.Chart(chart_df_platform_melted).mark_text(
                    align='center',
                    baseline='bottom',
                    dy=-8,
                    fontSize=10
                ).encode(
                    x=alt.X('月份:O', title='月份'),
                    y=alt.Y('使用率:Q', title='使用率 (%)', axis=alt.Axis(format='.1f')),
                    text=alt.Text('使用率標籤:N'),
                    color=alt.Color('媒體平台:N', legend=None)
                )
                
                chart_platform = (line_chart + text_chart).properties(width=700, height=400)
                img_data = _chart_to_image(chart_platform)
                if img_data:
                    img = ExcelImage(img_data)
                    img.width = 700
                    img.height = 400
                    ws1.add_image(img, 'A3')
            except Exception:
                pass
            
            # 添加表格（從第25行開始，給圖表留空間）
            _add_dataframe_to_sheet(ws1, top_df, start_row=25, apply_color=True)
        
        # ② 各秒數類型使用比例
        ws2 = wb.create_sheet("②秒數類型比例")
        ws2['A1'] = f"② 各秒數類型使用比例隨時間變化趨勢 - {summary_year}"
        ws2['A1'].font = Font(bold=True, size=14)
        ws2.merge_cells('A1:N1')
        
        by_type_agg = None
        for ent in ANNUAL_SUMMARY_ENTITY_LABELS:
            block = annual_viz.get('entities', {}).get(ent)
            if not block or block.get('by_type_df') is None:
                continue
            bt = block['by_type_df'].set_index("項目")[month_cols]
            if by_type_agg is None:
                by_type_agg = bt.copy()
            else:
                by_type_agg = by_type_agg + bt
        
        if by_type_agg is not None and not by_type_agg.empty:
            # 計算百分比
            monthly_total = by_type_agg.sum(axis=0)
            proportion = by_type_agg.copy()
            for c in month_cols:
                if monthly_total.get(c, 0) and monthly_total[c] > 0:
                    proportion[c] = (by_type_agg[c] / monthly_total[c] * 100)
                else:
                    proportion[c] = 0
            
            # 確保每個月份的比例加總為100%
            for col in proportion.columns:
                monthly_sum = proportion[col].sum()
                if monthly_sum > 0 and abs(monthly_sum - 100) > 0.01:
                    proportion[col] = proportion[col] / monthly_sum * 100
            
            chart_df_type = proportion.T
            chart_df_type.index.name = "月份"
            
            # 創建堆疊長條圖
            try:
                chart_df_type_melted = chart_df_type.reset_index().melt(id_vars='月份', var_name='秒數類型', value_name='比例')
                chart_df_type_melted['比例'] = pd.to_numeric(chart_df_type_melted['比例'], errors='coerce').fillna(0)
                chart_df_type_melted['比例'] = chart_df_type_melted['比例'].clip(lower=0)
                
                # 確保每個月份都有所有秒數類型的數據
                all_types = chart_df_type_melted['秒數類型'].unique()
                all_months = chart_df_type_melted['月份'].unique()
                complete_data = []
                for month in all_months:
                    for sec_type in all_types:
                        existing = chart_df_type_melted[(chart_df_type_melted['月份'] == month) & 
                                                       (chart_df_type_melted['秒數類型'] == sec_type)]
                        if existing.empty:
                            complete_data.append({'月份': month, '秒數類型': sec_type, '比例': 0})
                        else:
                            complete_data.append(existing.iloc[0].to_dict())
                chart_df_type_melted = pd.DataFrame(complete_data)
                
                chart_df_type_melted['比例'] = chart_df_type_melted.groupby('月份')['比例'].transform(
                    lambda x: (x / x.sum() * 100) if x.sum() > 0 else 0
                )
                
                chart_df_type_melted['比例標籤'] = chart_df_type_melted.apply(
                    lambda row: f"{row['比例']:.1f}%" if pd.notna(row['比例']) and row['比例'] > 2 else "", 
                    axis=1
                )
                
                chart_df_type_melted_sorted = chart_df_type_melted.sort_values(['月份', '秒數類型']).copy()
                chart_df_type_melted_sorted = chart_df_type_melted_sorted.reset_index(drop=True)
                chart_df_type_melted_sorted['累積起始'] = chart_df_type_melted_sorted.groupby('月份')['比例'].transform(
                    lambda x: x.shift(1).fillna(0).cumsum()
                )
                chart_df_type_melted_sorted['段中間位置'] = (
                    chart_df_type_melted_sorted['累積起始'] + chart_df_type_melted_sorted['比例'] / 2
                )
                
                bar_chart = alt.Chart(chart_df_type_melted_sorted).mark_bar(size=38).encode(
                    x=alt.X('月份:O', title='月份'),
                    y=alt.Y('比例:Q', title='比例 (%)', 
                           axis=alt.Axis(format='.1f'),
                           stack=True,
                           scale=alt.Scale(domain=[0, 100])),
                    color=alt.Color('秒數類型:N', title='秒數類型', 
                                  sort=alt.SortField('秒數類型', order='ascending'),
                                  legend=alt.Legend(
                        title='秒數類型',
                        orient='right',
                        titleFontSize=12,
                        labelFontSize=10
                    )),
                    order=alt.Order('秒數類型:O', sort='ascending'),
                    tooltip=['月份', '秒數類型', alt.Tooltip('比例:Q', format='.1f', title='比例 (%)')]
                ).properties(width=700, height=400)
                
                text_chart = alt.Chart(chart_df_type_melted_sorted[chart_df_type_melted_sorted['比例標籤'] != '']).mark_text(
                    align='center',
                    baseline='middle',
                    fontSize=10,
                    fontWeight='bold',
                    fill='white'
                ).encode(
                    x=alt.X('月份:O', title='月份'),
                    y=alt.Y('段中間位置:Q', title='比例 (%)', 
                           axis=alt.Axis(format='.1f'),
                           scale=alt.Scale(domain=[0, 100])),
                    text=alt.Text('比例標籤:N'),
                    color=alt.Color('秒數類型:N', legend=None)
                )
                
                chart_type = (bar_chart + text_chart).properties(width=700, height=400)
                img_data = _chart_to_image(chart_type)
                if img_data:
                    img = ExcelImage(img_data)
                    img.width = 700
                    img.height = 400
                    ws2.add_image(img, 'A3')
            except Exception:
                pass
            
            # 添加表格
            proportion_df = proportion.reset_index()
            proportion_df.columns = ['秒數類型'] + month_cols
            _add_dataframe_to_sheet(ws2, proportion_df, start_row=25, apply_color=False)
        
        # 總結表數字：各實體
        for ent in ANNUAL_SUMMARY_ENTITY_LABELS:
            block = annual_viz.get('entities', {}).get(ent)
            if not block:
                continue
            
            ws_ent = wb.create_sheet(f"{ent}")
            ws_ent['A1'] = f"{summary_year} {ent}"
            ws_ent['A1'].font = Font(bold=True, size=12)
            ws_ent.merge_cells('A1:N1')
            
            # 秒數用途分列
            _bt = block.get('by_type_df')
            if _bt is not None and not _bt.empty:
                ws_ent['A3'] = f"{ent} 秒數用途分列（1月～12月）"
                ws_ent['A3'].font = Font(bold=True, size=11)
                _add_dataframe_to_sheet(ws_ent, _bt, start_row=4, apply_color=False)
            
            # 使用/未使用/使用率
            summary_table = pd.DataFrame([
                block.get('used_row', {}),
                block.get('unused_row', {}),
                block.get('usage_rate_row', {}),
            ])
            if not summary_table.empty:
                start_row = len(_bt) + 6 if _bt is not None and not _bt.empty else 4
                ws_ent.cell(row=start_row, column=1).value = f"{ent} 使用/未使用/使用率（1月～12月）"
                ws_ent.cell(row=start_row, column=1).font = Font(bold=True, size=11)
                _add_dataframe_to_sheet(ws_ent, summary_table, start_row=start_row + 1, apply_color=True)
        
        wb.save(buf)
        buf.seek(0)
        return buf.getvalue()
    except Exception as e:
        return None

def _build_visualization_summary_pdf(annual_viz, summary_year):
    """
    將總結表視覺化分頁產出為 PDF 二進位內容。
    包含：① 各媒體平台使用率 ② 各秒數類型使用比例，以及對應的總結表數字。
    包含完整的圖表（轉換為圖片）。
    回傳 bytes，若失敗回傳 None。
    """
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak, Image
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.ttfonts import TTFont
        import altair as alt
        try:
            import vl_convert as vlc
        except ImportError:
            vlc = None
    except ImportError:
        return None
    
    buf = io.BytesIO()
    # 註冊中文字型
    pdf_font_name = None
    windir = os.environ.get('WINDIR', 'C:/Windows')
    font_candidates = [
        (os.path.join(windir, 'Fonts', 'msjh.ttf'), 'CJK'),
        (os.path.join(windir, 'Fonts', 'mingliu.ttc'), 'CJK'),
        (os.path.join(windir, 'Fonts', 'msjh.ttc'), 'CJK'),
        (os.path.join(windir, 'Fonts', 'simsun.ttc'), 'CJK'),
        (os.path.join(windir, 'Fonts', 'simhei.ttf'), 'CJK'),
        ('/System/Library/Fonts/PingFang.ttc', 'CJK'),
        ('/usr/share/fonts/truetype/wqy/wqy-zenhei.ttc', 'CJK'),
    ]
    for font_path, name in font_candidates:
        if not os.path.isfile(font_path):
            continue
        try:
            if font_path.lower().endswith('.ttc'):
                pdfmetrics.registerFont(TTFont(name, font_path, subfontIndex=0))
            else:
                pdfmetrics.registerFont(TTFont(name, font_path))
            pdf_font_name = name
            break
        except Exception:
            continue
    if not pdf_font_name:
        try:
            from reportlab.pdfbase.cidfonts import UnicodeCIDFont
            pdfmetrics.registerFont(UnicodeCIDFont('HeiseiMin-W3'))
            pdf_font_name = 'HeiseiMin-W3'
        except Exception:
            pass
    if not pdf_font_name:
        return None
    
    try:
        # 調整頁面邊距，給表格更多空間
        doc = SimpleDocTemplate(buf, pagesize=A4, rightMargin=30, leftMargin=30, topMargin=40, bottomMargin=30)
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            name='CJKTitle',
            parent=styles['Title'],
            fontName=pdf_font_name,
            fontSize=16,
        )
        heading_style = ParagraphStyle(
            name='CJKHeading2',
            parent=styles['Heading2'],
            fontName=pdf_font_name,
            fontSize=12,
        )
        normal_style = ParagraphStyle(
            name='CJKNormal',
            parent=styles['Normal'],
            fontName=pdf_font_name,
            fontSize=9,
        )
        story = []
        
        # 標題
        title = Paragraph(f"<b>📉 總結表視覺化 {summary_year}</b>", title_style)
        story.append(title)
        story.append(Spacer(1, 12))
        
        month_cols = [f"{m}月" for m in range(1, 13)]
        
        def _df_to_table_data(df):
            if df is None or df.empty:
                return []
            data = []
            # 添加列標題
            if not df.empty:
                # 將列標題轉換為較短的顯示名稱（如果太長）
                header = []
                for col in df.columns:
                    col_str = str(col)
                    # 如果列名太長，截斷或簡化
                    if len(col_str) > 8:
                        col_str = col_str[:6] + '..'
                    header.append(col_str)
                data.append(header)
                # 添加數據行
                for idx, row in df.iterrows():
                    row_data = []
                    for val in row.values:
                        val_str = str(val)
                        # 格式化數字：如果是百分比，保留小數點後1位
                        try:
                            val_float = float(val)
                            if '%' in str(val) or (isinstance(val, str) and val.endswith('%')):
                                val_str = f"{val_float:.1f}%"
                            elif abs(val_float) >= 1000:
                                val_str = f"{val_float:,.0f}"
                            elif abs(val_float) >= 1:
                                val_str = f"{val_float:.1f}"
                            else:
                                val_str = f"{val_float:.2f}"
                        except (ValueError, TypeError):
                            # 如果不是數字，保持原樣但限制長度
                            if len(val_str) > 12:
                                val_str = val_str[:10] + '..'
                        row_data.append(val_str)
                    data.append(row_data)
            return data
        
        def _get_cell_color(val_str, row_name=""):
            """根據數值返回單元格背景色和文字顏色"""
            try:
                # 移除百分號和逗號，轉換為浮點數
                val_clean = str(val_str).replace('%', '').replace(',', '').strip()
                val_float = float(val_clean)
                
                # 如果是使用率相關的行（包含"使用率"關鍵字）
                if '使用率' in str(row_name) or '%' in str(val_str):
                    if val_float >= 100:
                        return colors.HexColor('#ff6b6b'), colors.white  # 紅色背景，白色文字
                    elif val_float >= 70:
                        return colors.HexColor('#ffd93d'), colors.black  # 黃色背景，黑色文字
                    elif val_float >= 50:
                        return colors.HexColor('#6bcf7f'), colors.black  # 綠色背景，黑色文字
                    else:
                        return colors.white, colors.black  # 白色背景，黑色文字
                else:
                    # 非使用率數據：使用淺色背景
                    return colors.white, colors.black
            except (ValueError, TypeError):
                return colors.white, colors.black
        
        def _add_table(data, col_widths=None, title=None, apply_color=False):
            if not data:
                return
            if title:
                story.append(Paragraph(f"<b>{title}</b>", heading_style))
                story.append(Spacer(1, 6))
            
            ncols = len(data[0]) if data else 0
            if ncols == 0:
                return
            
            # 計算可用寬度（A4 寬度 - 左右邊距）
            available_width = A4[0] - 60  # 左右各30
            
            # 根據列數智能分配列寬
            if col_widths is None:
                if ncols <= 4:
                    # 少列數：平均分配
                    col_widths = [available_width / ncols] * ncols
                elif ncols <= 13:
                    # 月份表格：第一列（項目名稱）較寬，其他列較窄
                    first_col_width = available_width * 0.25
                    other_col_width = (available_width - first_col_width) / (ncols - 1)
                    col_widths = [first_col_width] + [other_col_width] * (ncols - 1)
                else:
                    # 很多列：使用最小寬度
                    min_col_width = max(25, available_width / ncols)
                    col_widths = [min_col_width] * ncols
            
            # 確保總寬度不超過可用寬度
            total_width = sum(col_widths)
            if total_width > available_width:
                scale = available_width / total_width
                col_widths = [w * scale for w in col_widths]
            
            t = Table(data, colWidths=col_widths, repeatRows=1)  # repeatRows=1 讓標題在每頁重複
            
            # 基礎表格樣式
            table_style = [
                # 字體和大小
                ('FONTNAME', (0, 0), (-1, -1), pdf_font_name),
                ('FONTSIZE', (0, 0), (-1, 0), 9),  # 標題行稍大
                ('FONTSIZE', (0, 1), (-1, -1), 7),  # 數據行較小
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
                
                # 邊框
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('LINEBELOW', (0, 0), (-1, 0), 1, colors.black),  # 標題行下方粗線
                
                # 標題行背景色
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#e0e0e0')),  # 稍深的灰色
                
                # 對齊
                ('ALIGN', (0, 0), (0, -1), 'LEFT'),  # 第一列左對齊（通常是項目名稱）
                ('ALIGN', (1, 0), (-1, -1), 'CENTER'),  # 其他列居中
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                
                # 內邊距
                ('LEFTPADDING', (0, 0), (-1, -1), 4),
                ('RIGHTPADDING', (0, 0), (-1, -1), 4),
                ('TOPPADDING', (0, 0), (-1, -1), 3),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
            ]
            
            # 如果啟用顏色，為數據單元格添加條件顏色
            if apply_color and len(data) > 1:
                for row_idx in range(1, len(data)):  # 跳過標題行
                    row_name = data[row_idx][0] if data[row_idx] else ""  # 第一列通常是項目名稱
                    for col_idx in range(len(data[row_idx])):
                        if col_idx == 0:
                            # 第一列（項目名稱）：淺灰背景
                            table_style.append(('BACKGROUND', (col_idx, row_idx), (col_idx, row_idx), colors.HexColor('#f5f5f5')))
                            table_style.append(('TEXTCOLOR', (col_idx, row_idx), (col_idx, row_idx), colors.black))
                        else:
                            # 數據列：根據數值應用顏色
                            cell_val = data[row_idx][col_idx] if col_idx < len(data[row_idx]) else ""
                            bg_color, text_color = _get_cell_color(cell_val, row_name)
                            table_style.append(('BACKGROUND', (col_idx, row_idx), (col_idx, row_idx), bg_color))
                            table_style.append(('TEXTCOLOR', (col_idx, row_idx), (col_idx, row_idx), text_color))
            else:
                # 不啟用顏色時，使用交替行顏色
                table_style.append(('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f9f9f9')]))
            
            t.setStyle(TableStyle(table_style))
            story.append(t)
            story.append(Spacer(1, 10))
        
        # 輔助函數：將 Altair 圖表轉換為圖片並添加到 PDF
        def _add_chart_image(chart, width=500):
            """將 Altair 圖表轉換為圖片並添加到 PDF"""
            if vlc is None:
                return  # 如果沒有 vl-convert，跳過圖表
            try:
                # 將圖表轉換為 PNG
                png_data = vlc.vegalite_to_png(chart.to_json(), scale=2)
                # 將 PNG 數據保存到臨時 BytesIO
                img_buf = io.BytesIO(png_data)
                # 創建 ReportLab Image 對象
                img = Image(img_buf, width=width, height=width * 0.57)  # 保持 700:400 的比例
                story.append(img)
                story.append(Spacer(1, 10))
            except Exception:
                pass  # 如果轉換失敗，跳過圖表
        
        # ① 各媒體平台使用率
        story.append(Paragraph("<b>① 各媒體平台使用率隨時間變化趨勢</b>", heading_style))
        story.append(Spacer(1, 6))
        if annual_viz.get('top_usage_df') is not None and not annual_viz['top_usage_df'].empty:
            top_df = annual_viz['top_usage_df'].copy()
            top_df['媒體平台'] = top_df['項目'].str.replace("使用率", "", regex=False)
            chart_df_platform = top_df.set_index("媒體平台")[month_cols].T
            chart_df_platform.index.name = "月份"
            
            # 創建折線圖
            try:
                chart_df_platform_melted = chart_df_platform.reset_index().melt(id_vars='月份', var_name='媒體平台', value_name='使用率')
                chart_df_platform_melted['使用率標籤'] = chart_df_platform_melted['使用率'].apply(lambda x: f"{x:.1f}%" if pd.notna(x) else "")
                
                line_chart = alt.Chart(chart_df_platform_melted).mark_line(point=True).encode(
                    x=alt.X('月份:O', title='月份'),
                    y=alt.Y('使用率:Q', title='使用率 (%)', axis=alt.Axis(format='.1f')),
                    color=alt.Color('媒體平台:N', title='媒體平台'),
                    tooltip=['月份', '媒體平台', alt.Tooltip('使用率:Q', format='.1f', title='使用率 (%)')]
                ).properties(width=700, height=400)
                
                text_chart = alt.Chart(chart_df_platform_melted).mark_text(
                    align='center',
                    baseline='bottom',
                    dy=-8,
                    fontSize=10
                ).encode(
                    x=alt.X('月份:O', title='月份'),
                    y=alt.Y('使用率:Q', title='使用率 (%)', axis=alt.Axis(format='.1f')),
                    text=alt.Text('使用率標籤:N'),
                    color=alt.Color('媒體平台:N', legend=None)
                )
                
                chart_platform = (line_chart + text_chart).properties(width=700, height=400)
                _add_chart_image(chart_platform, width=500)
            except Exception:
                pass
            
            # 添加表格（啟用顏色美化）
            data = _df_to_table_data(top_df)
            _add_table(data, title="對應數字表：年度使用率（各實體 × 1月~12月）", apply_color=True)
        
        story.append(PageBreak())
        
        # ② 各秒數類型使用比例
        story.append(Paragraph("<b>② 各秒數類型使用比例隨時間變化趨勢</b>", heading_style))
        story.append(Spacer(1, 6))
        by_type_agg = None
        for ent in ANNUAL_SUMMARY_ENTITY_LABELS:
            block = annual_viz.get('entities', {}).get(ent)
            if not block or block.get('by_type_df') is None:
                continue
            bt = block['by_type_df'].set_index("項目")[month_cols]
            if by_type_agg is None:
                by_type_agg = bt.copy()
            else:
                by_type_agg = by_type_agg + bt
        
        if by_type_agg is not None and not by_type_agg.empty:
            # 計算百分比
            monthly_total = by_type_agg.sum(axis=0)
            proportion = by_type_agg.copy()
            for c in month_cols:
                if monthly_total.get(c, 0) and monthly_total[c] > 0:
                    proportion[c] = (by_type_agg[c] / monthly_total[c] * 100)
                else:
                    proportion[c] = 0
            
            # 確保每個月份的比例加總為100%
            for col in proportion.columns:
                monthly_sum = proportion[col].sum()
                if monthly_sum > 0 and abs(monthly_sum - 100) > 0.01:
                    proportion[col] = proportion[col] / monthly_sum * 100
            
            chart_df_type = proportion.T
            chart_df_type.index.name = "月份"
            
            # 創建堆疊長條圖
            try:
                chart_df_type_melted = chart_df_type.reset_index().melt(id_vars='月份', var_name='秒數類型', value_name='比例')
                chart_df_type_melted['比例'] = pd.to_numeric(chart_df_type_melted['比例'], errors='coerce').fillna(0)
                chart_df_type_melted['比例'] = chart_df_type_melted['比例'].clip(lower=0)
                
                # 確保每個月份都有所有秒數類型的數據
                all_types = chart_df_type_melted['秒數類型'].unique()
                all_months = chart_df_type_melted['月份'].unique()
                complete_data = []
                for month in all_months:
                    for sec_type in all_types:
                        existing = chart_df_type_melted[(chart_df_type_melted['月份'] == month) & 
                                                       (chart_df_type_melted['秒數類型'] == sec_type)]
                        if existing.empty:
                            complete_data.append({'月份': month, '秒數類型': sec_type, '比例': 0})
                        else:
                            complete_data.append(existing.iloc[0].to_dict())
                chart_df_type_melted = pd.DataFrame(complete_data)
                
                # 重新計算比例百分比
                chart_df_type_melted['比例'] = chart_df_type_melted.groupby('月份')['比例'].transform(
                    lambda x: (x / x.sum() * 100) if x.sum() > 0 else 0
                )
                
                chart_df_type_melted['比例標籤'] = chart_df_type_melted.apply(
                    lambda row: f"{row['比例']:.1f}%" if pd.notna(row['比例']) and row['比例'] > 2 else "", 
                    axis=1
                )
                
                chart_df_type_melted_sorted = chart_df_type_melted.sort_values(['月份', '秒數類型']).copy()
                chart_df_type_melted_sorted = chart_df_type_melted_sorted.reset_index(drop=True)
                chart_df_type_melted_sorted['累積起始'] = chart_df_type_melted_sorted.groupby('月份')['比例'].transform(
                    lambda x: x.shift(1).fillna(0).cumsum()
                )
                chart_df_type_melted_sorted['段中間位置'] = (
                    chart_df_type_melted_sorted['累積起始'] + chart_df_type_melted_sorted['比例'] / 2
                )
                
                bar_chart = alt.Chart(chart_df_type_melted_sorted).mark_bar(size=38).encode(
                    x=alt.X('月份:O', title='月份'),
                    y=alt.Y('比例:Q', title='比例 (%)', 
                           axis=alt.Axis(format='.1f'),
                           stack=True,
                           scale=alt.Scale(domain=[0, 100])),
                    color=alt.Color('秒數類型:N', title='秒數類型', 
                                  sort=alt.SortField('秒數類型', order='ascending'),
                                  legend=alt.Legend(
                        title='秒數類型',
                        orient='right',
                        titleFontSize=12,
                        labelFontSize=10
                    )),
                    order=alt.Order('秒數類型:O', sort='ascending'),
                    tooltip=['月份', '秒數類型', alt.Tooltip('比例:Q', format='.1f', title='比例 (%)')]
                ).properties(width=700, height=400)
                
                text_chart = alt.Chart(chart_df_type_melted_sorted[chart_df_type_melted_sorted['比例標籤'] != '']).mark_text(
                    align='center',
                    baseline='middle',
                    fontSize=10,
                    fontWeight='bold',
                    fill='white'
                ).encode(
                    x=alt.X('月份:O', title='月份'),
                    y=alt.Y('段中間位置:Q', title='比例 (%)', 
                           axis=alt.Axis(format='.1f'),
                           scale=alt.Scale(domain=[0, 100])),
                    text=alt.Text('比例標籤:N'),
                    color=alt.Color('秒數類型:N', legend=None)
                )
                
                chart_type = (bar_chart + text_chart).properties(width=700, height=400)
                _add_chart_image(chart_type, width=500)
            except Exception:
                pass
            
            # 添加表格（不啟用顏色，因為這是比例數據）
            proportion_df = proportion.reset_index()
            proportion_df.columns = ['秒數類型'] + month_cols
            data = _df_to_table_data(proportion_df)
            _add_table(data, apply_color=False)
        
        story.append(PageBreak())
        
        # 總結表數字：各實體
        story.append(Paragraph("<b>📊 總結表數字</b>", heading_style))
        story.append(Spacer(1, 6))
        
        for ent in ANNUAL_SUMMARY_ENTITY_LABELS:
            block = annual_viz.get('entities', {}).get(ent)
            if not block:
                continue
            
            story.append(Paragraph(f"<b>{summary_year} {ent}</b>", heading_style))
            story.append(Spacer(1, 6))
            
            # 秒數用途分列（不啟用顏色）
            _bt = block.get('by_type_df')
            if _bt is not None and not _bt.empty:
                data = _df_to_table_data(_bt)
                _add_table(data, title=f"{ent} 秒數用途分列（1月～12月）", apply_color=False)
            
            # 使用/未使用/使用率（啟用顏色美化，特別是使用率）
            summary_table = pd.DataFrame([
                block.get('used_row', {}),
                block.get('unused_row', {}),
                block.get('usage_rate_row', {}),
            ])
            if not summary_table.empty:
                data = _df_to_table_data(summary_table)
                _add_table(data, title=f"{ent} 使用/未使用/使用率（1月～12月）", apply_color=True)
            
            story.append(Spacer(1, 10))
        
        doc.build(story)
        buf.seek(0)
        return buf.getvalue()
    except Exception as e:
        return None

def _build_annual_summary_pdf(annual, summary_year):
    """
    將年度使用秒數總表產出為 PDF 二進位內容。使用系統中文字型以正確顯示中文。
    回傳 bytes，若失敗回傳 None。
    """
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.ttfonts import TTFont
    except ImportError:
        return None
    buf = io.BytesIO()
    # 註冊中文字型並決定表格/段落使用的字型名稱（優先 .ttf，.ttc 需指定 subfontIndex）
    pdf_font_name = None
    windir = os.environ.get('WINDIR', 'C:/Windows')
    font_candidates = [
        (os.path.join(windir, 'Fonts', 'msjh.ttf'), 'CJK'),      # 微軟正黑體 .ttf
        (os.path.join(windir, 'Fonts', 'mingliu.ttc'), 'CJK'),  # 細明體
        (os.path.join(windir, 'Fonts', 'msjh.ttc'), 'CJK'),     # 微軟正黑體 .ttc
        (os.path.join(windir, 'Fonts', 'simsun.ttc'), 'CJK'),   # 宋體
        (os.path.join(windir, 'Fonts', 'simhei.ttf'), 'CJK'),   # 黑體
        ('/System/Library/Fonts/PingFang.ttc', 'CJK'),
        ('/usr/share/fonts/truetype/wqy/wqy-zenhei.ttc', 'CJK'),
    ]
    for font_path, name in font_candidates:
        if not os.path.isfile(font_path):
            continue
        try:
            if font_path.lower().endswith('.ttc'):
                pdfmetrics.registerFont(TTFont(name, font_path, subfontIndex=0))
            else:
                pdfmetrics.registerFont(TTFont(name, font_path))
            pdf_font_name = name
            break
        except Exception:
            continue
    if not pdf_font_name:
        try:
            from reportlab.pdfbase.cidfonts import UnicodeCIDFont
            pdfmetrics.registerFont(UnicodeCIDFont('HeiseiMin-W3'))
            pdf_font_name = 'HeiseiMin-W3'
        except Exception:
            pass
    if not pdf_font_name:
        return None
    try:
        doc = SimpleDocTemplate(buf, pagesize=A4, rightMargin=40, leftMargin=40, topMargin=40, bottomMargin=40)
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            name='CJKTitle',
            parent=styles['Title'],
            fontName=pdf_font_name,
            fontSize=16,
        )
        heading_style = ParagraphStyle(
            name='CJKHeading2',
            parent=styles['Heading2'],
            fontName=pdf_font_name,
            fontSize=12,
        )
        story = []
        title = Paragraph(f"<b>年度使用秒數總表 {summary_year}</b>", title_style)
        story.append(title)
        story.append(Spacer(1, 12))

        def _df_to_table_data(df):
            if df is None or df.empty:
                return []
            return [[str(x) for x in row] for row in df.values.tolist()]

        def _add_table(data, col_widths=None):
            if not data:
                return
            t = Table(data)
            ncols = len(data[0]) if data else 0
            if col_widths is None:
                col_widths = [max(40, 400 // ncols)] * ncols if ncols else []
            t.setStyle(TableStyle([
                ('FONTNAME', (0, 0), (-1, -1), pdf_font_name),
                ('FONTSIZE', (0, 0), (-1, -1), 8),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ]))
            story.append(t)
            story.append(Spacer(1, 10))

        if annual.get('top_usage_df') is not None and not annual['top_usage_df'].empty:
            top_df = annual['top_usage_df']
            data = [list(top_df.columns)] + _df_to_table_data(top_df)
            _add_table(data)
        for ent in ANNUAL_SUMMARY_ENTITY_LABELS:
            block = annual['entities'].get(ent)
            if not block:
                continue
            story.append(Paragraph(f"<b>{summary_year} {ent}</b>", heading_style))
            story.append(Spacer(1, 6))
            by_type = block['by_type_df']
            data = [list(by_type.columns)] + _df_to_table_data(by_type)
            _add_table(data)
            summary_table = pd.DataFrame([
                block['used_row'],
                block['unused_row'],
                block['usage_rate_row'],
            ])
            data = [list(summary_table.columns)] + _df_to_table_data(summary_table)
            _add_table(data)
            story.append(Spacer(1, 8))
        doc.build(story)
        buf.seek(0)
        return buf.getvalue()
    except Exception:
        return None


def _build_table1_from_segments(df_segments: pd.DataFrame, custom_settings=None, df_orders_info=None) -> pd.DataFrame:
    """
    從 segments 表建立表1（segments 已經正確拆解）
    df_orders_info: 可選，含 id, updated_at, contract_id，有則不讀 DB
    """
    if df_segments.empty:
        return pd.DataFrame()
    
    df = df_segments.copy()
    
    # === 日期處理 ===
    df['start_date'] = pd.to_datetime(df['start_date'], errors='coerce')
    df['end_date'] = pd.to_datetime(df['end_date'], errors='coerce')
    df['走期天數'] = df['duration_days']
    
    # === 區域（已經在 segments 中）===
    df['區域'] = df['region']
    
    # === 店數 ===
    df['店數'] = df['store_count']
    
    # === 檔次與秒數計算 ===
    df['每天總檔次'] = df['spots']
    df['委刊總檔數'] = df['total_spots']
    df['總秒數'] = df['委刊總檔數'] * df['seconds']
    df['使用總秒數'] = df['total_store_seconds']
    
    # === 處理提交日與合約編號（從 orders 取得）===
    if df_orders_info is not None and not df_orders_info.empty:
        df = df.merge(df_orders_info, left_on='source_order_id', right_on='id', how='left', suffixes=('', '_order'))
        df['提交日'] = pd.to_datetime(df['updated_at'], errors='coerce').dt.strftime('%Y/%m/%d')
        df['提交日'] = df['提交日'].fillna('')
    else:
        conn = get_db_connection()
        try:
            try:
                df_orders_info = pd.read_sql("SELECT id, updated_at, contract_id FROM orders", conn)
            except Exception:
                df_orders_info = pd.read_sql("SELECT id, updated_at FROM orders", conn)
                df_orders_info['contract_id'] = None
            conn.close()
            df = df.merge(df_orders_info, left_on='source_order_id', right_on='id', how='left', suffixes=('', '_order'))
            df['提交日'] = pd.to_datetime(df['updated_at'], errors='coerce').dt.strftime('%Y/%m/%d')
            df['提交日'] = df['提交日'].fillna('')
        except Exception:
            df['提交日'] = ''
            if 'contract_id' not in df.columns:
                df['contract_id'] = None
    
    # === 先找出所有訂單涵蓋的日期範圍（用於建立所有日期欄位）===
    all_dates = set()
    for idx, row in df.iterrows():
        if pd.notna(row['start_date']) and pd.notna(row['end_date']):
            date_range = pd.date_range(row['start_date'], row['end_date'], freq='D')
            all_dates.update(date_range)
    
    # 建立所有日期欄位名稱（按日期排序，包含星期幾）
    date_column_names = []
    weekday_map = {0: '一', 1: '二', 2: '三', 3: '四', 4: '五', 5: '六', 6: '日'}
    if all_dates:
        sorted_dates = sorted(all_dates)
        for d in sorted_dates:
            # 欄位名稱格式：月/日(星期)（例如：1/1(四), 1/2(五), 2/14(一)）
            weekday = weekday_map[d.weekday()]
            date_key = f"{d.month}/{d.day}({weekday})"
            if date_key not in date_column_names:
                date_column_names.append(date_key)
    
    # === 建立結果列表 ===
    result_rows = []
    
    for idx, row in df.iterrows():
        # 基本欄位（含公司別：東吳、聲活、鉑霖；合約編號：有 contract_id 則顯示合約編號，否則顯示訂單 ID）
        _contract_id = row.get('contract_id')
        _display_contract = (_contract_id if (pd.notna(_contract_id) and _contract_id) else row.get('source_order_id', ''))
        base_row = {
            '_source_order_id': row.get('source_order_id'),  # 供合併 orders 取得實收金額用，稍後移除
            '業務': row.get('sales', ''),
            '主管': '',  # 預留欄位
            '合約編號': _display_contract,
            '公司': row.get('company', ''),
            '實收金額': 0,  # 由 orders 依 source_order_id 合併後填入
            '除佣實收': 0,
            '製作成本': '',  # 預留欄位
            '獎金%': '',  # 預留欄位
            '核定獎金': '',  # 預留欄位
            '加發獎金': '',  # 預留欄位
            '業務基金': '',  # 預留欄位
            '協力基金': '',  # 預留欄位
            '秒數用途': ('銷售秒數' if (row.get('seconds_type') or '') == '銷售' else (row.get('seconds_type') or '銷售秒數')),
            '提交日': df.loc[idx, '提交日'],
            'HYUNDAI_CUSTIN': row.get('client', ''),  # 客戶名稱
            '秒數': int(row.get('seconds', 0) or 0),
            '素材': row.get('product', ''),
            '起始日': row['start_date'].strftime('%Y/%m/%d') if pd.notna(row['start_date']) else '',
            '終止日': row['end_date'].strftime('%Y/%m/%d') if pd.notna(row['end_date']) else '',
            '走期天數': int(df.loc[idx, '走期天數']),
            '區域': df.loc[idx, '區域'],
            '媒體平台': row.get('media_platform') or get_media_platform_display(row.get('platform'), row.get('channel'), ''),
        }
        
        # === 每日24小時檔次分配（預留，暫時留空）===
        hour_columns = [6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 18, 19, 20, 21, 22, 23, 0, 1]
        for hour in hour_columns:
            base_row[str(hour)] = ''  # 暫時留空，未來可從 CUE 表取得
        
        # === 統計欄位 ===
        base_row['每天總檔次'] = int(df.loc[idx, '每天總檔次'])
        base_row['委刊總檔數'] = int(df.loc[idx, '委刊總檔數'])
        base_row['總秒數'] = int(df.loc[idx, '總秒數'])
        base_row['店數'] = int(df.loc[idx, '店數'])
        base_row['使用總秒數'] = int(df.loc[idx, '使用總秒數'])
        
        # === 日期欄位（每個日期為獨立欄位，對齊 Excel 格式）===
        # 先初始化所有日期欄位為空
        for date_key in date_column_names:
            base_row[date_key] = ''
        
        # 填入走期內的日期檔次
        if pd.notna(row['start_date']) and pd.notna(row['end_date']):
            date_range = pd.date_range(row['start_date'], row['end_date'], freq='D')
            daily_spots = df.loc[idx, '每天總檔次']
            weekday_map = {0: '一', 1: '二', 2: '三', 3: '四', 4: '五', 5: '六', 6: '日'}
            
            for d in date_range:
                weekday = weekday_map[d.weekday()]
                date_key = f"{d.month}/{d.day}({weekday})"
                if date_key in date_column_names:
                    base_row[date_key] = daily_spots
        
        result_rows.append(base_row)
    
    # === 轉換為 DataFrame ===
    df_excel = pd.DataFrame(result_rows)
    
    # === 從 orders 取得實收金額、專案實收金額、拆分金額（以 source_order_id 對應 orders.id）===
    conn = get_db_connection()
    try:
        df_orders_amount = pd.read_sql("SELECT id, amount_net, project_amount_net, split_amount FROM orders", conn)
        conn.close()
        df_excel = df_excel.merge(
            df_orders_amount, left_on='_source_order_id', right_on='id', how='left', suffixes=('', '_order')
        )
        df_excel['實收金額'] = df_excel['amount_net'].fillna(0).astype(int)
        df_excel['除佣實收'] = df_excel['amount_net'].fillna(0).astype(int)
        df_excel['專案實收金額'] = pd.to_numeric(df_excel['project_amount_net'], errors='coerce').fillna(0)
        df_excel['拆分金額'] = pd.to_numeric(df_excel['split_amount'], errors='coerce').fillna(0)
        df_excel = df_excel.drop(columns=['id', 'amount_net', 'project_amount_net', 'split_amount', '_source_order_id'], errors='ignore')
    except Exception:
        try:
            conn = get_db_connection()
            df_orders_amount = pd.read_sql("SELECT id, amount_net FROM orders", conn)
            conn.close()
            df_excel = df_excel.merge(df_orders_amount, left_on='_source_order_id', right_on='id', how='left')
            df_excel['實收金額'] = df_excel['amount_net'].fillna(0).astype(int)
            df_excel['除佣實收'] = df_excel['amount_net'].fillna(0).astype(int)
            df_excel = df_excel.drop(columns=['id', 'amount_net', '_source_order_id'], errors='ignore')
        except Exception:
            df_excel = df_excel.drop(columns=['_source_order_id'], errors='ignore')
        if '專案實收金額' not in df_excel.columns:
            df_excel['專案實收金額'] = 0
        if '拆分金額' not in df_excel.columns:
            df_excel['拆分金額'] = 0
    
    # === 重新排列欄位順序（含公司別、媒體平台、專案實收金額、拆分金額）===
    base_columns = ['業務', '主管', '合約編號', '公司', '實收金額', '除佣實收', '專案實收金額', '拆分金額', '製作成本', '獎金%', 
                    '核定獎金', '加發獎金', '業務基金', '協力基金', '秒數用途', '提交日', 
                    'HYUNDAI_CUSTIN', '秒數', '素材', '起始日', '終止日', '走期天數', '區域', '媒體平台']
    hour_columns = [str(h) for h in [6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 18, 19, 20, 21, 22, 23, 0, 1]]
    stat_columns = ['每天總檔次', '委刊總檔數', '總秒數', '店數', '使用總秒數']

    # 組合所有欄位：基本欄位 + 小時欄位 + 統計欄位 + 日期欄位（已排序）
    all_columns = base_columns + hour_columns + stat_columns + date_column_names
    
    # 只保留實際存在的欄位
    existing_columns = [col for col in all_columns if col in df_excel.columns]
    # 加上其他可能遺漏的欄位
    other_columns = [col for col in df_excel.columns if col not in existing_columns]
    df_excel = df_excel[existing_columns + other_columns]
    
    # === 排序（符合行政直覺：業務 → 合約編號 → 起始日）===
    sort_cols = []
    if '業務' in df_excel.columns:
        sort_cols.append('業務')
    if '合約編號' in df_excel.columns:
        sort_cols.append('合約編號')
    if '起始日' in df_excel.columns:
        sort_cols.append('起始日')
    if sort_cols:
        df_excel = df_excel.sort_values(
            by=sort_cols,
            ascending=[True] * len(sort_cols),
            na_position='last'
        )
    
    df_excel = df_excel.reset_index(drop=True)
    
    return df_excel

def build_excel_table1_view(df_orders: pd.DataFrame, custom_settings=None, use_segments=True, df_segments=None) -> pd.DataFrame:
    """
    Excel 表1-資料（訂單主表）高還原版
    目標：接近 100% 還原行政實際使用的表1
    包含：基本欄位、每日24小時檔次分配（預留）、月份欄位、星期欄位
    
    參數:
        df_orders: 訂單 DataFrame
        custom_settings: 自訂平台設定
        use_segments: 是否從 segments 表建立（預設 True，推薦）
        df_segments: 可選，已載入的 segments，有則不讀 DB（換月加速用）
    """
    # 優先從 segments 表建立（因為 segments 已經正確拆解）
    if use_segments:
        if df_segments is not None and not df_segments.empty:
            cols = ['id', 'updated_at', 'contract_id'] if 'contract_id' in df_orders.columns else ['id', 'updated_at']
            info = df_orders[cols].copy() if all(c in df_orders.columns for c in cols) else df_orders[['id', 'updated_at']].copy()
            if 'contract_id' not in info.columns:
                info['contract_id'] = None
            return _build_table1_from_segments(df_segments, custom_settings, df_orders_info=info)
        conn = get_db_connection()
        try:
            df_seg = pd.read_sql("SELECT * FROM ad_flight_segments", conn)
            conn.close()
            if not df_seg.empty:
                return _build_table1_from_segments(df_seg, custom_settings)
        except Exception:
            conn.close()
    
    # 從 orders 建立（需要自行拆解）
    if df_orders.empty:
        return pd.DataFrame()
    
    df = df_orders.copy()
    
    # === 日期處理 ===
    df['start_date'] = pd.to_datetime(df['start_date'], errors='coerce')
    df['end_date'] = pd.to_datetime(df['end_date'], errors='coerce')
    df['走期天數'] = (df['end_date'] - df['start_date']).dt.days + 1
    df['走期天數'] = df['走期天數'].fillna(0).astype(int)
    
    # === 平台解析（只用於顯示區域）===
    def extract_region(p):
        if pd.isna(p):
            return ''
        p_str = str(p)
        for r in ['全省', '北北基', '桃竹苗', '中彰投', '高高屏', '雲嘉南', '宜花東']:
            if r in p_str:
                return r
        return ''
    
    df['區域'] = df['platform'].apply(extract_region)
    
    # === 媒體平台（表一切換用：全家廣播(企頻)、全家新鮮視、家樂福超市、家樂福量販店）===
    def _media_platform(r):
        platform, channel, _ = parse_platform_region(r['platform'])
        return get_media_platform_display(platform, channel, r.get('platform', ''))
    df['媒體平台'] = df.apply(_media_platform, axis=1)
    
    # === 店數 ===
    def resolve_store(p):
        return get_store_count(p, custom_settings)
    
    df['店數'] = df['platform'].apply(resolve_store)
    
    # === 檔次與秒數計算 ===
    df['每天總檔次'] = df['spots'].fillna(0).astype(int)
    df['委刊總檔數'] = df['每天總檔次'] * df['走期天數']
    df['總秒數'] = df['委刊總檔數'] * df['seconds'].fillna(0).astype(int)
    df['使用總秒數'] = df['總秒數'] * df['店數']
    
    # === 處理 updated_at 日期格式 ===
    df['提交日'] = pd.to_datetime(df['updated_at'], errors='coerce').dt.strftime('%Y/%m/%d')
    df['提交日'] = df['提交日'].fillna('')
    
    # === 先找出所有訂單涵蓋的日期範圍（用於建立所有日期欄位）===
    all_dates = set()
    for idx, row in df.iterrows():
        if pd.notna(row['start_date']) and pd.notna(row['end_date']):
            date_range = pd.date_range(row['start_date'], row['end_date'], freq='D')
            all_dates.update(date_range)
    
    # 建立所有日期欄位名稱（按日期排序，包含星期幾）
    date_column_names = []
    weekday_map = {0: '一', 1: '二', 2: '三', 3: '四', 4: '五', 5: '六', 6: '日'}
    if all_dates:
        sorted_dates = sorted(all_dates)
        for d in sorted_dates:
            # 欄位名稱格式：月/日(星期)（例如：1/1(四), 1/2(五), 2/14(一)）
            weekday = weekday_map[d.weekday()]
            date_key = f"{d.month}/{d.day}({weekday})"
            if date_key not in date_column_names:
                date_column_names.append(date_key)
    
    # === 建立結果列表 ===
    result_rows = []
    
    for idx, row in df.iterrows():
        # 基本欄位（含公司別；合約編號：有 contract_id 則顯示，否則訂單 ID）
        _cid = row.get('contract_id')
        _display_contract = (_cid if (pd.notna(_cid) and _cid) else row.get('id', ''))
        base_row = {
            '業務': row.get('sales', ''),
            '主管': '',  # 預留欄位
            '合約編號': _display_contract,
            '公司': row.get('company', ''),
            '實收金額': int(row.get('amount_net', 0) or 0),
            '除佣實收': int(row.get('amount_net', 0) or 0),  # 暫同實收金額
            '製作成本': '',  # 預留欄位
            '獎金%': '',  # 預留欄位
            '核定獎金': '',  # 預留欄位
            '加發獎金': '',  # 預留欄位
            '業務基金': '',  # 預留欄位
            '協力基金': '',  # 預留欄位
            '秒數用途': row.get('seconds_type') or '銷售秒數',
            '提交日': df.loc[idx, '提交日'],
            'HYUNDAI_CUSTIN': row.get('client', ''),  # 客戶名稱
            '秒數': int(row.get('seconds', 0) or 0),
            '素材': row.get('product', ''),
            '起始日': row['start_date'].strftime('%Y/%m/%d') if pd.notna(row['start_date']) else '',
            '終止日': row['end_date'].strftime('%Y/%m/%d') if pd.notna(row['end_date']) else '',
            '走期天數': df.loc[idx, '走期天數'],
            '區域': df.loc[idx, '區域'],
            '媒體平台': df.loc[idx, '媒體平台'] if '媒體平台' in df.columns else '',
        }
        
        # === 每日24小時檔次分配（預留，暫時留空）===
        # 小時順序：6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 18, 19, 20, 21, 22, 23, 0, 1
        hour_columns = [6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 18, 19, 20, 21, 22, 23, 0, 1]
        for hour in hour_columns:
            base_row[str(hour)] = ''  # 暫時留空，未來可從 CUE 表取得
        
        # === 統計欄位 ===
        base_row['每天總檔次'] = df.loc[idx, '每天總檔次']
        base_row['委刊總檔數'] = df.loc[idx, '委刊總檔數']
        base_row['總秒數'] = df.loc[idx, '總秒數']
        base_row['店數'] = df.loc[idx, '店數']
        base_row['使用總秒數'] = df.loc[idx, '使用總秒數']
        
        # === 日期欄位（每個日期為獨立欄位，對齊 Excel 格式）===
        # 先初始化所有日期欄位為空
        for date_key in date_column_names:
            base_row[date_key] = ''
        
        # 填入走期內的日期檔次
        if pd.notna(row['start_date']) and pd.notna(row['end_date']):
            date_range = pd.date_range(row['start_date'], row['end_date'], freq='D')
            daily_spots = df.loc[idx, '每天總檔次']
            weekday_map = {0: '一', 1: '二', 2: '三', 3: '四', 4: '五', 5: '六', 6: '日'}
            
            for d in date_range:
                weekday = weekday_map[d.weekday()]
                date_key = f"{d.month}/{d.day}({weekday})"
                if date_key in date_column_names:
                    base_row[date_key] = daily_spots
        
        result_rows.append(base_row)
    
    # === 轉換為 DataFrame ===
    df_excel = pd.DataFrame(result_rows)
    
    # === 重新排列欄位順序（對齊 Excel，含公司別、媒體平台）===
    base_columns = ['業務', '主管', '合約編號', '公司', '實收金額', '除佣實收', '製作成本', '獎金%', 
                    '核定獎金', '加發獎金', '業務基金', '協力基金', '秒數用途', '提交日', 
                    'HYUNDAI_CUSTIN', '秒數', '素材', '起始日', '終止日', '走期天數', '區域', '媒體平台']
    hour_columns = [str(h) for h in [6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 18, 19, 20, 21, 22, 23, 0, 1]]
    stat_columns = ['每天總檔次', '委刊總檔數', '總秒數', '店數', '使用總秒數']
    all_columns = base_columns + hour_columns + stat_columns + date_column_names
    existing_columns = [col for col in all_columns if col in df_excel.columns]
    other_columns = [col for col in df_excel.columns if col not in existing_columns]
    df_excel = df_excel[existing_columns + other_columns]
    
    # === 排序（業務 → 合約編號 → 起始日）===
    sort_cols = [c for c in ['業務', '合約編號', '起始日'] if c in df_excel.columns]
    if sort_cols:
        df_excel = df_excel.sort_values(by=sort_cols, ascending=[True] * len(sort_cols), na_position='last')
    
    df_excel = df_excel.reset_index(drop=True)
    return df_excel

# ==========================================
# 3. 介面呈現區 (Streamlit UI)
# ==========================================

st.set_page_config(layout="wide", page_title="秒數控管系統", page_icon="📊")

init_db()

# --- 登入檢查 ---
if 'user' not in st.session_state or st.session_state.get('user') is None:
    st.markdown("### 🔐 登入")
    st.caption("請輸入帳號與密碼。（測試用：已預填行政主管 admin / admin123）")
    with st.form("login_form"):
        login_user = st.text_input("帳號", value="admin", placeholder="username")
        login_pass = st.text_input("密碼", type="password", value="admin123", placeholder="password")
        if st.form_submit_button("登入"):
            u = auth_verify(login_user, login_pass)
            if u:
                st.session_state['user'] = u
                st.success("登入成功")
                st.rerun()
            else:
                st.error("帳號或密碼錯誤")
    st.stop()

user = st.session_state['user']
role = user['role']

# 載入自訂平台設定
custom_settings = load_platform_settings()

# --- 側邊欄：控制區 ---
st.sidebar.title("⚙️ 控制台")
st.sidebar.caption(f"👤 {user['username']}（{role}）")
if st.sidebar.button("🚪 登出", key="btn_logout"):
    del st.session_state['user']
    st.rerun()

# 變更我的密碼（所有使用者）
with st.sidebar.expander("🔑 變更密碼", expanded=False):
    cur_p = st.text_input("目前密碼", type="password", key="chpwd_current")
    new_p1 = st.text_input("新密碼", type="password", key="chpwd_new1")
    new_p2 = st.text_input("確認新密碼", type="password", key="chpwd_new2")
    if st.button("💾 變更密碼", key="chpwd_btn"):
        u = auth_verify(user['username'], cur_p)
        if not u:
            st.error("目前密碼錯誤")
        elif not new_p1 or new_p1 != new_p2:
            st.error("新密碼不一致或為空")
        else:
            auth_change_password(user['username'], new_p1)
            st.success("已變更，請重新登入")
            del st.session_state['user']
            st.rerun()

# 帳號權限管理（僅行政主管）
if role == "行政主管":
    with st.sidebar.expander("👥 帳號管理", expanded=False):
        df_users = auth_list_users()
        st.dataframe(df_users[['username', 'role']], use_container_width=True, hide_index=True)
        st.caption("新增帳號")
        new_u = st.text_input("帳號", key="am_new_username", placeholder="username")
        new_p = st.text_input("密碼", type="password", key="am_new_password", placeholder="password")
        new_r = st.selectbox("權限", ROLES, key="am_new_role")
        if st.button("➕ 新增", key="am_btn_add"):
            ok, msg = auth_create_user(new_u, new_p, new_r)
            if ok:
                st.success("已新增")
                st.rerun()
            else:
                st.error(msg)
        st.caption("刪除帳號")
        del_u = st.selectbox("選擇要刪除的帳號", df_users['username'].tolist(), key="am_del_user")
        if st.button("🗑️ 刪除", key="am_btn_del"):
            if del_u == user['username']:
                st.error("無法刪除目前登入的帳號")
            else:
                auth_delete_user(del_u)
                st.success("已刪除")
                st.rerun()

# 模擬資料：一鍵產生訂單、容量、採購，供各分頁呈現（含表3 使用率、ROI 多樣性）
st.sidebar.markdown("### 📊 資料來源")
if st.sidebar.button("🎲 產生模擬資料", type="primary", help="產生 200 筆訂單＋容量＋採購，表3 使用率 50–120%、ROI 具正負多樣性", key="btn_mock_data"):
    with st.spinner("正在產生模擬資料（訂單、容量、採購）..."):
        success, msg = load_mock_data_with_capacity_to_db(n=200, year=2026)
        if success:
            to_del = [k for k in st.session_state if str(k).startswith("purchase_sec_") or str(k).startswith("purchase_price_")]
            for k in to_del:
                del st.session_state[k]
            st.sidebar.success(msg)
            time.sleep(0.5)
            st.rerun()
        else:
            st.sidebar.error(f"產生失敗: {msg}")

# 匯入 Google 試算表（表1結構）
with st.sidebar.expander("📥 匯入 Google 試算表（表1結構）", expanded=False):
    st.caption("貼上試算表網址或 ID，結構需含：平台、起始日、終止日、秒數、每天總檔次、客戶(HYUNDAI_CUSTIN)、素材、業務、公司、合約編號、實收金額、秒數用途等。")
    gs_url = st.text_input(
        "試算表網址或 ID",
        value="https://docs.google.com/spreadsheets/d/1x2cboM_xmB7nl9aA12O633BzmvPNyJnZoqPipOQhVY4/edit?usp=sharing",
        placeholder="https://docs.google.com/spreadsheets/d/xxx/edit 或 貼上 ID",
        key="gs_import_url"
    )
    gs_replace = st.checkbox("匯入時取代現有資料", value=True, key="gs_replace")
    if st.button("📥 匯入（表1結構）", key="gs_import_btn"):
        if not (gs_url or "").strip():
            st.warning("請輸入試算表網址或 ID")
        else:
            with st.spinner("正在讀取試算表並匯入..."):
                success, msg = import_google_sheet_to_orders(gs_url.strip(), replace_existing=gs_replace)
                if success:
                    st.success(msg)
                    time.sleep(0.5)
                    st.rerun()
                else:
                    st.error(msg)

# 重置資料庫按鈕（用於清除舊 DB 問題）
st.sidebar.markdown("---")
if st.sidebar.button("🧨 重置資料庫（刪除並重建）", help="⚠️ 警告：這會刪除所有現有資料"):
    try:
        conn = get_db_connection()
        conn.close()
        if os.path.exists(DB_FILE):
            os.remove(DB_FILE)
            st.sidebar.success("✅ 已刪除資料庫，將重新初始化")
            time.sleep(1)
            st.rerun()
        else:
            st.sidebar.info("資料庫檔案不存在，無需刪除")
    except Exception as e:
        st.sidebar.error(f"❌ 刪除失敗: {e}")

# 平台設定區
st.sidebar.markdown("---")
st.sidebar.markdown("### 📝 平台設定")
with st.sidebar.expander("設定平台店數與營業時間"):
    # 取得所有平台
    conn = get_db_connection()
    platforms = pd.read_sql("SELECT DISTINCT platform FROM orders", conn)
    conn.close()
    
    if not platforms.empty:
        sel_platform = st.selectbox("選擇平台", platforms['platform'].tolist())
        current_store = get_store_count(sel_platform, custom_settings)
        current_hours = PLATFORM_CAPACITY.get(sel_platform, 18)
        if custom_settings and sel_platform in custom_settings:
            current_hours = custom_settings[sel_platform]['daily_hours']
        
        new_store = st.number_input("店數", min_value=1, value=int(current_store), step=1)
        new_hours = st.number_input("每日營業小時數", min_value=1, max_value=24, value=int(current_hours), step=1)
        
        if st.button("💾 儲存設定"):
            save_platform_settings(sel_platform, new_store, new_hours)
            st.success("設定已儲存！")
            st.rerun()
    else:
        st.info("請先產生模擬資料")

# --- 主畫面：讀取資料（使用快取，切換年月/篩選時不重算）---
_db_mtime = os.path.getmtime(DB_FILE) if os.path.exists(DB_FILE) else 0
st.session_state['_db_mtime'] = _db_mtime  # 表3 fragment 重跑時用
df_orders = _load_orders_cached(_db_mtime)

if df_orders.empty:
    st.warning("📭 資料庫為空，請按左側按鈕產生模擬資料。")
    st.stop()

# 重新載入設定（確保最新）
custom_settings = load_platform_settings()

# 計算庫存：從快取載入 segments 並快取 explode，切換表3 年月時不重算
df_seg_main = _load_segments_cached(_db_mtime)
df_daily = _explode_segments_to_daily_cached(df_seg_main) if not df_seg_main.empty else pd.DataFrame()

# 如果 segments 表為空，嘗試建立
if df_daily.empty and not df_orders.empty:
    with st.spinner("正在建立檔次段..."):
        build_ad_flight_segments(df_orders, custom_settings, write_to_db=True)
        _db_mtime = os.path.getmtime(DB_FILE) if os.path.exists(DB_FILE) else _db_mtime
        df_seg_main = _load_segments_cached(_db_mtime)
        df_daily = _explode_segments_to_daily_cached(df_seg_main) if not df_seg_main.empty else pd.DataFrame()

# --- 分頁呈現（角色導向入口 + 只渲染當前分頁）---
TAB_OPTIONS = ["📋 表1-資料", "📅 表2-秒數明細", "📊 表3-每日庫存", "📉 總結表圖表", "📊 分公司×媒體 每月秒數", "📋 媒體秒數與採購", "📊 ROI", "🧪 Ragic抓取測試", "🧪 實驗分頁"]
# 各角色可見分頁：行政主管=全部(預設)、業務=表1+表3(唯讀)、總經理=總結表+表3+表2+分公司×媒體+ROI+實驗(不呈現表1、不呈現媒體秒數與採購)
TAB_OPTIONS_BY_ROLE = {
    "行政主管": TAB_OPTIONS,  # 擁有所有權限，預設角色
    "業務": ["📋 表1-資料", "📊 表3-每日庫存"],
    "總經理": ["📉 總結表圖表", "📊 表3-每日庫存", "📅 表2-秒數明細", "📊 分公司×媒體 每月秒數", "📊 ROI", "🧪 實驗分頁"],
}

role_label = {"行政主管": "🗂 行政主管", "業務": "🧑‍💼 業務", "總經理": "👔 總經理"}.get(role, role)
st.markdown(f"#### 目前身份：{role_label}")
tab_options_for_role = TAB_OPTIONS_BY_ROLE.get(role, TAB_OPTIONS)
# 若目前選中的分頁不在該角色清單內，預設選第一個
current_tab = st.session_state.get("main_tab", tab_options_for_role[0])
if current_tab not in tab_options_for_role:
    st.session_state["main_tab"] = tab_options_for_role[0]


@st.fragment
def _render_tab3(role_readonly=False):
    """表3 內容：fragment 重跑時只跑此函數。role_readonly=True 時不顯示容量設定（業務唯讀）。"""
    _db = st.session_state.get('_db_mtime', 0)
    df_seg_t3 = _load_segments_cached(_db)
    df_daily_t3 = _explode_segments_to_daily_cached(df_seg_t3) if not df_seg_t3.empty else pd.DataFrame()

    st.markdown("### 📊 每月秒數控管表（對齊 Excel 表3）")
    st.caption("依媒體平台區分：執行秒、可用秒數、使用率、可排日（綠 50%+／黃 70%+／紅 100%+）。可選年份月份。每日可用秒數請至「媒體秒數與採購」分頁設定。")

    default_year = datetime.now().year
    default_month = datetime.now().month
    if not df_daily_t3.empty and '日期' in df_daily_t3.columns:
        df_daily_t3['日期'] = pd.to_datetime(df_daily_t3['日期'], errors='coerce')
        valid = df_daily_t3['日期'].dropna()
        if len(valid) > 0:
            default_year = int(valid.min().year)
            default_month = int(valid.min().month)
    sel_year = st.number_input("年份", min_value=2020, max_value=2030, value=default_year, key="table3_year")
    sel_month = st.number_input("月份", min_value=1, max_value=12, value=default_month, key="table3_month")

    if df_daily_t3.empty or df_seg_t3.empty:
        st.warning("📭 尚無每日或檔次段資料，請先產生模擬資料。")
    elif '媒體平台' not in df_daily_t3.columns:
        st.warning("📭 每日資料尚無媒體平台欄位，請重新產生模擬資料。")
    else:
        monthly_cap = load_platform_monthly_capacity_for(sel_year, sel_month)
        cap_tuple = tuple(sorted(monthly_cap.items())) if monthly_cap else ()
        table3_data = _build_table3_monthly_control_cached(_db, sel_year, sel_month, cap_tuple)
        if not table3_data:
            st.info("該年該月尚無媒體平台資料可顯示。")
        else:
            st.markdown("#### 📺 媒體平台")
            options_mp = ['全部'] + [p for p in MEDIA_PLATFORM_OPTIONS if p in table3_data]
            if len(options_mp) == 1:
                options_mp = ['全部'] + list(table3_data.keys())
            sel_mp = st.radio("選擇媒體平台", options=options_mp, horizontal=True, key="table3_media_filter")

            def _util_color(u):
                if u >= 100:
                    return 'background-color: #ff6b6b; color: white'
                if u >= 70:
                    return 'background-color: #ffd93d'
                return 'background-color: #6bcf7f'

            to_show = list(table3_data.keys()) if sel_mp == '全部' else [sel_mp]
            for mp in to_show:
                if mp not in table3_data:
                    continue
                df_t3 = table3_data[mp].copy()
                date_cols_t3 = [c for c in df_t3.columns if c not in ('授權', '項目', '秒數', '%')]
                # --- 一句話解讀 Summary Bar ---
                row_util_vals = table3_data[mp].iloc[3]
                util_vals = [row_util_vals.get(c) for c in date_cols_t3 if isinstance(row_util_vals.get(c), (int, float)) and pd.notna(row_util_vals.get(c))]
                n_red = sum(1 for u in util_vals if u >= 100)
                n_yellow = sum(1 for u in util_vals if 70 <= u < 100)
                n_green = sum(1 for u in util_vals if u < 70)
                try:
                    _mu = table3_data[mp].iloc[0].get('%')
                    month_util = float(_mu) if _mu is not None and pd.notna(_mu) else (sum(util_vals) / len(util_vals) if util_vals else 0)
                except (TypeError, KeyError, ValueError):
                    month_util = sum(util_vals) / len(util_vals) if util_vals else 0
                util_label = f"{round(float(month_util), 1)}%" if isinstance(month_util, (int, float)) else "—"
                if isinstance(month_util, (int, float)) and month_util >= 100:
                    util_status = "🔴 已滿"
                    suggestion = "建議：避免再加全省案，僅可補區域。"
                elif isinstance(month_util, (int, float)) and month_util >= 70:
                    util_status = "⚠️ 偏高"
                    suggestion = "建議：注意檔期集中，可考慮分散排程。"
                else:
                    util_status = "✅ 尚可"
                    suggestion = "建議：可排新案，留意熱門日期。"
                st.markdown(f"**{mp}**")
                st.markdown(f"📌 **{sel_year}/{sel_month} {mp}**  ")
                st.markdown(f"- 本月使用率：**{util_label}**（{util_status}）  ")
                st.markdown(f"- 🔴 紅色天數：{n_red} 天　🟡 黃色天數：{n_yellow} 天　🟢 綠色天數：{n_green} 天  ")
                st.markdown(f"- {suggestion}")
                st.markdown("")

                for col in date_cols_t3:
                    val = df_t3.at[3, col]
                    if isinstance(val, (int, float)) and pd.notna(val):
                        df_t3.at[3, col] = f"{round(float(val), 1)}%"
                orig_row4 = table3_data[mp].iloc[3].copy()
                fixed_cols_t3 = ['授權', '項目', '秒數', '%']
                # 依週分塊（每塊 6 天），欄數少、避免左右滑動；日期欄為 月/日(星期)
                chunk_size = 6
                date_chunks = [date_cols_t3[i:i + chunk_size] for i in range(0, len(date_cols_t3), chunk_size)]
                def _style_chunk(row, chunk_dates):
                    out = [''] * len(row)
                    if row.name != 3:
                        return out
                    for i, c in enumerate(row.index):
                        if c in chunk_dates:
                            orig = orig_row4.get(c)
                            if isinstance(orig, (int, float)) and pd.notna(orig):
                                out[i] = _util_color(orig)
                    return out
                # 可排日圖例拉出表格外，避免長文字造成左右滑動
                st.caption("🟢 綠 &lt;70%　🟡 黃 70%+　🔴 紅 100%+")
                # 一週一列：每週一塊表（全寬），垂直排列，不需左右滑動
                for chunk in date_chunks:
                    sub = df_t3[fixed_cols_t3 + chunk]
                    st.caption(f"**{chunk[0]} ～ {chunk[-1]}**")
                    _num_cols_sub = sub.select_dtypes(include=[np.number]).columns.tolist()
                    _fmt_sub = {c: "{:,.1f}" for c in _num_cols_sub} if _num_cols_sub else {}
                    st.dataframe(sub.style.format(_fmt_sub).apply(lambda row: _style_chunk(row, chunk), axis=1), use_container_width=True)

                # --- 點擊日期展開當日明細（可排日互動）---
                if not df_daily_t3.empty and '媒體平台' in df_daily_t3.columns and '日期' in df_daily_t3.columns:
                    df_daily_t3['日期'] = pd.to_datetime(df_daily_t3['日期'], errors='coerce')
                    month_dates = [d for d in df_daily_t3['日期'].dropna().unique() if d.year == sel_year and d.month == sel_month]
                    month_dates = sorted(month_dates)
                    if month_dates:
                        date_options = ["— 選擇日期查看當日明細 —"] + [f"{d.month}/{d.day}" for d in month_dates]
                        sel_date_str = st.selectbox(f"選擇日期（{mp}）", date_options, key=f"table3_sel_date_{mp}_{sel_year}_{sel_month}")
                        if sel_date_str != "— 選擇日期查看當日明細 —":
                            try:
                                parts = sel_date_str.split("/")
                                day = int(parts[1])
                                target_d = pd.Timestamp(sel_year, sel_month, day)
                                dd = df_daily_t3[(df_daily_t3['媒體平台'] == mp) & (df_daily_t3['日期'].dt.normalize() == target_d)]
                                if not dd.empty:
                                    show_cols = [c for c in ['日期', '媒體平台', '公司', '業務', '客戶', '產品', '使用店秒', '秒數', '檔次'] if c in dd.columns]
                                    _dd_show = dd[show_cols] if show_cols else dd
                                    st.dataframe(_styler_one_decimal(_dd_show), use_container_width=True, height=min(200, 80 + len(dd) * 38))
                                else:
                                    st.caption("該日無使用紀錄")
                            except Exception:
                                pass
                st.markdown("---")

            st.markdown("#### 🎨 可排日顏色說明")
            c1, c2, c3 = st.columns(3)
            with c1:
                st.markdown("🟢 **綠**：使用率 &lt; 70%")
            with c2:
                st.markdown("🟡 **黃**：70% ≤ 使用率 &lt; 100%")
            with c3:
                st.markdown("🔴 **紅**：使用率 ≥ 100%")


# --- 分頁按鈕區（與角色區隔、按鈕較大較清楚）---
st.markdown("---")
st.markdown("### 選擇分頁")
selected_tab = st.session_state.get("main_tab", tab_options_for_role[0])
tab_cols = st.columns(len(tab_options_for_role))
for i, tab in enumerate(tab_options_for_role):
    with tab_cols[i]:
        is_selected = (selected_tab == tab)
        if st.button(tab, key=f"tab_btn_{i}", type="primary" if is_selected else "secondary", use_container_width=True):
            st.session_state["main_tab"] = tab
            st.rerun()

st.markdown("---")

if selected_tab == "📋 表1-資料":
    st.markdown("### 📋 表1－資料（訂單主表）")
    st.caption("此表對應 Excel：秒數管理表 → 表1-資料，為行政與業務對帳用之訂單主表。目前使用模擬資料呈現。")
    # 換月份時不重算：表1 依 _db_mtime 快取於 session_state
    if st.session_state.get('_table1_cache_key') == _db_mtime and '_table1_cache' in st.session_state:
        df_table1 = st.session_state['_table1_cache']
    else:
        df_table1 = build_excel_table1_view(df_orders, custom_settings, use_segments=True, df_segments=df_seg_main)
        st.session_state['_table1_cache'] = df_table1
        st.session_state['_table1_cache_key'] = _db_mtime
    
    if df_table1.empty:
        st.warning("📭 尚無訂單資料")
        st.stop()
    
    # === 實收金額顯示模式：同一合約常只收一筆，表1卻拆成多列，可選「依合約合併」只於每合約第一列顯示總額 ===
    if '實收金額' in df_table1.columns and '合約編號' in df_table1.columns:
        amount_display_mode = st.radio(
            "實收金額顯示",
            options=["依訂單列（每列顯示該筆訂單金額）", "依合約合併（每合約只顯示一筆總額於第一列）"],
            index=0,
            horizontal=True,
            key="table1_amount_display_mode",
        )
        if "依合約合併" in amount_display_mode:
            contract_total = df_table1.groupby('合約編號')['實收金額'].transform('sum')
            first_in_contract = ~df_table1.duplicated('合約編號', keep='first')
            df_table1 = df_table1.copy()
            df_table1['實收金額'] = np.where(first_in_contract, contract_total, 0)
            df_table1['除佣實收'] = df_table1['實收金額']
    
    # === 媒體平台切換按鈕（全家廣播(企頻)、全家新鮮視、家樂福超市、家樂福量販店）===
    if '媒體平台' in df_table1.columns:
        st.markdown("#### 📺 媒體平台切換")
        platform_options = ['全部'] + [p for p in MEDIA_PLATFORM_OPTIONS if p in df_table1['媒體平台'].unique().tolist()]
        if len(platform_options) == 1:
            platform_options = ['全部'] + list(MEDIA_PLATFORM_OPTIONS)
        selected_platform = st.radio(
            "選擇要顯示的媒體平台",
            options=platform_options,
            horizontal=True,
            key="table1_media_platform_filter"
        )
        
        if selected_platform != '全部':
            df_table1 = df_table1[df_table1['媒體平台'] == selected_platform]
            if df_table1.empty:
                st.info(f"📭 媒體平台「{selected_platform}」目前沒有資料")
                st.stop()
    elif '平台分類' in df_table1.columns:
        st.markdown("#### 📺 平台篩選")
        platform_categories = ['全部', '全家新鮮視', '全家廣播', '家樂福', '診所', '其他']
        selected_platform = st.radio(
            "選擇要顯示的平台",
            options=platform_categories,
            horizontal=True,
            key="table1_platform_filter"
        )
        if selected_platform != '全部':
            df_table1 = df_table1[df_table1['平台分類'] == selected_platform]
            if df_table1.empty:
                st.info(f"📭 平台「{selected_platform}」目前沒有資料")
                st.stop()
    
    # === KPI 區 ===
    col1, col2, col3, col4 = st.columns(4)
    with col1:
        st.metric("訂單筆數", len(df_table1))
    with col2:
        st.metric("客戶數", df_table1['客戶'].nunique() if '客戶' in df_table1.columns else (df_table1['HYUNDAI_CUSTIN'].nunique() if 'HYUNDAI_CUSTIN' in df_table1.columns else 0))
    with col3:
        if '媒體平台' in df_table1.columns:
            st.metric("媒體平台數", df_table1['媒體平台'].nunique())
        elif '平台分類' in df_table1.columns:
            st.metric("平台數", df_table1['平台分類'].nunique())
        else:
            st.metric("平台數", df_table1['平台'].nunique() if '平台' in df_table1.columns else 0)
    with col4:
        total_amount = df_table1['實收金額'].sum() if '實收金額' in df_table1.columns else 0
        st.metric("實收金額總計", f"{total_amount:,}")
    
    # === 篩選區（Excel 沒有，但網頁必備）===
    with st.expander("🔍 篩選條件", expanded=False):
        col1, col2, col3 = st.columns(3)
        with col1:
            if '公司' in df_table1.columns:
                sel_company = st.selectbox("公司", ['全部'] + sorted(df_table1['公司'].unique().tolist()))
            else:
                sel_company = '全部'
        with col2:
            if '業務' in df_table1.columns:
                sel_sales = st.selectbox("業務", ['全部'] + sorted(df_table1['業務'].unique().tolist()))
            else:
                sel_sales = '全部'
        with col3:
            client_col_filter = "客戶" if "客戶" in df_table1.columns else "HYUNDAI_CUSTIN"
            if client_col_filter in df_table1.columns:
                sel_client = st.selectbox("客戶", ['全部'] + sorted(df_table1[client_col_filter].dropna().unique().astype(str).tolist()))
            else:
                sel_client = '全部'
    
    # 套用篩選
    df_filtered = df_table1.copy()
    if sel_company != '全部' and '公司' in df_filtered.columns:
        df_filtered = df_filtered[df_filtered['公司'] == sel_company]
    if sel_sales != '全部' and '業務' in df_filtered.columns:
        df_filtered = df_filtered[df_filtered['業務'] == sel_sales]
    client_col_filter = "客戶" if "客戶" in df_table1.columns else "HYUNDAI_CUSTIN"
    if sel_client != '全部' and client_col_filter in df_filtered.columns:
        df_filtered = df_filtered[df_filtered[client_col_filter].astype(str) == sel_client]

    # === 表1 三段式視圖：精簡（業務）/ 行政 / 完整；行政主管一進入即預設「完整」===
    table1_default_index = 2 if role == "行政主管" else 0  # 完整=2, 精簡=0
    view_mode = st.radio(
        "顯示模式",
        ["精簡", "行政", "完整"],
        format_func=lambda x: {"精簡": "① 精簡（業務：合約/客戶/平台/秒數/檔次/起訖/使用秒數）", "行政": "② 行政（+ 日期欄位、店數、委刊總檔數）", "完整": "③ 完整（全部欄位）"}[x],
        index=table1_default_index,
        horizontal=True,
        key="table1_view_mode"
    )
    # 精簡欄位：業務、合約編號、客戶、媒體平台、秒數、每天總檔次、起始日、終止日、使用總秒數
    client_col = "客戶" if "客戶" in df_filtered.columns else "HYUNDAI_CUSTIN"
    cols_simple = [c for c in ['業務', '合約編號', client_col, '媒體平台', '秒數', '每天總檔次', '起始日', '終止日', '使用總秒數'] if c in df_filtered.columns]
    # 行政 = 精簡 + 店數、委刊總檔數 + 日期欄位（格式 月/日(星期)）
    date_cols_t1 = [c for c in df_filtered.columns if re.match(r'^\d{1,2}/\d{1,2}\([一二三四五六日]\)$', str(c))]
    cols_admin = cols_simple + [c for c in ['店數', '委刊總檔數'] if c in df_filtered.columns] + date_cols_t1
    if view_mode == "精簡":
        show_cols = cols_simple
    elif view_mode == "行政":
        show_cols = [c for c in cols_admin if c in df_filtered.columns]
    else:
        show_cols = list(df_filtered.columns)
    df_display = df_filtered[[c for c in show_cols if c in df_filtered.columns]]

    # === 顯示表格（橫向滾動支援）===
    st.markdown("#### 📊 表1-資料（可橫向滾動查看完整欄位）")

    st.dataframe(
        _styler_one_decimal(df_display),
        use_container_width=True,
        height=650
    )

    # 提示資訊
    st.info("💡 **提示**：此表格較寬，請使用橫向滾動查看完整內容。\n"
            "- 每日24小時檔次分配欄位（6-23, 0-1點）目前為預留，未來可從 CUE 表取得詳細資料\n"
            "- 月份欄位顯示該月每天的檔次數\n"
            "- 星期序列顯示走期內每天的星期標記")
    
    # === 下載功能（對齊 Excel 使用情境，依目前顯示模式）===
    st.markdown("#### 📥 下載資料")
    st.info("💡 **提示**：建議下載 Excel 格式以避免編碼問題。下載內容依目前顯示模式（精簡/行政/完整）。")
    
    col1, col2 = st.columns(2)
    with col1:
        # Excel 格式（推薦）
        excel_bytes = df_to_excel_bytes(df_display, sheet_name="表1-資料")
        st.download_button(
            label="📥 下載 Excel（推薦）",
            data=excel_bytes,
            file_name=f"表1_資料_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    with col2:
        # CSV 格式（備選）
        csv = df_display.to_csv(index=False, encoding='utf-8-sig')
        st.download_button(
            label="📥 下載 CSV（備選）",
            data=csv,
            file_name=f"表1_資料_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv",
            mime="text/csv; charset=utf-8"
        )

    # === 訂單逐筆管理（新增／編輯／刪除，直接做在表1內）===
    st.markdown("---")
    st.markdown("#### 📝 訂單逐筆管理（新增／編輯／刪除）")
    st.caption("新增一筆：於下方表單填寫後儲存。每列可點「編輯」修改或「刪除」移除該筆訂單；變更後會自動重建檔次段。")
    conn_crud = get_db_connection()
    df_orders_crud = pd.read_sql("SELECT * FROM orders", conn_crud)
    conn_crud.close()

    def _idx(lst, val, default=0):
        try:
            return lst.index(val) if val in lst else default
        except (ValueError, TypeError):
            return default

    # 新增一筆：多出一列對應欄位自己填
    with st.expander("➕ 新增一筆訂單（填寫欄位後儲存）", expanded=False):
        c1, c2 = st.columns(2)
        with c1:
            new_id = st.text_input("訂單 ID（唯一）", key="crud_new_id", placeholder="例如 mock_2026_c001_01")
            new_contract_id = st.text_input("所屬合約編號（選填）", key="crud_new_contract_id", placeholder="同合約多列填相同值")
            new_platform = st.selectbox("平台", MOCK_PLATFORM_RAW, key="crud_new_platform")
            new_client = st.text_input("客戶", key="crud_new_client", value="")
            new_product = st.text_input("產品名稱", key="crud_new_product", value="")
            new_sales = st.selectbox("業務", MOCK_SALES, key="crud_new_sales")
            new_company = st.selectbox("公司", MOCK_COMPANY, key="crud_new_company")
            new_seconds_type = st.selectbox("秒數用途", SECONDS_USAGE_TYPES, key="crud_new_seconds_type")
        with c2:
            new_start = st.date_input("開始日", value=datetime(2026, 1, 1), key="crud_new_start")
            new_end = st.date_input("結束日", value=datetime(2026, 1, 31), key="crud_new_end")
            new_seconds = st.selectbox("秒數", MOCK_SECONDS, key="crud_new_seconds")
            new_spots = st.number_input("檔次", min_value=2, value=10, step=2, key="crud_new_spots")
            new_amount = st.number_input("實收金額（未稅）", min_value=0, value=100000, step=10000, key="crud_new_amount")
            new_project_amount = st.number_input("專案實收金額（同專案填同一數字，選填）", min_value=0, value=0, step=10000, key="crud_new_project_amount", help="同一合約編號多筆時填一次總額即可，系統會依使用秒數比例計算「拆分金額」")
            new_split_amount = st.number_input("拆分金額（選填，或由專案實收自動計算）", min_value=0, value=0, step=10000, key="crud_new_split_amount", help="ROI 等計算使用此欄；有填專案實收時儲存後會自動依比例計算")
        if st.button("💾 儲存新增", key="crud_btn_add"):
            if not new_id or not new_client or not new_product:
                st.error("請填寫訂單 ID、客戶、產品名稱")
            elif not df_orders_crud.empty and new_id in df_orders_crud['id'].tolist():
                st.error(f"訂單 ID「{new_id}」已存在")
            else:
                conn_ins = get_db_connection()
                try:
                    contract_id_val = (new_contract_id or '').strip() or None
                    project_val = float(new_project_amount) if new_project_amount else None
                    split_val = float(new_split_amount) if new_split_amount else None
                    conn_ins.execute("""
                        INSERT INTO orders (id, platform, client, product, sales, company, start_date, end_date, seconds, spots, amount_net, updated_at, contract_id, seconds_type, project_amount_net, split_amount)
                        VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
                    """, (new_id, new_platform, new_client, new_product, new_sales, new_company,
                          new_start.strftime("%Y-%m-%d"), new_end.strftime("%Y-%m-%d"),
                          int(new_seconds), int(new_spots), float(new_amount), datetime.now().strftime("%Y-%m-%d %H:%M:%S"), contract_id_val, new_seconds_type, project_val, split_val))
                    conn_ins.commit()
                    df_after = pd.read_sql("SELECT * FROM orders", conn_ins)
                    conn_ins.close()
                    build_ad_flight_segments(df_after, load_platform_settings(), write_to_db=True)
                    if project_val and project_val > 0 and contract_id_val:
                        _compute_and_save_split_amount_for_contract(contract_id_val)
                    st.success("✅ 已新增一筆")
                    if '_table1_cache_key' in st.session_state:
                        del st.session_state['_table1_cache_key']
                    st.rerun()
                except Exception as e:
                    conn_ins.rollback()
                    conn_ins.close()
                    st.error(f"新增失敗: {e}")

    # 編輯表單（僅在點選某列「編輯」時顯示）
    crud_edit_id = st.session_state.get("crud_edit_id")
    if crud_edit_id and not df_orders_crud.empty:
        edit_match = df_orders_crud[df_orders_crud['id'] == crud_edit_id]
        if not edit_match.empty:
            selected_row = edit_match.iloc[0]
            with st.expander("✏️ 編輯此筆訂單", expanded=True):
                col_edit_a, col_edit_b = st.columns(2)
                with col_edit_a:
                    edit_contract_id = st.text_input("所屬合約編號（選填）", value=(str(selected_row.get('contract_id')) if (pd.notna(selected_row.get('contract_id')) and selected_row.get('contract_id')) else ''), key="crud_edit_contract_id")
                    edit_platform = st.selectbox("平台", MOCK_PLATFORM_RAW, index=_idx(MOCK_PLATFORM_RAW, selected_row.get('platform')), key="crud_edit_platform")
                    edit_client = st.text_input("客戶", value=selected_row.get('client', '') or '', key="crud_edit_client")
                    edit_product = st.text_input("產品名稱", value=selected_row.get('product', '') or '', key="crud_edit_product")
                    edit_sales = st.selectbox("業務", MOCK_SALES, index=_idx(MOCK_SALES, selected_row.get('sales')), key="crud_edit_sales")
                    edit_company = st.selectbox("公司", MOCK_COMPANY, index=_idx(MOCK_COMPANY, selected_row.get('company')), key="crud_edit_company")
                    edit_seconds_type = st.selectbox("秒數用途", SECONDS_USAGE_TYPES, index=_idx(SECONDS_USAGE_TYPES, selected_row.get('seconds_type') or '銷售秒數'), key="crud_edit_seconds_type")
                with col_edit_b:
                    try:
                        _start_val = pd.to_datetime(selected_row['start_date'], errors='coerce')
                        edit_start_val = _start_val.date() if pd.notna(_start_val) else datetime(2026, 1, 1).date()
                    except Exception:
                        edit_start_val = datetime(2026, 1, 1).date()
                    try:
                        _end_val = pd.to_datetime(selected_row['end_date'], errors='coerce')
                        edit_end_val = _end_val.date() if pd.notna(_end_val) else datetime(2026, 1, 31).date()
                    except Exception:
                        edit_end_val = datetime(2026, 1, 31).date()
                    edit_start = st.date_input("開始日", value=edit_start_val, key="crud_edit_start")
                    edit_end = st.date_input("結束日", value=edit_end_val, key="crud_edit_end")
                    edit_seconds = st.number_input("秒數", min_value=5, max_value=60, value=int(selected_row['seconds']), key="crud_edit_seconds")
                    edit_spots = st.number_input("檔次", min_value=2, value=int(selected_row['spots']), step=2, key="crud_edit_spots")
                    edit_amount = st.number_input("實收金額（未稅）", min_value=0, value=int(selected_row['amount_net']), step=10000, key="crud_edit_amount")
                    _proj = selected_row.get('project_amount_net')
                    edit_project_amount = st.number_input("專案實收金額（同專案填同一數字，選填）", min_value=0, value=int(_proj) if pd.notna(_proj) and _proj else 0, step=10000, key="crud_edit_project_amount")
                    _split = selected_row.get('split_amount')
                    edit_split_amount = st.number_input("拆分金額（選填，或由專案實收自動計算）", min_value=0, value=int(_split) if pd.notna(_split) and _split else 0, step=10000, key="crud_edit_split_amount")
                col_save, col_cancel, _ = st.columns([1, 1, 2])
                with col_save:
                    if st.button("💾 儲存編輯", key="crud_btn_edit"):
                        conn_up = get_db_connection()
                        try:
                            edit_contract_id_val = (edit_contract_id or '').strip() or None
                            project_val = float(edit_project_amount) if edit_project_amount else None
                            split_val = float(edit_split_amount) if edit_split_amount else None
                            conn_up.execute("""
                                UPDATE orders SET platform=?, client=?, product=?, sales=?, company=?, start_date=?, end_date=?, seconds=?, spots=?, amount_net=?, updated_at=?, contract_id=?, seconds_type=?, project_amount_net=?, split_amount=?
                                WHERE id=?
                            """, (edit_platform, edit_client, edit_product, edit_sales, edit_company,
                                  edit_start.strftime("%Y-%m-%d"), edit_end.strftime("%Y-%m-%d"),
                                  int(edit_seconds), int(edit_spots), float(edit_amount), datetime.now().strftime("%Y-%m-%d %H:%M:%S"), edit_contract_id_val, edit_seconds_type, project_val, split_val, selected_row['id']))
                            conn_up.commit()
                            df_after = pd.read_sql("SELECT * FROM orders", conn_up)
                            conn_up.close()
                            build_ad_flight_segments(df_after, load_platform_settings(), write_to_db=True)
                            if project_val and project_val > 0 and edit_contract_id_val:
                                _compute_and_save_split_amount_for_contract(edit_contract_id_val)
                            if "crud_edit_id" in st.session_state:
                                del st.session_state["crud_edit_id"]
                            if '_table1_cache_key' in st.session_state:
                                del st.session_state['_table1_cache_key']
                            st.success("✅ 已更新")
                            st.rerun()
                        except Exception as e:
                            conn_up.rollback()
                            conn_up.close()
                            st.error(f"更新失敗: {e}")
                with col_cancel:
                    if st.button("取消", key="crud_btn_cancel"):
                        if "crud_edit_id" in st.session_state:
                            del st.session_state["crud_edit_id"]
                        st.rerun()

    # 每列一筆訂單，可直接編輯／刪除（同一行：左側欄位、右側 編輯／刪除）
    if df_orders_crud.empty:
        st.info("📭 尚無訂單資料，請於上方「新增一筆訂單」填寫後儲存。")
    else:
        st.markdown("**每列一筆訂單，可點「編輯」或「刪除」**")
        # 表頭列
        hcols = st.columns([2, 1, 2, 2, 1, 1, 1, 1, 1, 1, 1, 1])
        for i, label in enumerate(["ID", "合約編號", "平台", "客戶", "產品", "起日", "訖日", "秒數", "檔次", "金額", "編輯", "刪除"]):
            with hcols[i]:
                st.markdown(f"**{label}**")
        st.markdown("---")
        for idx, row in df_orders_crud.iterrows():
            cols = st.columns([2, 1, 2, 2, 1, 1, 1, 1, 1, 1, 1, 1])
            _cid = (row.get('contract_id') or '-') if pd.notna(row.get('contract_id')) and row.get('contract_id') else '-'
            with cols[0]:
                st.text(str(row['id'])[:20])
            with cols[1]:
                st.text(str(_cid)[:12])
            with cols[2]:
                st.text(str(row.get('platform', ''))[:14])
            with cols[3]:
                st.text(str(row.get('client', ''))[:14])
            with cols[4]:
                st.text(str(row.get('product', ''))[:10])
            with cols[5]:
                st.text(str(row.get('start_date', ''))[:10])
            with cols[6]:
                st.text(str(row.get('end_date', ''))[:10])
            with cols[7]:
                st.text(str(row.get('seconds', '')))
            with cols[8]:
                st.text(str(row.get('spots', '')))
            with cols[9]:
                st.text(str(int(row.get('amount_net', 0) or 0)))
            with cols[10]:
                if st.button("✏️ 編輯", key=f"edit_{row['id']}"):
                    st.session_state["crud_edit_id"] = row['id']
                    st.rerun()
            with cols[11]:
                if st.button("🗑️ 刪除", key=f"del_{row['id']}", type="primary"):
                    conn_del = get_db_connection()
                    try:
                        conn_del.execute("DELETE FROM orders WHERE id=?", (row['id'],))
                        conn_del.commit()
                        df_after = pd.read_sql("SELECT * FROM orders", conn_del)
                        conn_del.close()
                        build_ad_flight_segments(df_after, load_platform_settings(), write_to_db=True)
                        if "crud_edit_id" in st.session_state and st.session_state.get("crud_edit_id") == row['id']:
                            del st.session_state["crud_edit_id"]
                        if '_table1_cache_key' in st.session_state:
                            del st.session_state['_table1_cache_key']
                        st.success("✅ 已刪除")
                        st.rerun()
                    except Exception as e:
                        conn_del.rollback()
                        conn_del.close()
                        st.error(f"刪除失敗: {e}")
            st.markdown("---")

elif selected_tab == "📅 表2-秒數明細":
    st.markdown("### 📅 表2－秒數明細（對齊 Excel 表2）")
    st.caption("依公司統計總覽、依業務統計明細（平台／合約／客戶／每日使用店秒），含小計。")
    
    df_seg_t2 = _load_segments_cached(_db_mtime)
    if df_seg_t2.empty or df_daily.empty:
        st.warning("📭 尚無檔次段或每日資料，請先產生模擬資料。")
    else:
        # 換月份時不重算：表2 依 _db_mtime 快取於 session_state
        if st.session_state.get('_table2_cache_key') == _db_mtime and '_table2_summary' in st.session_state and '_table2_details' in st.session_state:
            summary_t2 = st.session_state['_table2_summary']
            details_t2 = st.session_state['_table2_details']
            if '_table2_summary_fresh' not in st.session_state or '_table2_summary_qi' not in st.session_state:
                st.session_state['_table2_summary_fresh'] = build_table2_summary_by_company(df_seg_t2, df_daily, df_orders, media_platform='全家新鮮視')
                st.session_state['_table2_summary_qi'] = build_table2_summary_by_company(df_seg_t2, df_daily, df_orders, media_platform='全家廣播(企頻)')
        else:
            summary_t2 = build_table2_summary_by_company(df_seg_t2, df_daily, df_orders)
            details_t2 = build_table2_details_by_company(df_seg_t2, df_daily, df_orders)
            summary_t2_fresh = build_table2_summary_by_company(df_seg_t2, df_daily, df_orders, media_platform='全家新鮮視')
            summary_t2_qi = build_table2_summary_by_company(df_seg_t2, df_daily, df_orders, media_platform='全家廣播(企頻)')
            st.session_state['_table2_summary'] = summary_t2
            st.session_state['_table2_summary_fresh'] = summary_t2_fresh
            st.session_state['_table2_summary_qi'] = summary_t2_qi
            st.session_state['_table2_details'] = details_t2
            st.session_state['_table2_cache_key'] = _db_mtime
        summary_t2_fresh = st.session_state.get('_table2_summary_fresh', pd.DataFrame())
        summary_t2_qi = st.session_state.get('_table2_summary_qi', pd.DataFrame())
        # 區塊一：依公司統計總覽（拆成 新鮮視、企頻 兩表，左欄「公司」固定不捲動）
        def _render_summary_table(summary_df, label):
            if summary_df.empty:
                st.info(f"尚無{label}公司彙總資料")
                return
            col_company = summary_df[['公司']].copy()
            col_rest = summary_df.drop(columns=['公司'])
            tbl_h = min(400, 80 + len(summary_df) * 38)
            c_left, c_right = st.columns([0.5, 7])
            with c_left:
                st.dataframe(_styler_one_decimal(col_company), use_container_width=True, height=tbl_h, hide_index=True)
            with c_right:
                st.dataframe(_styler_one_decimal(col_rest), use_container_width=True, height=tbl_h, hide_index=True)
        st.markdown("#### 依公司統計（新鮮視）")
        _render_summary_table(summary_t2_fresh, "新鮮視")
        st.markdown("#### 依公司統計（企頻）")
        _render_summary_table(summary_t2_qi, "企頻")
        if not summary_t2_fresh.empty or not summary_t2_qi.empty:
            st.caption("(使用店秒) = 每天檔數 × 秒數 × 店數")
        
        # 區塊二／三：依業務統計明細（依公司分組）
        st.markdown("#### 依業務統計明細")
        if details_t2:
            for company_name, detail_df in details_t2.items():
                with st.expander(f"**{company_name}**", expanded=True):
                    st.dataframe(_styler_one_decimal(detail_df), use_container_width=True, height=min(400, 80 + len(detail_df) * 38), hide_index=True)
        else:
            st.info("尚無依公司明細資料")
        
        # 下載：合併為單一 Excel（多工作表）
        st.markdown("#### 📥 下載表2")
        try:
            from io import BytesIO
            buf = BytesIO()
            with pd.ExcelWriter(buf, engine='openpyxl') as w:
                if not summary_t2_fresh.empty:
                    summary_t2_fresh.to_excel(w, sheet_name='依公司統計-新鮮視', index=False)
                if not summary_t2_qi.empty:
                    summary_t2_qi.to_excel(w, sheet_name='依公司統計-企頻', index=False)
                for company_name, detail_df in details_t2.items():
                    sheet_name = str(company_name)[:31]
                    detail_df.to_excel(w, sheet_name=sheet_name, index=False)
            buf.seek(0)
            st.download_button(
                label="📥 下載表2 Excel（含依公司統計與各公司明細）",
                data=buf.getvalue(),
                file_name=f"表2_秒數明細_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        except Exception as e:
            st.caption(f"下載 Excel 時發生錯誤：{e}")

elif selected_tab == "📊 表3-每日庫存":
    _render_tab3(role_readonly=(role == "業務"))

elif selected_tab == "📉 總結表圖表":
    st.markdown("### 📉 總結表視覺化")
    st.caption("圖表與數字表格一併呈現：① 各媒體平台使用率 ② 各秒數類型使用比例；下方為對應的總結表數字。")
    
    summary_year_viz = datetime.now().year
    if not df_daily.empty and '日期' in df_daily.columns:
        df_daily_viz = df_daily.copy()
        df_daily_viz['日期'] = pd.to_datetime(df_daily_viz['日期'], errors='coerce')
        valid = df_daily_viz['日期'].dropna()
        if len(valid) > 0:
            summary_year_viz = int(valid.min().year)
    summary_year_viz = st.number_input("年度", min_value=2020, max_value=2030, value=summary_year_viz, key="summary_year_viz")
    
    if not df_daily.empty and '使用店秒' in df_daily.columns:
        def _monthly_cap_viz(mp, y, m):
            return get_platform_monthly_capacity(mp, y, m)
        annual_viz = build_annual_seconds_summary(df_daily, summary_year_viz, monthly_capacity_loader=_monthly_cap_viz)
        
        if annual_viz:
            month_cols = [f"{m}月" for m in range(1, 13)]
            
            def _style_pct_viz(val):
                if not isinstance(val, (int, float)) or pd.isna(val):
                    return ''
                if val >= 100:
                    return 'background-color: #ff6b6b; color: white'
                if val >= 70:
                    return 'background-color: #ffd93d'
                if val >= 50:
                    return 'background-color: #6bcf7f'
                return ''
            
            # === 區塊 1：媒體平台使用率（圖 + 表）===
            st.markdown("#### ① 各媒體平台使用率隨時間變化趨勢")
            if annual_viz.get('top_usage_df') is not None and not annual_viz['top_usage_df'].empty:
                top_df = annual_viz['top_usage_df'].copy()
                top_df['媒體平台'] = top_df['項目'].str.replace("使用率", "", regex=False)
                chart_df_platform = top_df.set_index("媒體平台")[month_cols].T
                chart_df_platform.index.name = "月份"
                # 使用 Altair 顯示使用率圖表，Y 軸加上 % 符號
                try:
                    import altair as alt
                    chart_df_platform_melted = chart_df_platform.reset_index().melt(id_vars='月份', var_name='媒體平台', value_name='使用率')
                    
                    # 創建帶數據標籤的折線圖
                    # 先創建折線圖
                    line_chart = alt.Chart(chart_df_platform_melted).mark_line(point=True).encode(
                        x=alt.X('月份:O', title='月份'),
                        y=alt.Y('使用率:Q', title='使用率 (%)', axis=alt.Axis(format='.1f')),
                        color=alt.Color('媒體平台:N', title='媒體平台'),
                        tooltip=['月份', '媒體平台', alt.Tooltip('使用率:Q', format='.1f', title='使用率 (%)')]
                    ).properties(width=700, height=400)
                    
                    # 添加數據標籤（在每個點上方顯示數值，格式為 "33.0%"）
                    # 創建格式化後的標籤文字
                    chart_df_platform_melted['使用率標籤'] = chart_df_platform_melted['使用率'].apply(lambda x: f"{x:.1f}%" if pd.notna(x) else "")
                    text_chart = alt.Chart(chart_df_platform_melted).mark_text(
                        align='center',
                        baseline='bottom',
                        dy=-8,  # 標籤位置在點的上方
                        fontSize=10
                    ).encode(
                        x=alt.X('月份:O', title='月份'),
                        y=alt.Y('使用率:Q', title='使用率 (%)', axis=alt.Axis(format='.1f')),
                        text=alt.Text('使用率標籤:N'),  # 使用格式化後的標籤文字
                        color=alt.Color('媒體平台:N', legend=None)  # 標籤顏色與線條一致，但不顯示圖例
                    )
                    
                    # 合併折線圖和標籤
                    chart = (line_chart + text_chart).properties(width=700, height=400)
                    st.altair_chart(chart, use_container_width=True)
                except ImportError:
                    st.line_chart(chart_df_platform)
                st.markdown("**對應數字表：年度使用率（各實體 × 1月~12月）**")
                st.caption("🟢 50%+　🟡 70%+　🔴 100%+；若某媒體整年皆為 0%，請至「媒體秒數與採購」為該媒體設定該年各月每日可用秒數（例如全家廣播(企頻)）。")
                top_tbl = annual_viz['top_usage_df'].copy()
                _month_cols_viz = [c for c in top_tbl.columns if c != '項目']
                def _style_top_table(df_subset):
                    _sub_month_cols = [c for c in _month_cols_viz if c in df_subset.columns]
                    return df_subset.style.format({c: "{:.1f}%" for c in _sub_month_cols}).apply(lambda row: [_style_pct_viz(row.get(c)) for c in df_subset.columns], axis=1)
                _display_monthly_table_split(top_tbl, _month_cols_viz, style_func=_style_top_table, height=180, key_prefix="top_usage")
            else:
                st.info("尚無各媒體平台使用率資料（請於「媒體秒數與採購」分頁為各媒體設定當月每日可用秒數，例如「全家廣播(企頻)」1～3 月，使用率才會顯示）。")
            
            # === 區塊 2：秒數類型使用比例（圖 + 表）===
            st.markdown("#### ② 各秒數類型使用比例隨時間變化趨勢")
            by_type_agg = None
            for ent in ANNUAL_SUMMARY_ENTITY_LABELS:
                block = annual_viz.get('entities', {}).get(ent)
                if not block or block.get('by_type_df') is None:
                    continue
                bt = block['by_type_df'].set_index("項目")[month_cols]
                if by_type_agg is None:
                    by_type_agg = bt.copy()
                else:
                    by_type_agg = by_type_agg + bt
            if by_type_agg is not None and not by_type_agg.empty:
                # 計算百分比數據
                monthly_total = by_type_agg.sum(axis=0)
                proportion = by_type_agg.copy()
                for c in month_cols:
                    if monthly_total.get(c, 0) and monthly_total[c] > 0:
                        proportion[c] = (by_type_agg[c] / monthly_total[c] * 100)
                    else:
                        proportion[c] = 0
                
                # 確保每個月份的比例加總為100%（處理浮點數誤差）
                for col in proportion.columns:
                    monthly_sum = proportion[col].sum()
                    if monthly_sum > 0 and abs(monthly_sum - 100) > 0.01:
                        proportion[col] = proportion[col] / monthly_sum * 100
                
                chart_df_type = proportion.T
                chart_df_type.index.name = "月份"
                
                # 使用 Altair 顯示堆疊長條圖（比例圖表），並添加數據標籤
                try:
                    import altair as alt
                    chart_df_type_melted = chart_df_type.reset_index().melt(id_vars='月份', var_name='秒數類型', value_name='比例')
                    
                    # 確保數值為數值類型，並處理NaN，同時確保所有值都是正數
                    chart_df_type_melted['比例'] = pd.to_numeric(chart_df_type_melted['比例'], errors='coerce').fillna(0)
                    chart_df_type_melted['比例'] = chart_df_type_melted['比例'].clip(lower=0)  # 確保沒有負數
                    
                    # 確保每個月份都有所有秒數類型的數據（如果缺少則補0）
                    all_types = chart_df_type_melted['秒數類型'].unique()
                    all_months = chart_df_type_melted['月份'].unique()
                    complete_data = []
                    for month in all_months:
                        for sec_type in all_types:
                            existing = chart_df_type_melted[(chart_df_type_melted['月份'] == month) & 
                                                           (chart_df_type_melted['秒數類型'] == sec_type)]
                            if existing.empty:
                                complete_data.append({'月份': month, '秒數類型': sec_type, '比例': 0})
                            else:
                                complete_data.append(existing.iloc[0].to_dict())
                    chart_df_type_melted = pd.DataFrame(complete_data)
                    
                    # 確保每個月份的比例加總為100%（再次檢查）
                    chart_df_type_melted['比例'] = chart_df_type_melted.groupby('月份')['比例'].transform(
                        lambda x: (x / x.sum() * 100) if x.sum() > 0 else 0
                    )
                    
                    # 創建格式化後的標籤文字（格式為 "33.0%"），只顯示比例大於 2% 的標籤，避免標籤過小
                    chart_df_type_melted['比例標籤'] = chart_df_type_melted.apply(
                        lambda row: f"{row['比例']:.1f}%" if pd.notna(row['比例']) and row['比例'] > 2 else "", 
                        axis=1
                    )
                    
                    # 計算每個堆疊段的中間位置（用於標籤定位）
                    # 需要按照與長條圖相同的順序排序，然後計算累積位置
                    # 按照秒數類型排序，確保堆疊順序一致
                    chart_df_type_melted_sorted = chart_df_type_melted.sort_values(['月份', '秒數類型']).copy()
                    chart_df_type_melted_sorted = chart_df_type_melted_sorted.reset_index(drop=True)
                    # 計算每個月份內，每個秒數類型之前的累積比例（作為段的起始位置）
                    chart_df_type_melted_sorted['累積起始'] = chart_df_type_melted_sorted.groupby('月份')['比例'].transform(
                        lambda x: x.shift(1).fillna(0).cumsum()
                    )
                    # 段中間位置 = 累積起始位置 + 當前段高度的一半
                    chart_df_type_melted_sorted['段中間位置'] = (
                        chart_df_type_melted_sorted['累積起始'] + chart_df_type_melted_sorted['比例'] / 2
                    )
                    
                    # 創建堆疊長條圖（使用百分比數據直接堆疊）
                    # 因為數據已經是百分比（0-100），使用 stack=True 啟用堆疊
                    bar_chart = alt.Chart(chart_df_type_melted_sorted).mark_bar(size=38).encode(
                        x=alt.X('月份:O', title='月份'),
                        y=alt.Y('比例:Q', title='比例 (%)', 
                               axis=alt.Axis(format='.1f'),
                               stack=True,  # 啟用堆疊功能
                               scale=alt.Scale(domain=[0, 100])),  # 明確設置Y軸範圍為0-100%
                        color=alt.Color('秒數類型:N', title='秒數類型', 
                                      sort=alt.SortField('秒數類型', order='ascending'),  # 確保顏色順序一致
                                      legend=alt.Legend(
                            title='秒數類型',
                            orient='right',
                            titleFontSize=12,
                            labelFontSize=10
                        )),
                        order=alt.Order('秒數類型:O', sort='ascending'),  # 確保堆疊順序
                        tooltip=['月份', '秒數類型', alt.Tooltip('比例:Q', format='.1f', title='比例 (%)')]
                    ).properties(width=700, height=400)
                    
                    # 添加數據標籤（在每個堆疊段的中間位置顯示，只顯示比例 > 2% 的標籤）
                    # 標籤圖表必須使用與長條圖相同的Y軸配置
                    text_chart = alt.Chart(chart_df_type_melted_sorted[chart_df_type_melted_sorted['比例標籤'] != '']).mark_text(
                        align='center',
                        baseline='middle',
                        fontSize=10,
                        fontWeight='bold',
                        fill='white'  # 白色文字更明顯
                    ).encode(
                        x=alt.X('月份:O', title='月份'),
                        y=alt.Y('段中間位置:Q', title='比例 (%)', 
                               axis=alt.Axis(format='.1f'),
                               scale=alt.Scale(domain=[0, 100])),  # Y 軸範圍 0-100，與堆疊圖一致
                        text=alt.Text('比例標籤:N'),  # 使用格式化後的標籤文字
                        color=alt.Color('秒數類型:N', legend=None)  # 標籤不顯示圖例（圖例由 bar_chart 提供）
                    )
                    
                    # 合併長條圖和標籤
                    chart = (bar_chart + text_chart).properties(width=700, height=400)
                    st.altair_chart(chart, use_container_width=True)
                except ImportError:
                    st.bar_chart(chart_df_type)
            else:
                st.info("尚無各秒數類型使用資料。")
            
            # === 總結表數字表格：各實體區塊（秒數用途分列、使用／未使用／使用率）===
            st.markdown("---")
            st.markdown("#### 📊 總結表數字")
            
            # 使用 st.tabs() 讓用戶輕鬆切換不同實體（最簡單且用戶體驗最好）
            entity_tabs = st.tabs([f"📍 {ent}" for ent in ANNUAL_SUMMARY_ENTITY_LABELS])
            
            for idx, ent in enumerate(ANNUAL_SUMMARY_ENTITY_LABELS):
                with entity_tabs[idx]:
                    block = annual_viz.get('entities', {}).get(ent)
                    if not block:
                        st.info(f"尚無 {ent} 的資料")
                        continue
                    st.markdown(f"**{summary_year_viz} {ent}**")
                    st.caption(f"平均每月店秒：{block['avg_monthly_seconds']:,.0f}" if block['avg_monthly_seconds'] else f"{ent} 當月每日可用秒數請於表3 設定。")
                    
                    # === 該實體的圖表（使用率趨勢、使用/未使用堆疊、秒數用途分列）===
                    # 圖1：使用率趨勢（1月~12月）
                    rate_row = block.get('usage_rate_row', {})
                    if rate_row and any(c.endswith('月') for c in rate_row.keys()):
                        rate_data = {c: rate_row.get(c, 0) for c in month_cols if c in rate_row}
                        if rate_data and any(v > 0 for v in rate_data.values()):
                            df_rate = pd.DataFrame([rate_data], index=[f"{ent}使用率"])
                            st.markdown(f"**{ent} 使用率趨勢（1月～12月）**")
                            # 使用 Altair 顯示使用率圖表，Y 軸加上 % 符號
                            try:
                                import altair as alt
                                df_rate_melted = df_rate.T.reset_index()
                                df_rate_melted.columns = ['月份', '使用率']
                                chart = alt.Chart(df_rate_melted).mark_line(point=True).encode(
                                    x=alt.X('月份:O', title='月份'),
                                    y=alt.Y('使用率:Q', title='使用率 (%)', axis=alt.Axis(format='.1f')),
                                    tooltip=['月份', alt.Tooltip('使用率:Q', format='.1f', title='使用率 (%)')]
                                ).properties(width=700, height=300)
                                st.altair_chart(chart, use_container_width=True)
                            except ImportError:
                                st.line_chart(df_rate.T)
                    
                    # 圖2：使用/未使用秒數堆疊（1月~12月）
                    used_row = block.get('used_row', {})
                    unused_row = block.get('unused_row', {})
                    if used_row and unused_row:
                        used_data = {c: used_row.get(c, 0) for c in month_cols if c in used_row}
                        unused_data = {c: unused_row.get(c, 0) for c in month_cols if c in unused_row}
                        if used_data or unused_data:
                            df_usage = pd.DataFrame({
                                '使用秒數': [used_data.get(c, 0) for c in month_cols],
                                '未使用秒數': [unused_data.get(c, 0) for c in month_cols]
                            }, index=month_cols)
                            st.markdown(f"**{ent} 使用/未使用秒數（1月～12月）**")
                            st.area_chart(df_usage)
                    
                    # 圖3：秒數用途分列趨勢（1月~12月）
                    _bt = block['by_type_df']
                    if not _bt.empty and '項目' in _bt.columns:
                        _bt_chart = _bt.set_index('項目')[month_cols].T
                        if not _bt_chart.empty and _bt_chart.sum().sum() > 0:
                            st.markdown(f"**{ent} 秒數用途分列趨勢（1月～12月）**")
                            st.area_chart(_bt_chart)
                    
                    # === 該實體的數字表格 ===
                    _bt_month_cols = [c for c in _bt.columns if c != '項目']
                    def _style_by_type_table(df_subset):
                        _sub_month_cols = [c for c in _bt_month_cols if c in df_subset.columns]
                        return df_subset.style.format({c: "{:,.1f}" for c in _sub_month_cols}) if _sub_month_cols else df_subset.style
                    st.markdown(f"**{ent} 秒數用途分列（1月～12月）**")
                    _display_monthly_table_split(_bt, _bt_month_cols, style_func=_style_by_type_table, height=220, key_prefix=f"by_type_{ent}")
                    
                    summary_table = pd.DataFrame([
                        block['used_row'],
                        block['unused_row'],
                        block['usage_rate_row'],
                    ])
                    _sum_month_cols = [c for c in summary_table.columns if c.endswith('月')]
                    def _style_summary_table(df_subset):
                        _sub_month_cols = [c for c in _sum_month_cols if c in df_subset.columns]
                        # 複製 DataFrame 以便修改顯示值
                        df_display = df_subset.copy()
                        # 保留原始數值用於顏色判斷
                        original_values = {}
                        # 對於使用率行，將數值轉換為帶 % 的字串
                        for idx, row in df_display.iterrows():
                            row_name = str(row.get('項目', ''))
                            original_values[idx] = {}
                            if row_name.endswith('使用率'):
                                for col in _sub_month_cols:
                                    if col in df_display.columns:
                                        val = row[col]
                                        original_values[idx][col] = val  # 保留原始值
                                        if isinstance(val, (int, float)) and not pd.isna(val):
                                            df_display.at[idx, col] = f"{val:.1f}%"
                            else:
                                # 對於其他行（使用秒數、未使用秒數），保持數字格式
                                for col in _sub_month_cols:
                                    if col in df_display.columns:
                                        val = row[col]
                                        original_values[idx][col] = val  # 保留原始值
                                        if isinstance(val, (int, float)) and not pd.isna(val):
                                            df_display.at[idx, col] = f"{val:,.1f}"
                        # 套用顏色樣式（使用原始數值判斷）
                        def _apply_color(row):
                            row_name = str(row.get('項目', ''))
                            idx = row.name
                            colors = []
                            for c in df_subset.columns:
                                if row_name.endswith('使用率') and c.endswith('月'):
                                    # 使用原始數值判斷顏色
                                    orig_val = original_values.get(idx, {}).get(c, row.get(c))
                                    colors.append(_style_pct_viz(orig_val))
                                else:
                                    colors.append('')
                            return colors
                        styled = df_display.style.apply(_apply_color, axis=1)
                        return styled
                    st.markdown(f"**{ent} 使用/未使用/使用率（1月～12月）**")
                    _display_monthly_table_split(summary_table, _sum_month_cols, style_func=_style_summary_table, height=140, key_prefix=f"summary_{ent}")
            
            # PDF 和 Excel 下載按鈕
            st.markdown("---")
            st.markdown("#### 📥 下載報告")
            col_pdf, col_excel = st.columns(2)
            
            with col_pdf:
                pdf_bytes = _build_visualization_summary_pdf(annual_viz, summary_year_viz)
                if pdf_bytes:
                    st.download_button(
                        label="📥 下載 PDF",
                        data=pdf_bytes,
                        file_name=f"總結表視覺化_{summary_year_viz}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf",
                        mime="application/pdf",
                        key="download_viz_pdf",
                        use_container_width=True
                    )
                else:
                    st.caption("PDF 生成失敗（可能缺少中文字型支援）")
            
            with col_excel:
                excel_bytes = _build_visualization_summary_excel(annual_viz, summary_year_viz)
                if excel_bytes:
                    st.download_button(
                        label="📥 下載 Excel",
                        data=excel_bytes,
                        file_name=f"總結表視覺化_{summary_year_viz}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        key="download_viz_excel",
                        use_container_width=True
                    )
                else:
                    st.caption("Excel 生成失敗")
            
            # 下載年度總結（Excel）
            st.markdown("#### 📥 下載年度總結（Excel）")
            try:
                from io import BytesIO
                buf = BytesIO()
                with pd.ExcelWriter(buf, engine='openpyxl') as w:
                    if annual_viz.get('top_usage_df') is not None and not annual_viz['top_usage_df'].empty:
                        annual_viz['top_usage_df'].to_excel(w, sheet_name='年度使用率', index=False)
                    for ent in ANNUAL_SUMMARY_ENTITY_LABELS:
                        block = annual_viz.get('entities', {}).get(ent)
                        if block:
                            block['by_type_df'].to_excel(w, sheet_name=f'{ent}_秒數用途', index=False)
                            pd.DataFrame([block['used_row'], block['unused_row'], block['usage_rate_row']]).to_excel(w, sheet_name=f'{ent}_使用未使用率', index=False)
                buf.seek(0)
                st.download_button(
                    label="📥 下載年度使用秒數總表 Excel",
                    data=buf.getvalue(),
                    file_name=f"年度使用秒數總表_{summary_year_viz}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="dl_annual_summary_viz"
                )
            except Exception as e:
                st.caption(f"下載 Excel 時發生錯誤：{e}")
        else:
            st.warning("📭 尚無每日資料或媒體平台欄位，請先產生模擬資料。")
    else:
        st.warning("📭 尚無每日資料，請先產生模擬資料。")

elif selected_tab == "📊 分公司×媒體 每月秒數":
    st.markdown("### 📊 分公司 × 媒體平台 使用總秒數")
    st.caption("多種圖表回答不同決策問題：結構、總量、誰用最多、是否失衡、趨勢。")
    if df_daily.empty or '使用店秒' not in df_daily.columns or '公司' not in df_daily.columns or '媒體平台' not in df_daily.columns:
        st.warning("📭 尚無每日資料或缺少「公司」「媒體平台」「使用店秒」欄位，請先產生模擬資料。")
    else:
        df_v = df_daily.copy()
        df_v['日期'] = pd.to_datetime(df_v['日期'], errors='coerce')
        df_v = df_v.dropna(subset=['日期'])
        df_v['年'] = df_v['日期'].dt.year
        df_v['月'] = df_v['日期'].dt.month
        agg = df_v.groupby(['年', '月', '公司', '媒體平台'], dropna=False)['使用店秒'].sum().reset_index()
        years_avail = sorted(agg['年'].dropna().unique().astype(int).tolist()) if not agg.empty else [datetime.now().year]
        viz_year = st.number_input("年度", min_value=2020, max_value=2030, value=years_avail[0] if years_avail else datetime.now().year, key="viz_branch_media_year")
        agg_y = agg[agg['年'] == viz_year]
        companies_avail = sorted(agg_y['公司'].dropna().unique().tolist()) if not agg_y.empty else []
        companies_avail = [c for c in companies_avail if c]
        media_avail = sorted(agg_y['媒體平台'].dropna().unique().tolist()) if not agg_y.empty else []
        media_avail = [m for m in media_avail if m]

        time_scope = st.radio("時間範圍", options=["全年合計", "指定月份"], horizontal=True, key="viz_branch_scope")
        if time_scope == "指定月份":
            month_choice = st.selectbox("選擇月份", options=list(range(1, 13)), format_func=lambda x: f"{x}月", key="viz_branch_month")
            agg_scope = agg_y[agg_y['月'] == month_choice]
            scope_label = f"{viz_year} 年 {month_choice} 月"
        else:
            agg_scope = agg_y
            scope_label = f"{viz_year} 年 全年合計"

        if not agg_scope.empty and companies_avail:
            pivot_t = agg_scope.pivot_table(index='公司', columns='媒體平台', values='使用店秒', aggfunc='sum').reindex(companies_avail).fillna(0)
            pivot_t = pivot_t.reindex(columns=media_avail, fill_value=0) if media_avail else pivot_t
            total_scope = pivot_t.sum().sum()
        else:
            pivot_t = pd.DataFrame()
            total_scope = 0
            scope_label = f"{viz_year} 年"

        if not pivot_t.empty and total_scope > 0:
            st.markdown("---")
            st.markdown("#### ① 各分公司 × 媒體平台 — 總秒數堆疊圖")
            st.caption("每根長條為一分公司，各段為各媒體使用總秒數（堆疊為實際秒數，非占比）。")
            try:
                import altair as alt
                cols = pivot_t.reset_index().columns.tolist()
                melt_t = pivot_t.reset_index().melt(id_vars=[cols[0]], var_name="媒體", value_name="秒數").rename(columns={cols[0]: "分公司"})
                chart1 = alt.Chart(melt_t).mark_bar(size=38).encode(
                    x=alt.X("分公司:N", title="分公司"),
                    y=alt.Y("秒數:Q", title="秒數"),
                    color=alt.Color("媒體:N", title="媒體"),
                    tooltip=["分公司", "媒體", alt.Tooltip("秒數:Q", format=",.0f")]
                ).properties(width=700, height=400)
                st.altair_chart(chart1, use_container_width=True)
            except ImportError:
                st.bar_chart(pivot_t)
            st.dataframe(_styler_one_decimal(pivot_t.reset_index()), use_container_width=True, height=min(220, 80 + len(pivot_t) * 36))

            st.markdown("---")
            st.markdown("#### ② 分公司 × 平台 使用秒數（先分公司、再分平台，同平台同色）")
            st.caption("X 軸依序為 分公司-平台（東吳-企頻、東吳-新鮮視…）；同一平台顏色一致，方便比較不同分公司。")
            # 列 = 分公司-平台（先分公司再分平台），欄 = 平台（僅該格有值，其餘 0 → 同平台同色）
            bar_labels = [f"{co}-{mp}" for co in companies_avail for mp in media_avail]
            df_bars = pd.DataFrame(0.0, index=bar_labels, columns=media_avail)
            for co in companies_avail:
                for mp in media_avail:
                    df_bars.loc[f"{co}-{mp}", mp] = float(pivot_t.loc[co, mp]) if co in pivot_t.index and mp in pivot_t.columns else 0.0
            try:
                import altair as alt
                melt_bars = df_bars.reset_index().melt(id_vars=["index"], var_name="媒體", value_name="秒數").rename(columns={"index": "分公司-平台"})
                if not melt_bars.empty:
                    chart2 = alt.Chart(melt_bars).mark_bar(size=38).encode(
                        x=alt.X("分公司-平台:N", title="分公司-平台", sort=bar_labels),
                        y=alt.Y("秒數:Q", title="秒數"),
                        color=alt.Color("媒體:N", title="媒體"),
                        tooltip=["分公司-平台", "媒體", alt.Tooltip("秒數:Q", format=",.0f")]
                    ).properties(width=700, height=400)
                    st.altair_chart(chart2, use_container_width=True)
                else:
                    st.bar_chart(df_bars)
            except ImportError:
                st.bar_chart(df_bars)
            st.dataframe(_styler_one_decimal(pivot_t.reset_index()), use_container_width=True, height=min(220, 80 + len(pivot_t) * 36))

            st.markdown("---")
            st.markdown("#### ③ 某媒體「誰用最多」— 媒體 × 分公司矩陣表 / heatmap")
            st.caption("列＝媒體平台、欄＝分公司；顏色越深表示該媒體在該分公司用量越高（可看出單一媒體誰用最多）。")
            pivot_media_company = pivot_t.T.astype(float)
            # 手動依列做漸層上色（不依賴 matplotlib）
            def _heatmap_row_style(row):
                mn, mx = row.min(), row.max()
                if mx <= mn or pd.isna(mx):
                    return [""] * len(row)
                out = []
                for v in row:
                    if not isinstance(v, (int, float)) or pd.isna(v) or v <= 0:
                        out.append("")
                        continue
                    r = (v - mn) / (mx - mn)
                    # 淺黃 -> 深紅
                    R = 255
                    G = int(255 - 138 * r)
                    B = int(240 - 133 * r)
                    out.append(f"background-color: rgb({R},{max(0,G)},{max(0,B)})")
                return out
            heatmap_styled = pivot_media_company.style.apply(_heatmap_row_style, axis=1).format("{:,.0f}")
            st.dataframe(heatmap_styled, use_container_width=True, height=min(320, 100 + len(pivot_media_company) * 38))

            st.markdown("---")
            st.markdown("#### ④ 資源是否失衡 — 占比 + 警示色")
            st.caption("各分公司內各媒體占比；🔴 單一媒體佔該分公司 ≥50% 可能過度集中、🟡 30–50%、🟢 較分散。")
            row_sum_ = pivot_t.sum(axis=1)
            pct_t = pivot_t.div(row_sum_.replace(0, np.nan), axis=0).fillna(0) * 100

            def _cell_balance_style(v):
                if not isinstance(v, (int, float)) or pd.isna(v):
                    return ""
                if v >= 50:
                    return "background-color: #ff6b6b; color: white"
                if v >= 30:
                    return "background-color: #ffd93d"
                if v > 0:
                    return "background-color: #90EE90"
                return ""

            pct_display = pct_t.reset_index()
            def _balance_color(row):
                return ["" if c == "公司" else _cell_balance_style(row.get(c)) for c in pct_display.columns]
            st.dataframe(pct_display.style.format({c: "{:,.1f}%" for c in media_avail if c in pct_display.columns}).apply(_balance_color, axis=1), use_container_width=True, height=min(280, 80 + len(pct_display) * 36))

            st.markdown("---")
            st.markdown("#### ⑤ 年度 vs 月份趨勢 — 小 multiples 折線圖")
            st.caption("各分公司在 1～12 月、各媒體使用秒數的變化（每區塊一分公司）。")
            if not agg_y.empty and companies_avail and media_avail:
                n_cols = min(3, len(companies_avail))
                for i in range(0, len(companies_avail), n_cols):
                    cols = st.columns(n_cols)
                    for j in range(n_cols):
                        idx = i + j
                        if idx >= len(companies_avail):
                            break
                        co = companies_avail[idx]
                        with cols[j]:
                            agg_co = agg_y[agg_y['公司'] == co].pivot_table(index='月', columns='媒體平台', values='使用店秒', aggfunc='sum').reindex(range(1, 13)).fillna(0)
                            agg_co.index = [f"{int(m)}月" for m in agg_co.index]
                            if not agg_co.empty and agg_co.sum().sum() > 0:
                                st.caption(f"**{co}**")
                                st.line_chart(agg_co)
                            else:
                                st.caption(f"**{co}**（無資料）")

            st.markdown("---")
            st.markdown("#### ⑥ 全年趨勢合併圖（所有分公司-平台 一次看）")
            st.caption("圖⑤ 的折線合併成一張圖，每條線為一個「分公司-平台」；顏色採易辨識配置。")
            if not agg_y.empty and companies_avail and media_avail:
                series_order = [f"{co}-{mp}" for co in companies_avail for mp in media_avail]
                long_rows = []
                for _, r in agg_y.iterrows():
                    key = f"{r['公司']}-{r['媒體平台']}"
                    if key in series_order:
                        long_rows.append({"月": f"{int(r['月'])}月", "分公司-平台": key, "使用秒數": float(r["使用店秒"])})
                if long_rows:
                    df_lines = pd.DataFrame(long_rows)
                    pivot_lines = df_lines.pivot_table(index="月", columns="分公司-平台", values="使用秒數", aggfunc="sum").reindex(
                        [f"{m}月" for m in range(1, 13)], fill_value=0
                    ).fillna(0)
                    for c in series_order:
                        if c not in pivot_lines.columns:
                            pivot_lines[c] = 0
                    pivot_lines = pivot_lines[[c for c in series_order if c in pivot_lines.columns]]
                    if not pivot_lines.empty and pivot_lines.sum().sum() > 0:
                        try:
                            import altair as alt
                            import colorsys
                            # 同一分公司用同一色系（同 hue）、不同平台用深淺區分（不同 lightness）
                            palette = []
                            n_c, n_m = len(companies_avail), len(media_avail)
                            for i in range(n_c):
                                hue = (i / max(1, n_c)) * 0.82
                                for j in range(n_m):
                                    lightness = 0.38 + 0.4 * (j / max(1, n_m))
                                    r, g, b = colorsys.hls_to_rgb(hue, lightness, 0.75)
                                    palette.append("#{:02x}{:02x}{:02x}".format(int(r * 255), int(g * 255), int(b * 255)))
                            source = pivot_lines.reset_index().melt(id_vars=["月"], var_name="分公司-平台", value_name="使用秒數")
                            month_order = [f"{m}月" for m in range(1, 13)]
                            source["月序"] = source["月"].map(lambda x: month_order.index(x) if x in month_order else 0)
                            lines = (
                                alt.Chart(source)
                                .mark_line(strokeWidth=2.5, point=alt.OverlayMarkDef(size=50, filled=True))
                                .encode(
                                    x=alt.X("月:O", title="月份", sort=month_order),
                                    y=alt.Y("使用秒數:Q", title="使用秒數"),
                                    color=alt.Color("分公司-平台:N", legend=alt.Legend(title="分公司-平台"), scale=alt.Scale(range=palette)),
                                    order="月序"
                                )
                                .properties(width=700, height=400)
                            )
                            st.altair_chart(lines, use_container_width=True)
                        except ImportError:
                            st.line_chart(pivot_lines)
                            st.caption("（安裝 altair 可顯示自訂易辨識顏色：pip install altair）")
                    else:
                        st.caption("該年度無使用資料")
                else:
                    st.caption("該年度無使用資料")
            else:
                st.caption("尚無分公司或媒體資料。")

        else:
            st.caption("尚無分公司或媒體資料，或該時間範圍無使用資料，請先產生模擬資料。")

elif selected_tab == "📋 媒體秒數與採購":
    st.markdown("### 📋 媒體秒數與採購")
    st.caption("輸入各媒體平台「一年 12 個月」的購買秒數與購買價格；儲存後會同步更新表3 的當月每日可用秒數，並供 ROI 分頁計算成本。")
    purchase_year = st.number_input("年度", min_value=2020, max_value=2030, value=datetime.now().year, key="purchase_year")
    if st.button("🎲 產生模擬採購資料（測試用）", type="secondary", key="gen_mock_purchase", help="為上述年度、所有媒體產生 1～12 月合理模擬數據，方便測試表3 與 ROI 分頁"):
        with st.spinner("正在產生模擬採購資料..."):
            ok, msg = generate_mock_platform_purchase_for_year(purchase_year)
            if ok:
                # 清除所有相關的 session_state，確保輸入框會重新載入資料庫的值
                to_del = [k for k in st.session_state if str(k).startswith("purchase_sec_") or str(k).startswith("purchase_price_")]
                for k in to_del:
                    del st.session_state[k]
                st.success(msg)
                time.sleep(0.3)  # 短暫延遲確保資料庫寫入完成
                st.rerun()
            else:
                st.error(f"產生失敗：{msg}")
    # 每次頁面載入時都重新從資料庫讀取最新資料
    existing = load_platform_monthly_purchase_all_media_for_year(purchase_year)
    import calendar
    for mp in MEDIA_PLATFORM_OPTIONS:
        st.markdown(f"#### {mp}")
        data = existing.get(mp, {})
        cols = st.columns(12)
        inputs_sec = {}
        inputs_price = {}
        for m in range(1, 13):
            with cols[m - 1]:
                st.markdown(f"**{m}月**")
                sec, pr = data.get(m, (0, 0.0))
                key_sec = f"purchase_sec_{mp}_{m}"
                key_price = f"purchase_price_{mp}_{m}"
                # 如果 session_state 中沒有該 key，使用資料庫的值；否則使用 session_state 的值（保留用戶輸入）
                default_sec = int(sec) if sec else 0
                default_price = float(pr) if pr else 0.0
                # 當 session_state 中沒有值時，使用資料庫的值作為預設值
                # 這樣當產生模擬資料後清除 session_state，輸入框會自動顯示新的資料庫值
                inputs_sec[m] = st.number_input(
                    "購買秒數",
                    min_value=0,
                    value=default_sec,
                    step=5000,
                    key=key_sec,
                )
                inputs_price[m] = st.number_input(
                    "購買價格（元）",
                    min_value=0.0,
                    value=default_price,
                    step=1000.0,
                    format="%.0f",
                    key=key_price,
                )
        if st.button(f"儲存 {mp}", key=f"save_purchase_{mp}"):
            for m in range(1, 13):
                set_platform_monthly_purchase(mp, purchase_year, m, inputs_sec[m], inputs_price[m])
            st.success(f"已儲存 {mp} {purchase_year} 年 1~12 月資料（並已同步表3 每日可用秒數）。")
            st.rerun()
    st.markdown("---")
    st.caption("儲存後，ROI 分頁將依「購買價格 ÷ 購買秒數」計算成本並產生投報率。")

elif selected_tab == "🧪 Ragic抓取測試":
    from ui_ragic_test import render_ragic_test_tab

    render_ragic_test_tab(ragic_fields=RAGIC_FIELDS, parse_cue_excel_for_table1=parse_cue_excel_for_table1)

elif selected_tab == "🧪 實驗分頁":
    # 鎖定分頁：切換「分析對象」等控件會觸發 rerun，避免跳到其他分頁
    st.session_state["main_tab"] = "🧪 實驗分頁"
    st.markdown("### 🧪 依時間的庫存警示與分析（實驗）")
    with st.expander("📌 系統前提（核心假設）", expanded=True):
        st.markdown("""
- **當月秒數若未使用，於月底結算時視為 100% 浪費（不可逆）**
- 秒數的價值會隨時間接近月底而**快速衰減**
- 系統目標：**最小化月底浪費**（不是避免爆量）
- 爆量仍需監控，但屬次要風險
        """)
    today = datetime.now().date()
    exp_year = st.number_input("年度", min_value=2020, max_value=2030, value=today.year, key="exp_year")
    exp_month = st.number_input("月份", min_value=1, max_value=12, value=today.month, key="exp_month")
    emergency_days = st.slider("緊急期天數（T0 可補救窗口）", min_value=3, max_value=14, value=EMERGENCY_DAYS, key="exp_emergency_days")

    # 分析對象：全媒體合計 或 單一媒體平台
    exp_scope_options = ["全媒體合計"] + list(MEDIA_PLATFORM_OPTIONS)
    exp_scope = st.selectbox(
        "**分析對象**（本頁所有指標與圖表皆依此對象計算）",
        exp_scope_options,
        key="exp_scope"
    )
    exp_media_filter = None if exp_scope == "全媒體合計" else exp_scope

    if not df_daily.empty and '使用店秒' in df_daily.columns:
        def _cap_loader(mp, y, mo):
            return get_platform_monthly_capacity(mp, y, mo)
        daily_inv, metrics = build_daily_inventory_and_metrics(
            df_daily, exp_year, exp_month, today,
            emergency_days=emergency_days,
            monthly_capacity_loader=_cap_loader,
            media_platform=exp_media_filter,
        )
        month_cap = metrics["month_total_capacity"] or 1
        month_used = metrics["month_total_used"]
        past_wasted_pct = round(metrics["past_wasted_seconds"] / month_cap * 100, 1) if month_cap else 0

        # 明確標示目前顯示的對象
        st.markdown("---")
        st.info(f"📌 **目前顯示對象：{exp_scope}** — 以下浪費總覽、時間軸、救援壓力與戰略判斷皆為此對象。")
        # === 區塊 A：本月浪費總覽（Hero KPIs）===
        st.markdown("#### 🔝 本月浪費總覽")
        c1, c2, c3 = st.columns(3)
        with c1:
            st.metric("本月已成浪費率（過去日）", f"{past_wasted_pct}%")
        with c2:
            st.metric("已浪費秒數（Past）", _seconds_to_spot_label(metrics["past_wasted_seconds"], short=True))
        with c3:
            st.metric("尚可救援秒數（Emergency）", _seconds_to_spot_label(metrics["emergency_unused_seconds"], short=True))
        st.caption(f"尚可救援 = **{_seconds_to_spot_label(metrics['emergency_unused_seconds'], short=True)}** 未賣（全省 15 秒檔）")

        # === 區塊 B：時間價值條（視覺化）===
        st.markdown("#### 🧠 時間軸與未售庫存視覺化")
        past_sec = daily_inv[daily_inv["time_bucket"] == "past"]["unused_seconds"].sum()
        em_sec = daily_inv[daily_inv["time_bucket"] == "emergency"]["unused_seconds"].sum()
        buf_sec = daily_inv[daily_inv["time_bucket"] == "buffer"]["unused_seconds"].sum()
        total_unused = past_sec + em_sec + buf_sec or 1
        p_past = past_sec / total_unused * 100
        p_em = em_sec / total_unused * 100
        p_buf = buf_sec / total_unused * 100

        # B1：本月「時間軸」橫條（依「月/日」呈現 Past / Emergency / Buffer，方便閱讀）
        ndays = len(daily_inv)
        timeline_rows = []
        date_order_list = []
        for _, row in daily_inv.iterrows():
            d = row["date"]
            date_label = f"{d.month}/{d.day}"
            date_order_list.append(date_label)
            bucket = row["time_bucket"]
            label = "過去（浪費）" if bucket == "past" else ("緊急期（可救援）" if bucket == "emergency" else "緩衝期")
            timeline_rows.append({"日期": date_label, "bucket": bucket, "label": label})
        df_timeline = pd.DataFrame(timeline_rows)
        try:
            import altair as alt
            today_label = f"{today.month}/{today.day}" if (today.year == exp_year and today.month == exp_month) else None
            domain_bucket = ["past", "emergency", "buffer"]
            range_bucket = ["#c0392b", "#e67e22", "#27ae60"]
            chart_timeline = alt.Chart(df_timeline).mark_rect().encode(
                x=alt.X("日期:N", title="日期（月/日）", sort=date_order_list, axis=alt.Axis(labelFontSize=9, titleFontSize=10)),
                y=alt.value(0),
                y2=alt.value(60),
                color=alt.Color("bucket:N", title="區段", scale=alt.Scale(domain=domain_bucket, range=range_bucket), legend=alt.Legend(title="時間區段", labelFontSize=9, titleFontSize=10)),
                tooltip=[alt.Tooltip("日期:N", title="日期"), alt.Tooltip("label:N", title="說明")]
            ).properties(height=80, width=700, title=f"本月時間軸（{exp_month}月・左=月初 → 右=月底）")
            if today_label and today_label in date_order_list:
                rule_today = alt.Chart(pd.DataFrame([{"日期": today_label}])).mark_rule(color="white", strokeWidth=3).encode(x="日期:N")
                chart_timeline = alt.layer(chart_timeline, rule_today)
            st.altair_chart(chart_timeline, use_container_width=True)
            st.caption("🔴 今日為白線｜紅=已過（浪費）｜橙=緊急期（可救援）｜綠=緩衝期")
        except Exception:
            st.markdown(f"[🟥 Past 未售 {int(past_sec):,} 秒 ] [ 🟧 Emergency {int(em_sec):,} 秒 ] [ 🟩 Buffer {int(buf_sec):,} 秒 ]")
            st.caption("Past：已成浪費｜Emergency：可救援｜Buffer：可等待調度")

        # B2：未售庫存「結構條」（依未售秒數比例，感受量與緊迫度）
        try:
            import altair as alt
            df_bar = pd.DataFrame([
                {"label_short": "過去浪費", "segment": "過去浪費（不可逆）", "秒數": int(past_sec), "pct": p_past, "order": 1},
                {"label_short": "緊急期可救援", "segment": "緊急期未售（可救援）", "秒數": int(em_sec), "pct": p_em, "order": 2},
                {"label_short": "緩衝期未售", "segment": "緩衝期未售", "秒數": int(buf_sec), "pct": p_buf, "order": 3},
            ])
            df_bar = df_bar[df_bar["秒數"] > 0]
            if not df_bar.empty:
                label_order = ["過去浪費", "緊急期可救援", "緩衝期未售"]
                seg_range = ["#c0392b", "#e67e22", "#27ae60"]
                chart_bar = alt.Chart(df_bar).mark_bar(size=36).encode(
                    x=alt.X("pct:Q", title="佔未售比例（%）", scale=alt.Scale(domain=[0, 100])),
                    y=alt.Y("label_short:N", title="", sort=label_order, axis=alt.Axis(labelLimit=0, labelPadding=10)),
                    color=alt.Color("label_short:N", scale=alt.Scale(domain=label_order, range=seg_range), legend=None),
                    tooltip=[alt.Tooltip("segment:N", title="區段"), alt.Tooltip("秒數:Q", title="未售店秒", format=","), alt.Tooltip("pct:Q", title="比例%", format=".1f")]
                ).properties(height=180, title="未售庫存結構（可救援比例愈高愈需行動）").configure_axis(
                    labelFontSize=12, labelLimit=0
                )
                st.altair_chart(chart_bar, use_container_width=True)
        except Exception:
            pass

        # === 區塊 C：救援壓力面板（視覺化緊迫感）===
        st.markdown("#### ⏱ 救援壓力與緊迫程度")
        rem = metrics["remaining_days"]
        req = metrics["required_daily_seconds"]
        emergency_total_days = min(emergency_days, len([d for d in daily_inv["date"] if d >= today and d <= today + timedelta(days=emergency_days)])) or emergency_days
        rem_ratio = rem / emergency_total_days if emergency_total_days else 0
        daily_cap = (month_cap / ndays) if ndays else 1
        req_vs_cap = min(1.0, (req / daily_cap)) if daily_cap else 0

        c1, c2, c3 = st.columns(3)
        with c1:
            st.metric("剩餘緊急期天數", f"{rem} 天")
            st.progress(rem_ratio, text="緊急期剩餘時間" if rem_ratio <= 0.5 else "尚有緩衝")
        with c2:
            st.metric("尚未售出（緊急期內）", _seconds_to_spot_label(metrics["emergency_unused_seconds"], short=True))
            emergency_unused_ratio = metrics["emergency_unused_seconds"] / month_cap if month_cap else 0
            st.progress(min(1.0, emergency_unused_ratio), text="緊急期未售佔本月容量")
        with c3:
            st.metric("每日需賣", _seconds_to_spot_label(req, short=True))
            st.progress(req_vs_cap, text="每日需賣 vs 日容量（愈滿愈吃緊）")

        st.markdown("**每日需賣：** ≈ **{:,}** 店秒/日（{}）".format(int(req), _seconds_to_spot_label(req, short=True)))

        # === 區塊 C2：剩餘日子可賣多少（視覺化 + 與建議呼應的數字）===
        emergency_unused_sec = metrics["emergency_unused_seconds"]
        total_sellable_label = _seconds_to_spot_label(emergency_unused_sec, short=True)
        daily_target_label = _seconds_to_spot_label(req, short=True) if rem else "—"
        st.markdown("#### 📅 剩餘日子可賣多少")
        if rem > 0 and emergency_unused_sec > 0:
            try:
                import altair as alt
                emergency_dates = metrics.get("emergency_dates") or []
                target_per_day = req
                rows = []
                for d in sorted(emergency_dates):
                    date_label = f"{d.month}/{d.day}" if hasattr(d, "month") else str(d)
                    rows.append({"日期": date_label, "date_sort": d, "需賣": int(target_per_day), "需賣檔": round(target_per_day / SECONDS_PER_SPOT_15S, 1)})
                if not rows:
                    date_order = [today + timedelta(days=i) for i in range(rem)]
                    rows = [{"日期": f"{d.month}/{d.day}", "date_sort": d, "需賣": int(target_per_day), "需賣檔": round(target_per_day / SECONDS_PER_SPOT_15S, 1)} for d in date_order]
                df_days = pd.DataFrame(rows)
                date_order_str = df_days["日期"].tolist()
                chart_days = alt.Chart(df_days).mark_bar(color="#e67e22").encode(
                    x=alt.X("日期:N", title="日期", sort=date_order_str),
                    y=alt.Y("需賣:Q", title="店秒"),
                    tooltip=[
                        alt.Tooltip("日期:N", title="日期"),
                        alt.Tooltip("需賣:Q", title="當日建議需賣(店秒)", format=","),
                        alt.Tooltip("需賣檔:Q", title="約檔(15秒)", format=".1f")
                    ]
                ).properties(height=220, title=f"依日期・每日建議需賣 ≈ {daily_target_label}")
                st.altair_chart(chart_days, use_container_width=True)
            except Exception:
                pass
            st.markdown(f"**未來 {rem} 天**內可售總量 = **{total_sellable_label}**｜每日目標 ≈ **{daily_target_label}**（與下方行動建議一致）")
        else:
            st.caption("緊急期內無剩餘天數或無未售量，無需補救目標。")

        # === 區塊 D：即時戰略判斷（視覺化狀態與風險）===
        st.markdown("#### 🚦 即時戰略判斷")
        state = metrics["strategy_state"]
        state_label = {"SELL": "強推補檔", "HOLD": "限制接案", "NORMAL": "正常銷售", "ANOMALY": "檢查假設"}[state]
        state_color = {"SELL": "#e74c3c", "HOLD": "#f39c12", "NORMAL": "#27ae60", "ANOMALY": "#9b59b6"}
        state_bg = state_color.get(state, "#95a5a6")
        st.markdown(
            f'<div style="background:{state_bg};color:white;padding:12px 20px;border-radius:8px;font-size:1.1em;margin:8px 0;">'
            f'🎯 當前戰略：<strong>{state}</strong> — {state_label}'
            f'</div>',
            unsafe_allow_html=True
        )
        under_high = metrics["under_risk"] >= 0.5
        over_high = metrics["over_risk"] >= 0.5
        time_pressure_high = rem <= 3 and metrics["emergency_unused_seconds"] > 0
        risk_waste = min(1.0, metrics["under_risk"])
        risk_over = min(1.0, metrics["over_risk"])
        risk_time = 1.0 if time_pressure_high else (0.5 if rem <= 5 and metrics["emergency_unused_seconds"] > 0 else 0.0)
        r1, r2, r3 = st.columns(3)
        with r1:
            st.caption("浪費風險（未達目標使用率）")
            st.progress(risk_waste)
            st.caption("高" if under_high else "低")
        with r2:
            st.caption("爆量風險（超過安全上限）")
            st.progress(risk_over)
            st.caption("高" if over_high else "低")
        with r3:
            st.caption("時間壓力（緊急期內未售）")
            st.progress(risk_time)
            st.caption("高" if time_pressure_high else "中/低")

        # === 區塊 E：行動建議 ===
        st.markdown("#### 📌 行動建議")
        if state == "SELL":
            suggestions = ["15 秒短檔（區域）優先推", "舊客戶補檔／加購", "包量促銷或限時方案"]
        elif state == "HOLD":
            suggestions = ["暫緩新案接單", "以既有訂單消化為主", "觀察明日使用率再決定"]
        elif state == "ANOMALY":
            suggestions = ["檢查資料與假設是否正確", "確認容量設定與實際排程", "必要時人工覆核"]
        else:
            suggestions = ["維持正常銷售節奏", "留意緊急期內未售秒數", "可排日可彈性接案"]
        for i, s in enumerate(suggestions, 1):
            st.markdown(f"{i}. {s}")

        # === 系統語句生成 ===
        if metrics["emergency_unused_seconds"] > 0 and rem > 0 and rem <= 10:
            st.info(f"💬 **本月進入關鍵救援期，未來 {rem} 天為唯一補救窗口。**")
        st.caption("TWWI（時間加權浪費指數）= " + str(round(metrics["twwi"], 1)))

        # 可選：日粒度表
        with st.expander("📋 日粒度事實表（daily_inventory）", expanded=False):
            st.dataframe(_styler_one_decimal(daily_inv), use_container_width=True, height=400)
    else:
        st.warning("📭 尚無每日資料，請先產生模擬資料。")

elif selected_tab == "📊 ROI":
    st.markdown("### 📊 ROI 投報分析")
    st.caption("依現有採購與訂單資料計算各媒體之投報率，支援多時間維度檢視。")

    with st.expander("📖 資料來源說明", expanded=False):
        st.markdown("""
| 項目 | 來源 |
|------|------|
| **購買成本** | 「📋 媒體秒數與採購」分頁的購買價格，依選定時間維度彙總 |
| **實收金額** | 表1 訂單；依檔次段日期與選定區間重疊者計算，同一合約多媒體時依秒數比例或拆分金額分配 |
| **ROI** | (實收 - 購買成本) ÷ 購買成本 |
""")

    # 時間維度選擇
    roi_time_dim = st.radio(
        "時間維度",
        options=["month", "quarter", "year", "all"],
        format_func=lambda x: {"month": "📅 單月", "quarter": "📊 單季", "year": "📆 單年", "all": "🔄 累計至今"}[x],
        horizontal=True,
        key="roi_time_dim",
    )
    roi_year = datetime.now().year
    roi_month = 1
    roi_quarter = 1
    if roi_time_dim == "month":
        c1, c2 = st.columns(2)
        with c1:
            roi_year = st.number_input("參考年度", min_value=2020, max_value=2030, value=datetime.now().year, key="roi_year")
        with c2:
            roi_month = st.number_input("參考月份", min_value=1, max_value=12, value=datetime.now().month, key="roi_month")
        period_label = f"{roi_year}年{roi_month}月"
    elif roi_time_dim == "quarter":
        c1, c2 = st.columns(2)
        with c1:
            roi_year = st.number_input("參考年度", min_value=2020, max_value=2030, value=datetime.now().year, key="roi_year")
        with c2:
            roi_quarter = st.selectbox("參考季度", options=[1, 2, 3, 4], format_func=lambda x: f"Q{x}（{'1-3月' if x==1 else '4-6月' if x==2 else '7-9月' if x==3 else '10-12月'}）", key="roi_quarter")
        roi_month = (roi_quarter - 1) * 3 + 1
        period_label = f"{roi_year} Q{roi_quarter}"
    elif roi_time_dim == "year":
        roi_year = st.number_input("參考年度", min_value=2020, max_value=2030, value=datetime.now().year, key="roi_year")
        period_label = f"{roi_year}年"
    else:
        st.caption("將彙總所有採購與訂單資料，無需選擇年度或月份。")
        period_label = "累計至今"

    roi_rows = _calculate_roi_by_period(roi_time_dim, roi_year, roi_month if roi_time_dim in ("month", "quarter") else 1, period_label)

    if not roi_rows:
        st.warning("尚無採購資料或該區間無資料。請至「📋 媒體秒數與採購」分頁輸入購買秒數與購買價格。")
    else:
        roi_df = pd.DataFrame(roi_rows)
        display_label = period_label
        if period_label == "累計至今":
            range_start, range_end = _get_roi_all_period_date_range()
            if range_start and range_end:
                display_label = f"累計至今（{range_start} ～ {range_end}）"
        st.markdown(f"#### 媒體別 ROI 表 — {display_label}")

        # ROI 長條圖（依 ROI 正負上色）
        try:
            import altair as alt
            roi_chart_df = roi_df.copy()
            roi_chart_df["ROI色彩"] = roi_chart_df["ROI（投報率）"].apply(lambda x: "正報酬" if x >= 0 else "負報酬")
            chart_roi = alt.Chart(roi_chart_df).mark_bar(size=38).encode(
                x=alt.X("媒體:N", title="媒體", sort="-y"),
                y=alt.Y("ROI（投報率）:Q", title="ROI（投報率）", axis=alt.Axis(format="%")),
                color=alt.Color("ROI色彩:N", scale=alt.Scale(domain=["正報酬", "負報酬"], range=["#27ae60", "#e74c3c"]), legend=None),
                tooltip=[
                    alt.Tooltip("媒體:N", title="媒體"),
                    alt.Tooltip("ROI（投報率）:Q", format=".2%", title="ROI"),
                    alt.Tooltip("實收金額（元）:Q", format=",.0f"),
                    alt.Tooltip("購買成本（元）:Q", format=",.0f"),
                ]
            ).properties(width=700, height=350).configure_axisY(format="%")
            st.altair_chart(chart_roi, use_container_width=True)
        except Exception:
            st.bar_chart(roi_df.set_index("媒體")["ROI（投報率）"])

        st.dataframe(_styler_one_decimal(roi_df.drop(columns=["時間區間"], errors="ignore")), use_container_width=True, height=min(200, 60 + len(roi_rows) * 38))

        # 多維度一鍵比較（設計巧思）
        st.markdown("---")
        st.markdown("#### 🔀 多維度比較")
        st.caption("一次檢視「當月、當季、當年、累計」四種維度的 ROI，快速掌握各媒體在不同時間尺度下的表現。")
        if st.checkbox("顯示多維度比較表", value=False, key="roi_multi_compare"):
            all_rows = []
            for pt, pl in [("month", f"{roi_year}年{roi_month}月"), ("quarter", f"{roi_year} Q{(roi_month-1)//3+1}"), ("year", f"{roi_year}年"), ("all", "累計至今")]:
                r = _calculate_roi_by_period(pt, roi_year, roi_month if pt in ("month", "quarter") else 1, pl)
                for row in r:
                    row["時間區間"] = pl
                    all_rows.append(row)
            if all_rows:
                multi_df = pd.DataFrame(all_rows)
                # 樞紐：列=媒體，欄=時間區間，值=ROI
                pivot_roi = multi_df.pivot_table(index="媒體", columns="時間區間", values="ROI（投報率）", aggfunc="first")
                order_cols = [f"{roi_year}年{roi_month}月", f"{roi_year} Q{(roi_month-1)//3+1}", f"{roi_year}年", "累計至今"]
                pivot_roi = pivot_roi.reindex(columns=[c for c in order_cols if c in pivot_roi.columns])
                st.dataframe(pivot_roi.style.format("{:.2%}"), use_container_width=True)

# （檔次稽核、檔次拆解表 已移除）
